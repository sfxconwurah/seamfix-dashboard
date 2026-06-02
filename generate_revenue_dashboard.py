#!/usr/bin/env python3
"""
Seamfix Revenue & Fundability Dashboard Generator
Analyzes revenue streams (in USD), projections vs actuals, and funding gap vs budget (in Naira).
All revenue in USD, budget in Naira. Exchange rate: $1 = ₦1,450
Usage: python3 generate_revenue_dashboard.py [folder_path]
"""

import os, sys, json, re, glob
from datetime import datetime
from openpyxl import load_workbook
from theme import get_base_css, get_toggle_html, get_theme_js


# Currency constants
FX_RATE = 1450  # $1 USD = ₦1,450 NGN

# Monthly actual columns: M=Jan, N=Feb, ..., X=Dec (convention set by Finance)
MONTH_COLUMNS = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def sf(val):
    """Safe float conversion"""
    if val is None:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0


def fmt_usd(val):
    """Format USD with appropriate suffix"""
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000:
        return f"{sign}${v/1_000_000:.2f}M"
    elif v >= 1_000:
        return f"{sign}${v/1_000:.1f}K"
    else:
        return f"{sign}${v:,.0f}"


def fmt_naira(val):
    """Format Naira with appropriate suffix"""
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000_000:
        return f"{sign}₦{v/1_000_000_000:.2f}B"
    elif v >= 1_000_000:
        return f"{sign}₦{v/1_000_000:.1f}M"
    elif v >= 1_000:
        return f"{sign}₦{v/1_000:.0f}K"
    else:
        return f"{sign}₦{v:,.0f}"


def fmt_dual(val_usd):
    """Format as USD primary with Naira equivalent in parentheses"""
    return f"{fmt_usd(val_usd)} ({fmt_naira(val_usd * FX_RATE)})"


def safe_row_dict(row):
    """Convert openpyxl row to dict by column letter"""
    d = {}
    for c in row:
        try:
            d[c.column_letter] = c.value
        except (AttributeError, TypeError):
            pass
    return d


def find_file(folder, pattern):
    """Find first file matching pattern (case-insensitive)"""
    files = glob.glob(os.path.join(folder, f"*{pattern}*"), recursive=False)
    if files:
        return files[0]
    for f in os.listdir(folder):
        if pattern.lower() in f.lower() and f.endswith('.xlsx'):
            return os.path.join(folder, f)
    return None


def extract_revenue_data(revenue_file):
    """Extract revenue projections and actuals from Revenues sheet (ALL IN USD)"""
    wb = load_workbook(revenue_file, data_only=True)
    ws = wb['Revenues']

    revenues = []
    current_parent = None  # Track parent deal for sub-items

    # Section headers to skip (but not TOTAL — hitting TOTAL means we stop scanning)
    SECTION_HEADERS = {"ANCHOR DEALS", "EXISTING CUSTOMERS", "DEALS FROM 2025", "NEW BUSINESS"}

    # Scan up to row 80 — sheet previously ended at row 62 but grows as new deals are added.
    # We BREAK on the TOTAL row to avoid picking up secondary tables (e.g. monthly budget
    # breakdown) that sit below the main deals table in the same sheet.
    for row_idx in range(3, 80):
        row = ws[row_idx]
        cells = safe_row_dict(row)

        # Skip if E is None and this looks like a subtotal (Column A is empty)
        sn = sf(cells.get('A'))
        name = cells.get('B')

        # Stop at the TOTAL row — everything below is a different table
        if name and str(name).strip().upper() == 'TOTAL':
            break

        # Skip section headers
        if name and str(name).strip() in SECTION_HEADERS:
            current_parent = None
            continue

        # Skip empty rows
        if not name or not str(name).strip():
            continue

        name = str(name).strip()

        # Check if this is a main line (has S/N) or sub-item
        has_sn = sn > 0

        # If it has S/N, it's a parent item
        if has_sn:
            current_parent = name

        # Get rail/category
        rail = cells.get('C')
        if rail:
            rail = str(rail).strip()
        elif current_parent:
            # Sub-items inherit rail from context or stay generic
            rail = 'Others'
        else:
            rail = 'Others'

        # Get annual revenue (USD)
        annual_usd = sf(cells.get('E'))

        # Get actuals to check if row has any data
        has_actuals = any(sf(cells.get(col)) != 0 for col in ['M', 'N', 'O', 'P'])

        # Include rows with annual target OR actual revenue data
        if annual_usd == 0 and not has_actuals:
            continue

        # Get start date
        start_date = cells.get('D')
        if start_date:
            if hasattr(start_date, 'strftime'):
                start_date_str = start_date.strftime('%d %b %Y')
            else:
                start_date_str = str(start_date)
        else:
            start_date_str = 'N/A'

        # Get status
        status = cells.get('K')
        status = str(status).strip() if status else 'Unknown'

        # Get monthly actuals (USD) — dynamically read all 12 months (M=Jan..X=Dec)
        monthly = [sf(cells.get(col)) for col in MONTH_COLUMNS]

        # Deficit=Y, Surplus=Z (Finance added Apr–Dec columns, shifting these right)
        deficit = sf(cells.get('Y'))
        surplus = sf(cells.get('Z'))

        ytd_actual = sum(monthly)  # will be refined after we detect which months have data globally

        # achievement_pct, gap, ytd_target_pace calculated in post-processing below

        revenues.append({
            'sn': sn,
            'name': name,
            'rail': rail,
            'start_date': start_date_str,
            'start_date_obj': start_date if hasattr(start_date, 'month') else None,
            'annual_usd': annual_usd,
            'status': status,
            'monthly': monthly,
            'ytd_actual': ytd_actual,
            'deficit': deficit,
            'surplus': surplus,
        })

    wb.close()

    # ── POST-PROCESSING: detect months with data and calculate YTD/pace dynamically ──
    if revenues:
        monthly_totals = [sum(r['monthly'][i] for r in revenues) for i in range(12)]
        months_with_data = [i for i, v in enumerate(monthly_totals) if v > 0]
        last_data_month = max(months_with_data) if months_with_data else 0
        num_months = last_data_month + 1

        for r in revenues:
            # Recalculate YTD based on actual months with data
            r['ytd_actual'] = sum(r['monthly'][:num_months])

            # Calculate months_active based on start date
            months_active = num_months
            sd = r.get('start_date_obj')
            if sd and hasattr(sd, 'month'):
                if sd.year == 2026 and sd.month > num_months:
                    months_active = 0
                elif sd.year == 2026 and sd.month >= 1:
                    months_active = max(0, num_months + 1 - sd.month)
            elif r['start_date'] in ('Closed', 'Closed - Dec 31', 'Closed - Aug 31'):
                months_active = num_months

            r['months_active'] = months_active
            r['ytd_target_pace'] = r['annual_usd'] * months_active / 12
            # Achievement % = progress toward FULL annual target (not pace-adjusted)
            # Finance expects: earned full annual amount = 100%, not 240%
            r['achievement_pct'] = (r['ytd_actual'] / r['annual_usd'] * 100) if r['annual_usd'] > 0 else 0
            r['gap'] = r['annual_usd'] - r['ytd_actual']

    return revenues


def extract_budget_data(budget_file):
    """Extract total annual budget from Budget Summary sheet (IN NAIRA)"""
    wb = load_workbook(budget_file, data_only=True)
    ws = wb['Budget Summary']

    total_budget_ngn = 0
    for row_idx in range(1, ws.max_row + 1):
        row = ws[row_idx]
        cells = safe_row_dict(row)

        a_val = cells.get('A')
        if a_val and 'TOTAL EXPENSE BUDGETS' in str(a_val):
            total_budget_ngn = sf(cells.get('B'))
            break

    wb.close()
    return total_budget_ngn


def generate_critical_actions(revenues, budget_ngn, annual_revenue_usd):
    """Auto-generate specific, actionable recommendations"""
    actions = []

    # Convert budget to USD for comparison
    budget_usd = budget_ngn / FX_RATE

    # 1. Which deals MUST close for budget to be funded?
    fundability_ratio = annual_revenue_usd / budget_usd if budget_usd > 0 else 0
    funding_gap_usd = max(0, budget_usd - annual_revenue_usd)

    if fundability_ratio < 1.0:
        must_close = []
        for r in sorted(revenues, key=lambda x: -x['annual_usd']):
            if r['status'] not in ('On Track', 'Completed'):
                must_close.append(r)
                if len(must_close) >= 3:
                    break

        if must_close:
            deals_str = ', '.join([f"{m['name']} ({fmt_usd(m['annual_usd'])})" for m in must_close])
            actions.append({
                'type': 'CRITICAL',
                'title': 'Pipeline Closure Required',
                'description': f"To fund the {fmt_naira(budget_ngn)} budget, need {fmt_usd(funding_gap_usd)} ({fmt_naira(funding_gap_usd * FX_RATE)}) more revenue. Priority deals to close: {deals_str}"
            })

    # 2. Which revenue streams are underperforming vs projection?
    underperforming = [r for r in revenues if r['achievement_pct'] < 20 and r['annual_usd'] > 50_000]
    if underperforming:
        streams_bullets = ''.join([f"<li>{u['name']} — {u['achievement_pct']:.0f}% achieved ({fmt_usd(u['ytd_actual'])} of {fmt_usd(u['annual_usd'])} target)</li>" for u in underperforming[:5]])
        actions.append({
            'type': 'WARNING',
            'title': 'Underperforming Revenue Streams',
            'description': f"<p style='margin-bottom:8px'>These streams are below 20% of their annual target — investigate delays and re-baseline if needed:</p><ul style='margin:0 0 0 16px;padding:0;list-style:disc'>{streams_bullets}</ul>"
        })

    # 3. Monthly run-rate needed (dynamic)
    # Detect months with data from the revenues
    if revenues:
        monthly_totals = [sum(r['monthly'][i] for r in revenues) for i in range(12)]
        mwd = [i for i, v in enumerate(monthly_totals) if v > 0]
        months_elapsed = (max(mwd) + 1) if mwd else 1
    else:
        months_elapsed = 1
    months_remaining = 12 - months_elapsed
    if months_remaining > 0 and funding_gap_usd > 0:
        monthly_needed = funding_gap_usd / months_remaining
        current_monthly = annual_revenue_usd / months_elapsed
        actions.append({
            'type': 'INFO',
            'title': 'Required Monthly Run-Rate',
            'description': f"To close the {fmt_usd(funding_gap_usd)} gap in {months_remaining} months, need {fmt_usd(monthly_needed)}/month (vs current {fmt_usd(current_monthly)}/month)."
        })

    # 4. Upcoming start dates
    upcoming = [r for r in revenues if r['status'] != 'On Track' and r['status'] != 'Completed' and r['annual_usd'] > 50_000]
    if upcoming:
        dates_str = ', '.join([f"{u['name']} ({u['start_date']})" for u in upcoming[:3]])
        actions.append({
            'type': 'WATCH',
            'title': 'Upcoming Revenue Start Dates',
            'description': f"Monitor these deals starting soon: {dates_str}. Ensure implementation is on track."
        })

    # 5. Revenue concentration risk
    total_annual = sum(r['annual_usd'] for r in revenues)
    if total_annual > 0:
        top_stream = max(revenues, key=lambda x: x['annual_usd'])
        concentration = top_stream['annual_usd'] / total_annual * 100
        if concentration > 40:
            actions.append({
                'type': 'RISK',
                'title': 'Revenue Concentration Risk',
                'description': f"{top_stream['name']} represents {concentration:.0f}% of projected revenue. Diversification is critical—any delay directly impacts fundability."
            })

    return actions


def generate_html(revenues, budget_ngn, revenue_file_path, budget_file_path, output_path):
    """Generate interactive HTML dashboard"""
    from datetime import datetime as _dt

    # ── DYNAMIC MONTH DETECTION ──────────────────────────────────────
    monthly_totals = [sum(r['monthly'][i] for r in revenues) for i in range(12)]
    months_with_data_idx = [i for i, v in enumerate(monthly_totals) if v > 0]
    last_data_month = max(months_with_data_idx) if months_with_data_idx else 0
    num_months = last_data_month + 1
    today = _dt.now()
    is_partial = (last_data_month == today.month - 1)
    first_month_name = MONTH_NAMES[0]
    last_month_name = MONTH_NAMES[last_data_month]
    partial_note = f' ({last_month_name} partial)' if is_partial else ''
    ytd_label = f'{first_month_name} – {last_month_name} 2026{partial_note}'

    # Convert budget to USD for comparison
    budget_usd = budget_ngn / FX_RATE

    # Calculate KPIs (all revenue in USD)
    deal_bucket_total_usd = sum(r['annual_usd'] for r in revenues)  # $10M optimistic bucket
    annual_revenue_target_usd = 8_000_000  # $8M official company target
    ytd_actual_revenue_usd = sum(r['ytd_actual'] for r in revenues)
    # Achievement: actual vs full annual target (simple — what Finance expects)
    annual_progress_pct = (ytd_actual_revenue_usd / annual_revenue_target_usd * 100) if annual_revenue_target_usd > 0 else 0
    # Per-stream achievement vs their own annual targets (aggregate)
    total_stream_annual = sum(r['annual_usd'] for r in revenues)
    ytd_achievement_rate = (ytd_actual_revenue_usd / total_stream_annual * 100) if total_stream_annual > 0 else 0
    # Count active streams (those with annual target > 0)
    active_streams = len([r for r in revenues if r['annual_usd'] > 0])
    fundability_ratio = (annual_revenue_target_usd / budget_usd * 100) if budget_usd > 0 else 0

    # Scenarios
    optimistic_revenue_usd = annual_revenue_target_usd
    realistic_revenue_usd = annual_revenue_target_usd * 0.7
    conservative_revenue_usd = annual_revenue_target_usd * 0.4

    optimistic_coverage = (optimistic_revenue_usd / budget_usd * 100) if budget_usd > 0 else 0
    realistic_coverage = (realistic_revenue_usd / budget_usd * 100) if budget_usd > 0 else 0
    conservative_coverage = (conservative_revenue_usd / budget_usd * 100) if budget_usd > 0 else 0

    optimistic_gap_usd = max(0, budget_usd - optimistic_revenue_usd)
    realistic_gap_usd = max(0, budget_usd - realistic_revenue_usd)
    conservative_gap_usd = max(0, budget_usd - conservative_revenue_usd)

    # Revenue by Rail
    rails = {}
    for r in revenues:
        if r['rail'] not in rails:
            rails[r['rail']] = {'annual': 0, 'actual': 0}
        rails[r['rail']]['annual'] += r['annual_usd']
        rails[r['rail']]['actual'] += r['ytd_actual']

    rail_labels = list(rails.keys())
    rail_annual = [rails[r]['annual'] for r in rail_labels]
    rail_actual = [rails[r]['actual'] for r in rail_labels]

    # Status distribution
    status_counts = {}
    for r in revenues:
        status_counts[r['status']] = status_counts.get(r['status'], 0) + 1

    # Generate critical actions
    critical_actions = generate_critical_actions(revenues, budget_ngn, annual_revenue_target_usd)

    # Revenue streams sorted by annual amount
    revenues_sorted = sorted(revenues, key=lambda x: -x['annual_usd'])

    # Build revenue streams table
    revenue_rows = ""
    for r in revenues_sorted:
        # Use the ACTUAL status from the sheet (On Track, Off Track, At Risk, Closed)
        actual_status = r['status']
        status_color_map = {
            'On Track': '#00D4AA',
            'Closed': '#4ECDC4',
            'At Risk': '#FFE66D',
            'Off Track': '#FF6B6B',
        }
        status_color = status_color_map.get(actual_status, '#94a3b8')

        # Deficit/surplus display
        deficit_val = r.get('deficit', 0)
        surplus_val = r.get('surplus', 0)
        if surplus_val > 0:
            ds_display = f'<span class="positive">+{fmt_usd(surplus_val)}</span>'
        elif deficit_val > 0:
            ds_display = f'<span class="negative">-{fmt_usd(deficit_val)}</span>'
        else:
            ds_display = '<span style="color:var(--text-tertiary)">—</span>'

        revenue_rows += f"""<tr>
<td>{r['name']}</td>
<td>{r['rail']}</td>
<td>{r['start_date']}</td>
<td class="positive">{fmt_dual(r['annual_usd'])}</td>
<td class="positive">{fmt_dual(r['ytd_actual'])}</td>
<td>{r['achievement_pct']:.0f}%</td>
<td style="color:{status_color};font-weight:600">{actual_status}</td>
<td>{ds_display}</td>
</tr>"""

    # Build critical actions HTML
    actions_html = ""
    for action in critical_actions:
        if action['type'] == 'CRITICAL':
            badge_color = '#FF6B6B'
            badge_text = 'CRITICAL'
            icon = '🚨'
        elif action['type'] == 'WARNING':
            badge_color = '#FFE66D'
            badge_text = 'WARNING'
            icon = '⚠️'
        elif action['type'] == 'RISK':
            badge_color = '#FF6B6B'
            badge_text = 'RISK'
            icon = '⚠️'
        else:
            badge_color = '#4ECDC4'
            badge_text = 'WATCH'
            icon = '👁️'

        actions_html += f"""<div class="action-card" style="border-left-color:{badge_color}">
<div class="action-header">
<span class="action-icon">{icon}</span>
<span class="action-title">{action['title']}</span>
<span class="action-badge" style="background:rgba({int(badge_color[1:3], 16)},{int(badge_color[3:5], 16)},{int(badge_color[5:7], 16)},0.15);color:{badge_color}">{badge_text}</span>
</div>
<div class="action-description">{action['description']}</div>
</div>"""

    # Executive takeaways
    takeaways_html = ""

    # Takeaway 1: Fundability Assessment
    if fundability_ratio >= 100:
        fundability_headline = f"Projected revenue fully funds the budget"
        fundability_body = f"<ul style='margin:6px 0 0 16px;padding:0;list-style:disc'><li>Revenue target: {fmt_dual(annual_revenue_target_usd)}</li><li>Annual budget: {fmt_naira(budget_ngn)}</li><li>Funding ratio: {fundability_ratio:.0f}%</li><li>Strong revenue diversification and deal confidence</li></ul>"
        fundability_color = '#00D4AA'
        fundability_badge = 'HEALTHY'
    elif fundability_ratio >= 70:
        fundability_headline = f"Achievable shortfall at full conversion"
        fundability_body = f"<ul style='margin:6px 0 0 16px;padding:0;list-style:disc'><li>Revenue target: {fmt_dual(annual_revenue_target_usd)}</li><li>Annual budget: {fmt_naira(budget_ngn)}</li><li>Funding ratio: {fundability_ratio:.0f}%</li><li>At 70% conversion, gap is {fmt_dual(realistic_gap_usd)}</li><li>Focus on closing top deals to bridge the gap</li></ul>"
        fundability_color = '#FFE66D'
        fundability_badge = 'MONITOR'
    else:
        fundability_headline = f"Significant funding gap even at 100% conversion"
        fundability_body = f"<ul style='margin:6px 0 0 16px;padding:0;list-style:disc'><li>Revenue target: {fmt_dual(annual_revenue_target_usd)}</li><li>Annual budget: {fmt_naira(budget_ngn)}</li><li>Funding ratio: {fundability_ratio:.0f}%</li><li>New revenue streams or cost optimization required</li></ul>"
        fundability_color = '#FF6B6B'
        fundability_badge = 'CRITICAL'

    takeaways_html += f"""<div class="takeaway-card" style="border-left-color:{fundability_color}">
<div class="tw-header"><span class="tw-icon">💰</span><span class="tw-title">Fundability Assessment</span><span class="tw-badge" style="background:rgba({int(fundability_color[1:3], 16)},{int(fundability_color[3:5], 16)},{int(fundability_color[5:7], 16)},0.15);color:{fundability_color}">{fundability_badge}</span></div>
<div class="tw-headline" style="color:{fundability_color}">{fundability_headline}</div>
<div class="tw-body">{fundability_body}</div>
</div>"""

    # Takeaway 2: Revenue Health (using actual status from sheet)
    on_track = [r for r in revenues if r['status'] == 'On Track']
    at_risk = [r for r in revenues if r['status'] == 'At Risk']
    off_track = [r for r in revenues if r['status'] == 'Off Track']
    closed = [r for r in revenues if r['status'] == 'Closed']

    # Streams with surplus (exceeding targets)
    surplus_streams = [r for r in revenues if r.get('surplus', 0) > 0]
    # Streams with deficit (behind targets)
    deficit_streams = [r for r in revenues if r.get('deficit', 0) > 0]

    health_headline = f"{len(on_track)} On Track, {len(closed)} Closed, {len(at_risk)} At Risk, {len(off_track)} Off Track"
    health_bullets = []
    health_bullets.append(f"<li>Annual progress: {annual_progress_pct:.0f}% — {fmt_dual(ytd_actual_revenue_usd)} of {fmt_dual(annual_revenue_target_usd)}</li>")
    health_bullets.append(f"<li>Pipeline achievement: {ytd_achievement_rate:.0f}% of stream targets ({active_streams} active streams)</li>")
    if surplus_streams:
        total_surplus = sum(r.get('surplus', 0) for r in surplus_streams)
        health_bullets.append(f"<li>{len(surplus_streams)} streams exceeding targets (total surplus: {fmt_usd(total_surplus)})</li>")
    if deficit_streams:
        total_deficit = sum(r.get('deficit', 0) for r in deficit_streams)
        health_bullets.append(f"<li>{len(deficit_streams)} streams below target (total deficit: {fmt_usd(total_deficit)})</li>")
    performing = surplus_streams  # For the rest of the logic
    critical = off_track + at_risk
    # Strong performers = streams that are On Track AND have actually generated revenue
    actual_performers = sorted([r for r in on_track if r['ytd_actual'] > 0], key=lambda x: -x['ytd_actual'])
    if actual_performers:
        perf_str = ', '.join([f"{p['name']} ({fmt_usd(p['ytd_actual'])})" for p in actual_performers[:3]])
        health_bullets.append(f"<li><strong>Top performers:</strong> {perf_str}</li>")
    # Streams that are On Track but have zero revenue (pipeline only)
    pipeline_only = [r for r in on_track if r['ytd_actual'] == 0 and r['annual_usd'] > 50_000]
    if pipeline_only:
        health_bullets.append(f"<li>{len(pipeline_only)} On Track deals still in pipeline (no revenue yet)</li>")
    if critical:
        crit_str = ', '.join([c['name'] for c in critical[:3]])
        health_bullets.append(f"<li><strong>Needs attention:</strong> {crit_str}</li>")
    health_body = "<ul style='margin:6px 0 0 16px;padding:0;list-style:disc'>" + ''.join(health_bullets) + "</ul>"

    health_color = '#00D4AA' if len(critical) == 0 else '#FFE66D' if len(critical) <= 2 else '#FF6B6B'
    health_badge = 'HEALTHY' if len(critical) == 0 else 'MONITOR' if len(critical) <= 2 else 'NEEDS ACTION'

    takeaways_html += f"""<div class="takeaway-card" style="border-left-color:{health_color}">
<div class="tw-header"><span class="tw-icon">📊</span><span class="tw-title">Revenue Health</span><span class="tw-badge" style="background:rgba({int(health_color[1:3], 16)},{int(health_color[3:5], 16)},{int(health_color[5:7], 16)},0.15);color:{health_color}">{health_badge}</span></div>
<div class="tw-headline" style="color:{health_color}">{health_headline}</div>
<div class="tw-body">{health_body}</div>
</div>"""

    # Takeaway 3: Risk Alerts
    risk_alerts = []

    # Check concentration
    total_annual = sum(r['annual_usd'] for r in revenues)
    if total_annual > 0:
        top = max(revenues, key=lambda x: x['annual_usd'])
        top_conc = top['annual_usd'] / total_annual * 100
        if top_conc > 50:
            risk_alerts.append(f"<li><strong>Concentration Risk:</strong> {top['name']} is {top_conc:.0f}% of pipeline</li>")

    # Check underperformers
    underperf = [r for r in revenues if r['achievement_pct'] < 10 and r['annual_usd'] > 50_000]
    if underperf:
        underperf_str = ', '.join([u['name'] for u in underperf[:2]])
        risk_alerts.append(f"<li><strong>Underperformance:</strong> {underperf_str} significantly behind target</li>")

    # Check upcoming starts
    upcoming_risky = [r for r in revenues if r['status'] not in ('On Track', 'Completed') and r['annual_usd'] > 100_000]
    if upcoming_risky:
        risk_alerts.append(f"<li><strong>Execution Risk:</strong> {len(upcoming_risky)} major deals not yet started</li>")

    if not risk_alerts:
        risk_alerts.append("<li>No major risks detected — portfolio is well-balanced</li>")

    risk_body = "<ul style='margin:6px 0 0 16px;padding:0;list-style:disc'>" + ''.join(risk_alerts) + "</ul>"

    takeaways_html += f"""<div class="takeaway-card" style="border-left-color:#FF6B6B">
<div class="tw-header"><span class="tw-icon">⚠️</span><span class="tw-title">Risk Alerts</span><span class="tw-badge" style="background:rgba(255,107,107,0.15);color:#FF6B6B">MONITOR</span></div>
<div class="tw-headline" style="color:#FF6B6B">Key risks to fundability</div>
<div class="tw-body">{risk_body}</div>
</div>"""

    # Build chart data
    rail_colors = ['#00D4AA', '#4ECDC4', '#FFE66D', '#FF6B6B', '#A8E6CF', '#95E1D3', '#F38181', '#AA96DA']

    # Generate at timestamp
    generated_at = datetime.now().strftime('%d %b %Y %H:%M')

    # Theme system
    theme_css = get_base_css()
    theme_toggle = get_toggle_html()
    theme_js = get_theme_js()

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Seamfix Revenue & Fundability Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
{theme_css}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:var(--bg-body);color:var(--text-primary);padding:0;min-height:100vh}}
.container{{max-width:1600px;margin:0 auto;padding:0 28px 28px}}
.header{{margin-bottom:32px;padding-bottom:20px;border-bottom:2px solid var(--border-accent)}}
.header h1{{font-size:2.4em;font-weight:700;background:linear-gradient(135deg,var(--accent),var(--accent-secondary));-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:4px}}
.header .sub{{color:var(--text-secondary);font-size:0.95em}}
.header .meta{{color:var(--text-tertiary);font-size:0.8em;margin-top:6px}}

.nav-bar{{display:none!important}}

.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;margin-bottom:36px}}
.kpi-card{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:22px;transition:all 0.3s ease}}
.kpi-card:hover{{border-color:var(--border-accent);transform:translateY(-3px);box-shadow:var(--shadow-hover)}}
.kpi-label{{font-size:0.78em;color:var(--text-secondary);text-transform:uppercase;letter-spacing:1.2px;margin-bottom:8px;font-weight:600}}
.kpi-value{{font-size:1.4em;font-weight:700;color:var(--accent);margin-bottom:6px}}
.kpi-value.negative{{color:#FF6B6B}}
.kpi-secondary{{font-size:0.85em;color:var(--text-tertiary);margin-bottom:4px}}
.kpi-change{{font-size:0.82em;color:var(--text-secondary)}}

.charts-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(480px,1fr));gap:20px;margin-bottom:36px}}
.chart-box{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:22px;min-height:380px}}
.chart-box h3{{font-size:1em;font-weight:600;margin-bottom:16px;color:var(--text-heading)}}

.section{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:24px;margin-bottom:24px}}
.section h2{{font-size:1.3em;margin-bottom:16px;color:var(--accent)}}

.fundability-section{{border-color:rgba(78,205,196,0.2)}}
.fundability-section h2{{color:var(--accent-secondary)}}



.scenario-label{{font-size:0.8em;color:var(--text-secondary);text-transform:uppercase;margin-bottom:10px;font-weight:600}}
.scenario-value{{font-size:1.2em;font-weight:700;color:var(--accent);margin-bottom:4px}}
.scenario-secondary{{font-size:0.75em;color:var(--text-tertiary);margin-bottom:6px}}
.scenario-coverage{{font-size:0.9em;color:var(--accent-secondary);margin-bottom:4px}}
.scenario-gap{{font-size:0.85em;color:var(--warning)}}

.gauge{{width:100%;height:24px;background:var(--bg-gauge);border-radius:12px;overflow:hidden;margin-top:8px}}
.gauge-fill{{height:100%;background:linear-gradient(90deg,var(--danger),var(--warning),var(--accent));border-radius:12px;transition:width 0.3s}}

.action-card{{background:var(--danger-bg);border-left:4px solid var(--danger);border-radius:8px;padding:16px;margin-bottom:12px}}
.action-header{{display:flex;align-items:center;gap:8px;margin-bottom:8px}}
.action-icon{{font-size:1.2em}}
.action-title{{font-size:0.9em;font-weight:700;color:var(--text-heading);text-transform:uppercase;letter-spacing:0.5px}}
.action-badge{{font-size:0.7em;font-weight:700;padding:3px 10px;border-radius:12px;letter-spacing:0.5px}}
.action-description{{font-size:0.85em;color:var(--text-secondary);line-height:1.6}}

.takeaways-section{{margin-bottom:36px}}
.takeaways-section h2{{font-size:1.4em;margin-bottom:16px;color:var(--warning);display:flex;align-items:center;gap:10px}}
.takeaways-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(440px,1fr));gap:16px}}
.takeaway-card{{background:var(--bg-card);border:1px solid var(--border-light);border-left:4px solid var(--text-tertiary);border-radius:10px;padding:20px;transition:all 0.3s ease}}
.takeaway-card:hover{{transform:translateY(-2px);box-shadow:var(--shadow-hover)}}
.tw-header{{display:flex;align-items:center;gap:8px;margin-bottom:10px}}
.tw-icon{{font-size:1.2em}}
.tw-title{{font-size:0.82em;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:var(--text-secondary)}}
.tw-badge{{font-size:0.7em;font-weight:700;padding:3px 10px;border-radius:12px;letter-spacing:0.5px}}
.tw-headline{{font-size:1.05em;font-weight:600;margin-bottom:8px;line-height:1.4}}
.tw-body{{font-size:0.88em;color:var(--text-secondary);line-height:1.65}}

table{{width:100%;border-collapse:collapse;font-size:0.85em}}
thead{{background:var(--bg-table-header)}}
th{{padding:10px 12px;text-align:left;font-weight:600;color:var(--accent);border-bottom:2px solid var(--border-accent)}}
td{{padding:10px 12px;border-bottom:1px solid var(--border-light);color:var(--text-heading)}}
tbody tr:hover{{background:var(--bg-table-hover)}}
.positive{{color:#00D4AA}}
.negative{{color:#FF6B6B}}

th.sortable{{cursor:pointer;user-select:none}}
th.sortable:hover{{color:var(--warning)}}
th.sort-asc::after{{content:' ▲';font-size:0.7em}}
th.sort-desc::after{{content:' ▼';font-size:0.7em}}

.top-nav{{background:var(--bg-nav);border-bottom:1px solid var(--border-main);padding:0 24px;display:flex;align-items:center;height:48px;overflow-x:auto;position:sticky;top:0;z-index:200}}
.top-nav-brand{{color:var(--text-primary);font-weight:700;font-size:15px;margin-right:24px;white-space:nowrap;text-decoration:none}}
.top-nav-link{{color:var(--text-secondary);text-decoration:none;padding:0 14px;height:48px;display:flex;align-items:center;font-size:13px;border-bottom:2px solid transparent;white-space:nowrap;transition:color .2s}}
.top-nav-link:hover{{color:var(--text-primary)}}
.top-nav-link.active{{color:var(--text-primary);border-bottom-color:var(--accent);font-weight:500}}
.pdf-btn{{display:none!important}}

@media(max-width:1024px){{.charts-grid{{grid-template-columns:1fr}}.kpi-grid{{grid-template-columns:repeat(2,1fr)}}.takeaways-grid{{grid-template-columns:1fr}}}}
@media(max-width:640px){{.kpi-grid{{grid-template-columns:1fr}}.header h1{{font-size:1.6em}}.nav-bar{{flex-wrap:wrap}}}}

@media print{{
.pdf-btn{{display:none!important}}
#revenueSearch{{display:none!important}}
#searchResults{{display:none!important}}
*{{color-adjust:exact!important;-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}}
body{{background:#fff!important;color:#1e293b!important;padding:12px!important;font-size:9pt!important}}
.container{{max-width:100%}}
.header h1{{font-size:1.6em!important;background:none!important;-webkit-text-fill-color:#0f172a!important;color:#0f172a!important}}
.header .sub,.header .meta{{color:#475569!important}}
.nav-bar{{background:#f8fafc!important;border-color:#e2e8f0!important;display:none!important}}
.kpi-grid{{grid-template-columns:repeat(6,1fr)!important;gap:8px!important;margin-bottom:16px!important}}
.kpi-card{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:12px!important;box-shadow:none!important;page-break-inside:avoid}}
.kpi-label{{color:#475569!important;font-size:0.65em!important}}
.kpi-value{{color:#0f172a!important;font-size:1.2em!important}}
.kpi-value.negative{{color:#dc2626!important}}
.kpi-secondary{{color:#64748b!important;font-size:0.7em!important}}
.takeaways-section{{page-break-inside:avoid}}
.takeaways-section h2{{color:#0f172a!important;font-size:1.1em!important}}
.takeaways-grid{{grid-template-columns:repeat(2,1fr)!important}}
.takeaway-card{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:14px!important;page-break-inside:avoid}}
.tw-title{{color:#475569!important}}
.tw-body{{color:#475569!important}}
.charts-grid{{grid-template-columns:repeat(2,1fr)!important;margin-bottom:16px!important}}
.chart-box{{background:#fff!important;border:1px solid #e2e8f0!important;padding:14px!important;min-height:auto!important;page-break-inside:avoid}}
.section{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:16px!important;page-break-inside:avoid}}
.section h2{{color:#0f172a!important}}
.fundability-section{{border-color:#b3e5fc!important}}


.scenario-label{{color:#475569!important}}
.scenario-value{{color:#0f172a!important}}
.scenario-secondary{{color:#64748b!important}}
.scenario-coverage{{color:#0f172a!important}}
.scenario-gap{{color:#ea580c!important}}
.action-card{{background:#fef2f2!important;border-color:#fca5a5!important}}
.action-title{{color:#0f172a!important}}
.action-description{{color:#475569!important}}
table{{font-size:0.75em!important}}
thead{{background:#f1f5f9!important}}
th{{color:#0f172a!important;border-bottom-color:#cbd5e1!important}}
td{{color:#334155!important;border-bottom-color:#e2e8f0!important}}
tbody tr:hover{{background:transparent!important}}
}}
.dashboard-footer{{margin-top:48px;padding:20px 28px;border-top:1px solid var(--border-main);color:var(--text-secondary);font-size:12px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
.dashboard-footer span{{color:var(--text-secondary)}}
/* ── HEADER: match Pipeline Intelligence style ── */
.header{{padding:24px 28px 16px!important;border-bottom:1px solid var(--border-main)!important;margin-bottom:24px!important;background:none!important}}
.header h1{{font-size:22px!important;font-weight:700!important;background:none!important;-webkit-background-clip:unset!important;-webkit-text-fill-color:var(--text-primary)!important;color:var(--text-primary)!important;margin-bottom:4px!important}}
.header .sub{{font-size:13px!important;color:var(--text-secondary)!important}}
.header .meta{{font-size:12px!important;color:var(--text-secondary)!important;margin-top:4px!important}}
.header p{{font-size:13px!important;color:var(--text-secondary)!important;margin-top:4px!important}}
</style>
</head>
<body>
{theme_toggle}
<nav class="top-nav"><span class="top-nav-brand">⚡ Seamfix</span><a href="dashboard.html" class="top-nav-link ">Cash Overview</a><a href="expense_dashboard.html" class="top-nav-link ">Expense &amp; Vendor</a><a href="budget_dashboard.html" class="top-nav-link ">Budget vs Actual</a><a href="revenue_dashboard.html" class="top-nav-link active">Revenue &amp; Fundability</a><a href="pipeline_dashboard.html" class="top-nav-link ">Pipeline Intelligence</a></nav>
<!-- PDF button hidden per user request -->

<div class="container">
<div class="header">
<h1>Seamfix Revenue & Fundability Dashboard</h1>
<div class="sub">2026 Budget Coverage Analysis &mdash; Path to Revenue (All amounts in USD with Naira equivalents)</div>
<div class="meta">Data as of: <strong>{ytd_label}</strong> &nbsp;&bull;&nbsp; Generated: {generated_at}</div>
</div>

<!-- old nav-bar replaced by top-nav -->

<div class="kpi-grid">
<div class="kpi-card">
<div class="kpi-label">Annual Revenue Target</div>
<div class="kpi-value">{fmt_usd(annual_revenue_target_usd)}</div>
<div class="kpi-secondary">{fmt_naira(annual_revenue_target_usd * FX_RATE)}</div>
<div class="kpi-change">{len(revenues)} streams across pipeline</div>
</div>
<div class="kpi-card">
<div class="kpi-label">YTD Actual Revenue</div>
<div class="kpi-value">{fmt_usd(ytd_actual_revenue_usd)}</div>
<div class="kpi-secondary">{fmt_naira(ytd_actual_revenue_usd * FX_RATE)}</div>
<div class="kpi-change">{ytd_label}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Annual Progress</div>
<div class="kpi-value">{annual_progress_pct:.0f}%</div>
<div class="kpi-secondary">{fmt_usd(ytd_actual_revenue_usd)} of {fmt_usd(annual_revenue_target_usd)}</div>
<div class="kpi-change">Actual earned vs full-year target</div>
</div>
<div class="kpi-card">
<div class="kpi-label">YTD Achievement ({ytd_label})</div>
<div class="kpi-value {'negative' if ytd_achievement_rate < 20 else ''}">{ytd_achievement_rate:.0f}%</div>
<div class="kpi-secondary">{fmt_usd(ytd_actual_revenue_usd)} of {fmt_usd(total_stream_annual)} pipeline</div>
<div class="kpi-change">{active_streams} of {len(revenues)} streams active {ytd_label}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Budget Fundability Score</div>
<div class="kpi-value {'negative' if fundability_ratio < 100 else ''}">{fundability_ratio:.0f}%</div>
<div class="kpi-secondary">{fmt_naira(budget_ngn)}</div>
<div class="kpi-change">Annual revenue / budget</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Funding Gap (100% conversion)</div>
<div class="kpi-value negative">{fmt_usd(max(0, budget_usd - annual_revenue_target_usd))}</div>
<div class="kpi-secondary">{fmt_naira(max(0, budget_usd - annual_revenue_target_usd) * FX_RATE)}</div>
<div class="kpi-change">Additional revenue needed</div>
</div>
</div>

<div class="takeaways-section">
<h2>📋 Executive Takeaways</h2>
<div class="takeaways-grid">
{takeaways_html}
</div>
</div>

<div class="section fundability-section">
<h2>💼 Revenue Projection</h2>
<p style="color:var(--text-secondary);margin-bottom:16px">Status-weighted projections are on <a href="pipeline_dashboard.html" style="color:var(--accent);text-decoration:none">Pipeline Intelligence</a> — deal-level scenarios built from actual On Track / At Risk / Off Track weights are more accurate than flat conversion rates.</p>
<div style="background:var(--accent-bg);border:1px solid var(--border-accent);border-radius:10px;padding:16px 20px">
  <div style="display:flex;gap:32px;flex-wrap:wrap">
    <div><div style="font-size:11px;color:var(--text-tertiary);text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Annual Target</div><div style="font-size:1.4em;font-weight:700;color:var(--accent)">{fmt_usd(annual_revenue_target_usd)}</div></div>
    <div><div style="font-size:11px;color:var(--text-tertiary);text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Budget</div><div style="font-size:1.4em;font-weight:700;color:var(--text-primary)">{fmt_naira(budget_ngn)}</div></div>
    <div><div style="font-size:11px;color:var(--text-tertiary);text-transform:uppercase;letter-spacing:1px;margin-bottom:4px">Full-conversion Coverage</div><div style="font-size:1.4em;font-weight:700;color:#{'00D4AA' if optimistic_coverage >= 100 else 'FFE66D'}">{optimistic_coverage:.0f}%</div></div>
  </div>
</div>
</div>

<div class="section">
<h2>🚨 Critical Actions Engine</h2>
<p style="color:var(--text-secondary);margin-bottom:16px">Auto-generated recommendations based on pipeline health and fundability scenarios.</p>
{actions_html}
</div>

<div class="charts-grid">
<div class="chart-box"><h3>Revenue Pipeline by Rail (USD)</h3><canvas id="railChart"></canvas></div>
<div class="chart-box"><h3>YTD Performance vs Target (USD)</h3><canvas id="perfChart"></canvas></div>
</div>

<div class="section">
<h2>📊 Revenue Streams Overview</h2>
<p style="color:var(--text-secondary);margin-bottom:12px;font-size:0.9em">All revenue streams sorted by projected annual amount. Values in USD (₦ equivalent). Status from revenue tracker.</p>
<div style="margin-bottom:16px">
<input type="text" id="revenueSearch" placeholder="Search revenue streams (e.g. NIMC, MTN, Fixiam, Benin...)" style="width:100%;max-width:500px;padding:10px 16px;border-radius:8px;border:1px solid var(--border-accent);background:var(--bg-input);color:var(--text-primary);font-family:inherit;font-size:0.9em;outline:none" onfocus="this.style.borderColor='var(--accent)'" onblur="this.style.borderColor='var(--border-accent)'" oninput="filterRevenueTable()">
</div>
<table id="revenueTable">
<thead><tr>
<th class="sortable" onclick="sortTable('revenueTable',0,'text')">Revenue Stream</th>
<th class="sortable" onclick="sortTable('revenueTable',1,'text')">Rail</th>
<th class="sortable" onclick="sortTable('revenueTable',2,'text')">Start Date</th>
<th class="sortable" onclick="sortTable('revenueTable',3,'money')">Annual Target</th>
<th class="sortable" onclick="sortTable('revenueTable',4,'money')">YTD Actual</th>
<th class="sortable" onclick="sortTable('revenueTable',5,'pct')">Achievement %</th>
<th class="sortable" onclick="sortTable('revenueTable',6,'status')">Status</th>
<th class="sortable" onclick="sortTable('revenueTable',7,'money')">Surplus / Deficit</th>
</tr></thead>
<tbody id="revenueBody">
{revenue_rows}
</tbody>
<tfoot style="background:var(--bg-table-header);font-weight:700">
<tr>
<td colspan="3" style="color:var(--accent);font-weight:700">TOTAL ({len(revenues)} streams)</td>
<td class="positive">{fmt_dual(annual_revenue_target_usd)}</td>
<td class="positive">{fmt_dual(ytd_actual_revenue_usd)}</td>
<td>{annual_progress_pct:.0f}% of $8M target / {ytd_achievement_rate:.0f}% of pipeline</td>
<td></td>
<td></td>
</tr>
</tfoot>
</table>
<div id="searchResults" style="color:var(--text-tertiary);font-size:0.8em;margin-top:8px"></div>
</div>
</div>

<script>
function sortTable(tableId, colIdx, type) {{
    const table = document.getElementById(tableId);
    const tbody = table.querySelector('tbody');
    if (!tbody) return;

    const rows = Array.from(tbody.querySelectorAll('tr'));
    const headers = table.querySelectorAll('th.sortable');
    const tfoot = table.querySelector('tfoot');

    // Filter out tfoot rows
    const bodyRows = rows.filter(row => !tfoot || !tfoot.contains(row));

    // Determine sort direction
    const header = headers[colIdx];
    const isAsc = header.classList.contains('sort-asc');

    // Clear previous sort indicators
    headers.forEach(h => {{
        h.classList.remove('sort-asc', 'sort-desc');
    }});

    // Sort rows
    bodyRows.sort((a, b) => {{
        let aVal = a.cells[colIdx].textContent.trim();
        let bVal = b.cells[colIdx].textContent.trim();

        if (type === 'money') {{
            aVal = parseFloat(aVal.replace(/[₦$B M K(),\-]/g, '')) || 0;
            bVal = parseFloat(bVal.replace(/[₦$B M K(),\-]/g, '')) || 0;
        }} else if (type === 'pct') {{
            aVal = parseFloat(aVal) || 0;
            bVal = parseFloat(bVal) || 0;
        }} else if (type === 'status') {{
            const statusOrder = {{'On Track': 0, 'Closed': 1, 'At Risk': 2, 'Off Track': 3, 'Unknown': 4}};
            aVal = statusOrder[aVal] ?? 4;
            bVal = statusOrder[bVal] ?? 4;
        }} else {{
            aVal = aVal.toLowerCase();
            bVal = bVal.toLowerCase();
        }}

        return isAsc ? (aVal > bVal ? 1 : -1) : (aVal < bVal ? 1 : -1);
    }});

    // Re-append body rows (keep tfoot in place)
    bodyRows.forEach(row => tbody.appendChild(row));
    header.classList.add(isAsc ? 'sort-desc' : 'sort-asc');
}}

const C = {{p:'#00D4AA',s:'#4ECDC4',a:'#FFE66D',d:'#FF6B6B',g:'rgba(0,212,170,0.08)'}};
const baseOpts = {{
    responsive:true,
    maintainAspectRatio:true,
    plugins:{{legend:{{labels:{{color:'#94a3b8',font:{{family:"'Inter',sans-serif",size:11}}}}}}}},
    scales:{{x:{{grid:{{color:C.g}},ticks:{{color:'#64748b'}}}},y:{{grid:{{color:C.g}},ticks:{{color:'#64748b',callback:function(v){{if(v>=1e6)return'$'+(v/1e6).toFixed(1)+'M';if(v>=1e3)return'$'+(v/1e3).toFixed(0)+'K';return'$'+v}}}}}}}}
}};

const railLabels = {json.dumps(rail_labels)};
const railAnnual = {json.dumps(rail_annual)};
const railActual = {json.dumps(rail_actual)};

new Chart(document.getElementById('railChart'),{{
    type:'doughnut',
    data:{{
        labels:railLabels,
        datasets:[
            {{label:'Projected Annual (USD)',data:railAnnual,backgroundColor:['{rail_colors[0]}','{rail_colors[1]}','{rail_colors[2]}','{rail_colors[3]}','{rail_colors[4]}','{rail_colors[5]}','{rail_colors[6]}','{rail_colors[7]}']}}
        ]
    }},
    options:{{responsive:true,maintainAspectRatio:true,plugins:{{legend:{{labels:{{color:'#94a3b8',font:{{family:"'Inter',sans-serif",size:11}}}}}}}},}}
}});

const streams = {json.dumps([r['name'][:25] for r in revenues_sorted[:10]])};
const annuals = {json.dumps([r['annual_usd'] for r in revenues_sorted[:10]])};
const ytds = {json.dumps([r['ytd_actual'] for r in revenues_sorted[:10]])};

new Chart(document.getElementById('perfChart'),{{
    type:'bar',
    data:{{
        labels:streams,
        datasets:[
            {{label:'Annual Target (USD)',data:annuals,backgroundColor:'rgba(100,116,139,0.5)',borderColor:'rgba(100,116,139,0.8)',borderWidth:1,borderRadius:4}},
            {{label:'YTD Actual (USD)',data:ytds,backgroundColor:'rgba(0,212,170,0.8)',borderColor:'rgba(0,212,170,1)',borderWidth:1,borderRadius:4}}
        ]
    }},
    options:{{...baseOpts,indexAxis:undefined}}
}});

function filterRevenueTable() {{
    const q = document.getElementById('revenueSearch').value.toLowerCase();
    const rows = document.getElementById('revenueBody').getElementsByTagName('tr');
    let shown = 0;
    for (let i = 0; i < rows.length; i++) {{
        const text = rows[i].textContent.toLowerCase();
        if (q === '' || text.includes(q)) {{
            rows[i].style.display = '';
            shown++;
        }} else {{
            rows[i].style.display = 'none';
        }}
    }}
    const info = document.getElementById('searchResults');
    if (q) {{
        info.textContent = shown + ' of ' + rows.length + ' streams matching "' + document.getElementById('revenueSearch').value + '"';
    }} else {{
        info.textContent = '';
    }}
}}
{theme_js}
</script>
<div class="dashboard-footer">
<span>Seamfix Financial Intelligence &nbsp;·&nbsp; Powered by Claude Cowork</span>
<span>Generated: {generated_at} &nbsp;&bull;&nbsp; $1 = ₦1,450</span>
</div>
</body>
</html>"""

    with open(output_path, 'w') as f:
        f.write(html)


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else "/sessions/relaxed-busy-archimedes/mnt/Seamfix Cash Reports 2026/"

    if not os.path.isdir(folder):
        print(f"Error: {folder} not found")
        sys.exit(1)

    # Find revenue file
    revenue_file = find_file(folder, "Path to Revenue")
    if not revenue_file:
        revenue_file = find_file(folder, "Revenue")
    if not revenue_file:
        print("Error: Revenue file not found")
        sys.exit(1)

    print(f"Revenue file: {revenue_file}")

    # Find budget file
    budget_file = find_file(folder, "BUDGET")
    if not budget_file:
        budget_file = find_file(folder, "budget")
    if not budget_file:
        print("Error: Budget file not found")
        sys.exit(1)

    print(f"Budget file: {budget_file}")

    # Extract data
    print("\nExtracting revenue streams...")
    revenues = extract_revenue_data(revenue_file)
    print(f"Found {len(revenues)} revenue streams")

    print("Extracting budget...")
    annual_budget_ngn = extract_budget_data(budget_file)
    print(f"Annual budget: {fmt_naira(annual_budget_ngn)} ({fmt_usd(annual_budget_ngn / FX_RATE)})")

    # Generate dashboard
    output_file = os.path.join(folder, 'revenue_dashboard.html')
    print(f"\nGenerating dashboard...")
    generate_html(revenues, annual_budget_ngn, revenue_file, budget_file, output_file)

    print(f"✓ Dashboard generated: {output_file}")

    # Also save to outputs if available
    outputs = "/sessions/relaxed-busy-archimedes/mnt/outputs/"
    if os.path.isdir(outputs):
        import shutil
        shutil.copy2(output_file, os.path.join(outputs, 'revenue_dashboard.html'))
        print(f"✓ Also saved to: {outputs}revenue_dashboard.html")


if __name__ == '__main__':
    main()
