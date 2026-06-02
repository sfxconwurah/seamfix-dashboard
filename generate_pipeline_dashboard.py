#!/usr/bin/env python3
"""
Seamfix Pipeline Intelligence Dashboard Generator
Analyses revenue pipeline by status, momentum, and deal-level intelligence.
Usage: python3 generate_pipeline_dashboard.py [folder_path]
"""

import os, sys, json, re, glob
from datetime import datetime
from openpyxl import load_workbook
from theme import get_base_css, get_toggle_html, get_theme_js


FX_RATE = 1450  # $1 USD = ₦1,450 NGN

SECTION_HEADERS = {"ANCHOR DEALS", "EXISTING CUSTOMERS", "DEALS FROM 2025", "NEW BUSINESS", "TOTAL"}

# Monthly actual columns: M=Jan, N=Feb, ..., X=Dec (convention set by Finance)
MONTH_COLUMNS = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTH_COLORS = [
    '#3b82f6', '#10b981', '#8b5cf6', '#f59e0b', '#ec4899', '#06b6d4',
    '#84cc16', '#f97316', '#6366f1', '#14b8a6', '#e11d48', '#a855f7',
]

# Projection weights by status
STATUS_WEIGHTS = {
    'On Track':  1.00,
    'Closed':    1.00,
    'At Risk':   0.50,
    'Off Track': 0.10,
    '':          0.70,   # unknown/blank
    'Unknown':   0.70,
}

# Conservative scenario (if at-risk & off-track stay unresolved)
STATUS_WEIGHTS_CONSERVATIVE = {
    'On Track':  1.00,
    'Closed':    1.00,
    'At Risk':   0.00,
    'Off Track': 0.00,
    '':          0.50,
    'Unknown':   0.50,
}

LANDING_ZONE = 8_000_000  # $8M official annual target (deal bucket is $10M optimistic)


def sf(val):
    if val is None: return 0.0
    try: return float(val)
    except: return 0.0


def fmt_usd(val, decimals=2):
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000:
        return f"{sign}${v/1_000_000:.{decimals}f}M"
    elif v >= 1_000:
        return f"{sign}${v/1_000:.1f}K"
    else:
        return f"{sign}${v:,.0f}"


def fmt_naira(val):
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000_000:
        return f"{sign}₦{v/1_000_000_000:.2f}B"
    elif v >= 1_000_000:
        return f"{sign}₦{v/1_000_000:.1f}M"
    elif v >= 1_000:
        return f"{sign}₦{v/1_000:.0f}K"
    else:
        return f"{sign}₦{v:,.0f}"


def safe(row):
    d = {}
    for c in row:
        try:
            if c.column_letter:
                d[c.column_letter] = c.value
        except:
            pass
    return d


def extract_revenue_data(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb['Revenues']
    revenues = []
    current_parent = None
    current_parent_rail = ''

    # Scan to row 80 — sheet grows as deals are added. Break on TOTAL to avoid
    # picking up secondary budget tables that sit below the deals table.
    for row in ws.iter_rows(min_row=3, max_row=80):
        d = safe(row)
        name = str(d.get('B', '') or '').strip()
        if not name:
            continue
        if name.upper() == 'TOTAL':
            break  # Stop — everything below is a different table
        if name.upper() in SECTION_HEADERS:
            current_parent = None
            continue

        sn = sf(d.get('A'))
        has_sn = sn > 0

        if has_sn:
            current_parent = name
            current_parent_rail = str(d.get('C', '') or '').strip()

        rail = str(d.get('C', '') or '').strip() or current_parent_rail or 'Other'
        annual = sf(d.get('E'))
        status = str(d.get('K', '') or '').strip()
        comment = str(d.get('L', '') or '').strip()

        # Read all 12 month columns dynamically (M=Jan .. X=Dec)
        monthly = [sf(d.get(col)) for col in MONTH_COLUMNS]
        ytd = sum(monthly)

        # Only include rows with meaningful data
        if annual == 0 and ytd == 0:
            continue

        # Find the last month index with data across ALL deals (set later globally)
        # For momentum: find the two most recent months with any data for this deal
        non_zero_indices = [i for i, v in enumerate(monthly) if v > 0]
        if len(non_zero_indices) >= 2:
            latest_idx = non_zero_indices[-1]
            prev_idx = non_zero_indices[-2]
            latest = monthly[latest_idx]
            prev = monthly[prev_idx]
        elif len(non_zero_indices) == 1:
            latest_idx = non_zero_indices[0]
            latest = monthly[latest_idx]
            prev = 0
        else:
            latest_idx = -1
            latest = 0
            prev = 0

        # Determine momentum dynamically
        if ytd == 0:
            momentum = 'zero'
        elif len(non_zero_indices) == 1:
            momentum = 'new'
        elif prev > 0 and latest > prev * 1.1:
            momentum = 'growing'
        elif latest == 0 and prev > 0:
            momentum = 'stalled'
        elif len(non_zero_indices) >= 2 and non_zero_indices[-1] < max(non_zero_indices[-1], 0):
            momentum = 'stalled'
        else:
            momentum = 'steady'

        # Check for stalled: had early activity but last 2 months are zero
        all_months_with_data = [i for i in range(12) if monthly[i] > 0]
        if all_months_with_data and ytd > 0:
            latest_data_month = max(all_months_with_data)
            # If the most recent month with data is 2+ months behind the current data frontier
            # (detected globally later), mark as potentially stalled
            pass

        # Is this deal stalled despite On Track label?
        stalled_on_track = (status == 'On Track' and annual > 0 and ytd == 0)

        deal = {
            'sn': sn,
            'parent': current_parent if not has_sn else None,
            'name': name,
            'rail': rail,
            'annual_usd': annual,
            'status': status or 'Unknown',
            'comment': comment,
            'monthly': monthly,  # list of 12 monthly values
            'ytd': ytd,
            'momentum': momentum,
            'stalled_on_track': stalled_on_track,
        }
        # Also store individual months for backward compatibility
        for i, mname in enumerate(MONTH_NAMES):
            deal[mname.lower()] = monthly[i]

        revenues.append(deal)

    wb.close()
    return revenues


def build_recommendations(deal):
    """Generate specific, actionable recommendations per deal based on status and context."""
    name = deal['name']
    comment = deal['comment'].lower()
    annual = deal['annual_usd']
    status = deal['status']
    rail = deal['rail']
    ytd = deal['ytd']

    recs = []

    if status == 'At Risk':
        # Context-specific recommendations from comments
        if 'competitor' in comment or 'competition' in comment:
            recs.append("Conduct competitive analysis and prepare differentiation brief for the client — price alone won't win this.")
        if 'hold' in comment or 'paused' in comment or 'reconfigur' in comment:
            recs.append("Reactivate engagement: schedule a senior-level call this week to understand the blocker and agree a revised timeline.")
        if 'procurement' in comment or 'tender' in comment or 'bid' in comment:
            recs.append("Map procurement contacts and escalate internally — identify who can accelerate the procurement decision.")
        if 'price' in comment or 'pricing' in comment:
            recs.append("Revisit commercial structure. Consider phased pricing, pilot model, or value-based pricing narrative tied to ROI.")
        if 'meeting' in comment or 'president' in comment or 'engagement' in comment:
            recs.append("Convert the scheduled meeting into a concrete agreement — come with a proposal, timeline, and next-step commitments.")
        if 'awaiting' in comment or 'feedback' in comment or 'respond' in comment:
            recs.append("Escalate the follow-up — set a clear deadline with the client and involve a senior executive to accelerate response.")
        if 'interest' in comment or 'license' in comment:
            recs.append("Push for a signed LOI or purchase order to lock commitment before further slippage. Offer implementation support as sweetener.")
        if annual > 400_000:
            recs.append(f"This is a high-value deal ({fmt_usd(annual)}) — assign a dedicated deal champion and weekly executive review.")
        if ytd == 0:
            recs.append("No revenue recorded YTD. Urgently clarify whether the deal is still live or needs to be re-scoped/removed from the pipeline.")
        # Generic fallbacks if no specific context
        if not recs:
            recs.append("Schedule an urgent client review call this week to re-confirm interest and reset delivery expectations.")
            recs.append("Identify the single biggest blocker and assign a named owner with a resolution deadline.")

    elif status == 'Off Track':
        if 'suspended' in comment or 'cash' in comment:
            recs.append("Confirm whether the project is permanently cancelled or suspended — avoid carrying dead pipeline value.")
            recs.append("Explore if there's a minimal viable scope that can proceed within their cash constraints.")
        if 'infrastructure' in comment or 'pending' in comment:
            recs.append("Identify what Seamfix can control (e.g., documentation, staging prep) and move those forward while awaiting the blocker.")
        if not recs:
            recs.append("Assess whether this deal is recoverable in 2026 or should be moved to 2027 pipeline to improve forecast accuracy.")
            recs.append("Get written confirmation from the client on status — internal assumptions should not drive pipeline numbers.")

    return recs


def generate_html(revenues, output_path):
    generated_at = datetime.now().strftime('%d %b %Y, %H:%M')
    today = datetime.now().strftime('%d %b %Y')

    # ── LANDING ZONE: derive from live data, not a hardcoded constant ───
    # Uses the sum of all deal annual targets from the revenue file, so it
    # stays in sync with generate_revenue_dashboard.py. Falls back to the
    # module-level constant if the data is empty (shouldn't happen in prod).
    # LANDING_ZONE is the official $8M company target, not the $10M deal bucket sum
    deal_bucket_total = sum(r['annual_usd'] for r in revenues)

    # ── STATUS BUCKETS ──────────────────────────────────────────────
    buckets = {}
    for r in revenues:
        s = r['status']
        if s not in buckets:
            buckets[s] = {'count': 0, 'annual': 0, 'deals': []}
        buckets[s]['count'] += 1
        buckets[s]['annual'] += r['annual_usd']
        buckets[s]['deals'].append(r)

    on_track_count  = buckets.get('On Track',  {}).get('count',  0)
    on_track_val    = buckets.get('On Track',  {}).get('annual', 0)
    at_risk_count   = buckets.get('At Risk',   {}).get('count',  0)
    at_risk_val     = buckets.get('At Risk',   {}).get('annual', 0)
    off_track_count = buckets.get('Off Track', {}).get('count',  0)
    off_track_val   = buckets.get('Off Track', {}).get('annual', 0)
    closed_count    = buckets.get('Closed',    {}).get('count',  0)
    closed_val      = buckets.get('Closed',    {}).get('annual', 0)
    unknown_count   = buckets.get('Unknown',   {}).get('count',  0)
    unknown_val     = buckets.get('Unknown',   {}).get('annual', 0)

    # ── PROJECTIONS ─────────────────────────────────────────────────
    realistic_proj = sum(r['annual_usd'] * STATUS_WEIGHTS.get(r['status'], 0.5) for r in revenues)
    conservative_proj = sum(r['annual_usd'] * STATUS_WEIGHTS_CONSERVATIVE.get(r['status'], 0) for r in revenues)

    realistic_gap     = LANDING_ZONE - realistic_proj
    conservative_gap  = LANDING_ZONE - conservative_proj
    realistic_pct     = realistic_proj / LANDING_ZONE * 100
    conservative_pct  = conservative_proj / LANDING_ZONE * 100

    # ── DYNAMIC MONTH DETECTION ──────────────────────────────────────
    # Sum each month across all deals to find which months have data
    monthly_totals = [sum(r['monthly'][i] for r in revenues) for i in range(12)]
    # Find the last month with any data
    months_with_data = [i for i, v in enumerate(monthly_totals) if v > 0]
    last_data_month = max(months_with_data) if months_with_data else 0  # 0-indexed (0=Jan)
    num_months_with_data = last_data_month + 1  # count from Jan

    ytd_total = sum(monthly_totals[:num_months_with_data])

    # Run rate: use complete months (all except the last which may be partial)
    today = datetime.now()
    current_month_idx = today.month - 1  # 0-indexed
    # If the last data month is the current month, treat it as partial
    is_partial = (last_data_month == current_month_idx)
    complete_month_count = last_data_month if is_partial else num_months_with_data
    complete_months_total = sum(monthly_totals[:complete_month_count]) if complete_month_count > 0 else 0
    complete_months_avg = complete_months_total / complete_month_count if complete_month_count > 0 else 0

    if is_partial and monthly_totals[last_data_month] > 0:
        days_in_month = 30
        days_elapsed = min(today.day, days_in_month)
        partial_scaled = monthly_totals[last_data_month] * (days_in_month / max(days_elapsed, 1))
        monthly_run_rate = max(complete_months_avg, (complete_months_total + partial_scaled) / (complete_month_count + 1))
    else:
        monthly_run_rate = complete_months_avg

    annual_run_rate = monthly_run_rate * 12
    run_rate_pct = annual_run_rate / LANDING_ZONE * 100

    # Dynamic label
    first_month_name = MONTH_NAMES[0]
    last_month_name = MONTH_NAMES[last_data_month]
    partial_note = f' ({last_month_name} partial)' if is_partial else ''
    ytd_label = f'{first_month_name} – {last_month_name} 2026{partial_note}'

    # ── STALLED ON TRACK ────────────────────────────────────────────
    stalled = [r for r in revenues if r['stalled_on_track']]
    stalled_value = sum(r['annual_usd'] for r in stalled)

    # ── MOVERS vs NON-MOVERS (deals with any actuals) ───────────────
    movers    = [r for r in revenues if r['ytd'] > 0]
    non_movers = [r for r in revenues if r['ytd'] == 0 and r['annual_usd'] > 0 and r['status'] in ('On Track',)]

    # ── AT RISK DEALS ───────────────────────────────────────────────
    at_risk_deals  = sorted(buckets.get('At Risk',  {}).get('deals', []), key=lambda x: -x['annual_usd'])
    off_track_deals = sorted(buckets.get('Off Track',{}).get('deals', []), key=lambda x: -x['annual_usd'])

    # ── CHART DATA ──────────────────────────────────────────────────
    # Monthly momentum for top movers (dynamic — only months with data)
    top_movers = sorted(movers, key=lambda x: -x['ytd'])[:10]
    momentum_labels = [r['name'][:25] + ('…' if len(r['name']) > 25 else '') for r in top_movers]
    # Build per-month data arrays for each month that has data
    momentum_month_data = []
    for mi in range(num_months_with_data):
        values = [r['monthly'][mi] for r in top_movers]
        label = MONTH_NAMES[mi]
        if mi == last_data_month and is_partial:
            label += ' (partial)'
        momentum_month_data.append({
            'label': label,
            'data': values,
            'color': MONTH_COLORS[mi % len(MONTH_COLORS)],
        })

    # Projection bar data
    proj_labels = ['Annual Target', 'Realistic\nProjection', 'Conservative\n(If Risk Stays)', 'Run Rate\n(Annualised)']
    proj_values = [LANDING_ZONE, realistic_proj, conservative_proj, annual_run_rate]
    proj_colors = ['#3b82f6', '#10b981', '#f59e0b', '#8b5cf6']

    # Status donut (by value)
    donut_labels = []
    donut_values = []
    donut_colors_map = {'On Track': '#10b981', 'At Risk': '#f59e0b', 'Off Track': '#ef4444', 'Closed': '#3b82f6', 'Unknown': '#6b7280'}
    for status, color in donut_colors_map.items():
        val = buckets.get(status, {}).get('annual', 0)
        if val > 0:
            donut_labels.append(status)
            donut_values.append(val)

    # ── CLOSED & UNKNOWN DEAL LISTS ─────────────────────────────────
    closed_deals  = sorted(buckets.get('Closed',  {}).get('deals', []), key=lambda x: -x['annual_usd'])
    unknown_deals = sorted(buckets.get('Unknown', {}).get('deals', []), key=lambda x: -x['annual_usd'])

    # ── HTML ────────────────────────────────────────────────────────
    stalled_rows = ""
    for r in stalled:
        comment_short = r['comment'][:250] + '…' if len(r['comment']) > 250 else r['comment']
        stalled_rows += f"""
        <tr>
            <td>{r['name']}</td>
            <td>{r['rail']}</td>
            <td class="amount">{fmt_usd(r['annual_usd'])}</td>
            <td style="color:var(--text-secondary);font-style:italic;font-size:12px;max-width:400px;white-space:normal;line-height:1.4">{comment_short or '—'}</td>
        </tr>"""

    closed_rows = ""
    for r in closed_deals:
        comment_short = r['comment'][:250] + '…' if len(r['comment']) > 250 else r['comment']
        ytd_str = fmt_usd(r['ytd']) if r['ytd'] > 0 else '—'
        closed_rows += f"""
        <tr>
            <td>{r['name']}</td>
            <td>{r['rail']}</td>
            <td class="amount">{fmt_usd(r['annual_usd'])}</td>
            <td class="amount" style="color:#10b981">{ytd_str}</td>
            <td style="color:var(--text-secondary);font-style:italic;font-size:12px;max-width:400px;white-space:normal;line-height:1.4">{comment_short or '—'}</td>
        </tr>"""

    unknown_rows = ""
    for r in unknown_deals:
        comment_short = r['comment'][:250] + '…' if len(r['comment']) > 250 else r['comment']
        ytd_str = fmt_usd(r['ytd']) if r['ytd'] > 0 else '—'
        unknown_rows += f"""
        <tr>
            <td>{r['name']}</td>
            <td>{r['rail']}</td>
            <td class="amount">{fmt_usd(r['annual_usd'])}</td>
            <td class="amount" style="color:var(--text-secondary)">{ytd_str}</td>
            <td style="color:var(--text-secondary);font-style:italic;font-size:12px;max-width:400px;white-space:normal;line-height:1.4">{comment_short or '—'}</td>
        </tr>"""

    movers_rows = ""
    for r in sorted(movers, key=lambda x: -x['ytd']):
        monthly = r['monthly']
        status = r['status']

        # ── TREND LABEL (dynamic — uses most recent months with data) ────────
        non_zero = [i for i, v in enumerate(monthly) if v > 0]
        if len(non_zero) == 0:
            trend_key = 'mixed'
            trend = '<span style="color:var(--text-secondary)">– Mixed</span>'
        elif len(non_zero) == 1:
            trend_key = 'new'
            trend = '<span style="color:#3b82f6">✦ New</span>'
        else:
            latest_val = monthly[non_zero[-1]]
            prev_val = monthly[non_zero[-2]]
            latest_mi = non_zero[-1]
            # Check if the last month with data is NOT the most recent data month globally
            # (i.e. deal stopped producing revenue before others)
            if latest_mi < last_data_month - 1:
                trend_key = 'dropped'
                trend = f'<span style="color:#ef4444">▼ Dropped (last: {MONTH_NAMES[latest_mi]})</span>'
            elif latest_mi < last_data_month and monthly[last_data_month] == 0:
                trend_key = 'stalled'
                trend = f'<span style="color:#f59e0b">⚠ Stalled {MONTH_NAMES[last_data_month]}</span>'
            elif latest_val > prev_val * 1.05:
                trend_key = 'growing'
                trend = '<span style="color:#10b981">▲ Growing</span>'
            elif len(non_zero) >= 3:
                trend_key = 'consistent'
                trend = '<span style="color:#10b981">✓ Consistent</span>'
            elif len(non_zero) == 2 and non_zero[1] - non_zero[0] <= 1:
                trend_key = 'consistent'
                trend = '<span style="color:#10b981">✓ Consistent</span>'
            else:
                trend_key = 'mixed'
                trend = '<span style="color:var(--text-secondary)">– Mixed</span>'

        # ── INSIGHT NOTE ──────────────────────────────────────────────
        # Always produce a note: divergence cases get a strong flag,
        # aligned cases get Claude's rationale for what the data shows.
        excel_comment = r.get('comment', '').strip()
        contrast = ''

        # ── DIVERGENCE: status and trend contradict ──────────────────
        if status == 'At Risk' and trend_key == 'growing':
            contrast = 'Revenue is growing month-on-month, but flagged At Risk — contract may not yet be formalised. Verify whether the risk label needs updating.'
        elif status == 'At Risk' and trend_key == 'consistent':
            contrast = 'Revenue is flowing consistently despite At Risk label. Confirm whether the risk has been resolved and status needs updating.'
        elif status == 'On Track' and trend_key == 'dropped':
            contrast = 'Status says On Track but revenue has stopped. Verify deal health with the account manager — status may be stale.'
        elif status == 'On Track' and trend_key == 'stalled_apr':
            contrast = 'Had revenue through March but nothing recorded in April yet. Could be timing; monitor closely.'
        elif status == 'Closed' and trend_key == 'new':
            contrast = 'Deal is closed/secured and revenue just started recently — expected pattern for a recently closed deal.'
        elif status == 'Closed' and trend_key == 'mixed':
            contrast = 'Deal is closed but monthly revenue is irregular. Check invoice schedule or delivery milestones.'
        elif status == 'Closed' and trend_key == 'dropped':
            contrast = 'Closed deal had revenue earlier but has since stopped. Verify whether all invoices have been settled.'
        elif status == 'Off Track' and trend_key in ('growing', 'consistent'):
            contrast = 'Revenue is actually flowing despite Off Track label. Status may need to be revised upward.'

        # ── ALIGNED: status and trend are consistent — explain what we see ──
        elif status == 'At Risk' and trend_key == 'stalled_apr':
            contrast = 'At Risk status is consistent with the data — revenue stopped in April after some activity in March. Needs urgent follow-up.'
        elif status == 'At Risk' and trend_key == 'dropped':
            contrast = 'At Risk status confirmed by revenue data — revenue recorded earlier has since stopped entirely. Escalate or reclassify.'
        elif status == 'At Risk' and trend_key == 'new':
            contrast = 'At Risk but revenue just appeared — early signal that the deal may be activating. Monitor whether it sustains.'
        elif status == 'At Risk' and trend_key == 'mixed':
            contrast = 'At Risk with irregular revenue pattern. Deal health is unclear — needs direct client contact to clarify outlook.'
        elif status == 'On Track' and trend_key == 'growing':
            contrast = 'Strong alignment — On Track status confirmed by growing revenue. Keep momentum.'
        elif status == 'On Track' and trend_key == 'consistent':
            contrast = 'Healthy pattern — On Track status backed by steady monthly revenue. No action needed.'
        elif status == 'On Track' and trend_key == 'new':
            contrast = 'On Track and revenue just started flowing — early but positive. Confirm invoicing cadence is established.'
        elif status == 'On Track' and trend_key == 'mixed':
            contrast = 'On Track but revenue pattern is uneven. Check if delivery or invoicing timing explains the irregularity.'
        elif status == 'Closed' and trend_key == 'consistent':
            contrast = 'Closed and delivering steadily — performing as expected.'
        elif status == 'Closed' and trend_key == 'growing':
            contrast = 'Closed and revenue is increasing — may indicate expanded scope or usage-based growth.'
        elif status == 'Off Track' and trend_key == 'dropped':
            contrast = 'Off Track confirmed — revenue has stopped. Decide whether to invest in recovery or write off.'
        elif status == 'Off Track' and trend_key == 'stalled_apr':
            contrast = 'Off Track and stalled in April — deal appears to be winding down.'
        elif status == 'Off Track' and trend_key == 'mixed':
            contrast = 'Off Track with sporadic revenue. Unclear if the deal has residual value — clarify with the client.'
        elif status == 'Unknown':
            if r['ytd'] > 0:
                contrast = f'No status assigned but ${r["ytd"]:,.0f} collected YTD. Revenue team should classify this deal immediately.'
            else:
                contrast = 'No status and no revenue. Confirm whether this deal is still in the pipeline.'

        # ── FALLBACK: generic note if none of the above matched ──────
        if not contrast:
            if r['ytd'] == 0:
                contrast = f'No revenue recorded YTD against {fmt_usd(r["annual_usd"])} annual target. Confirm deal is still active.'
            else:
                pct = (r["ytd"]/r["annual_usd"]*100) if r["annual_usd"] > 0 else 0
                contrast = f'YTD revenue of {fmt_usd(r["ytd"])} against {fmt_usd(r["annual_usd"])} annual target ({pct:.0f}% collected).'

        # Append Excel comment if present (user context from the sheet)
        if excel_comment:
            contrast += f' <span style="color:var(--text-tertiary)">[Sheet note: {excel_comment}]</span>'

        comment_cell = f'<td style="font-size:12px;color:var(--text-secondary);font-style:italic;max-width:280px">{contrast}</td>'

        # ── ROW BACKGROUND by status ──────────────────────────────────
        row_bg = {
            'On Track':  'background:rgba(59,130,246,0.06)',
            'At Risk':   'background:rgba(245,158,11,0.07)',
            'Off Track': 'background:rgba(239,68,68,0.07)',
            'Closed':    'background:rgba(16,185,129,0.07)',
        }.get(status, '')

        # ── BADGE COLOURS ─────────────────────────────────────────────
        status_badge_color = {
            'On Track':  '#3b82f6',
            'At Risk':   '#f59e0b',
            'Off Track': '#ef4444',
            'Closed':    '#10b981',
        }.get(status, '#6b7280')

        # Build dynamic month cells
        month_cells = ''
        for mi in range(num_months_with_data):
            val = monthly[mi]
            partial_tag = f' <span style="font-size:10px">(partial)</span>' if (mi == last_data_month and is_partial) else ''
            style = ' style="color:var(--text-secondary)"' if (mi == last_data_month and is_partial) else ''
            month_cells += f'<td class="amount"{style}>{fmt_usd(val) if val else "–"}{partial_tag}</td>'

        movers_rows += f"""
        <tr style="{row_bg}">
            <td style="max-width:160px;white-space:normal;word-break:break-word">{r['name']}</td>
            <td><span class="badge" style="background:{status_badge_color}20;color:{status_badge_color};border:1px solid {status_badge_color}40">{status}</span></td>
            {month_cells}
            <td class="amount"><strong>{fmt_usd(r['ytd'])}</strong></td>
            <td>{trend}</td>
            {comment_cell}
        </tr>"""

    at_risk_cards = ""
    for r in at_risk_deals:
        recs = build_recommendations(r)
        recs_html = ''.join(f'<li>{rec}</li>' for rec in recs)
        comment_display = r['comment'] if r['comment'] else 'No notes recorded.'
        ytd_display = fmt_usd(r['ytd']) if r['ytd'] > 0 else '<span style="color:#ef4444">$0 — No YTD revenue</span>'
        at_risk_cards += f"""
        <div class="risk-card">
            <div class="risk-card-header">
                <div>
                    <div class="risk-deal-name">{r['name']}</div>
                    <div class="risk-deal-meta">{r['rail']} &nbsp;·&nbsp; Annual value: <strong>{fmt_usd(r['annual_usd'])}</strong> ({fmt_naira(r['annual_usd'] * FX_RATE)})</div>
                </div>
                <div class="risk-amount">{fmt_usd(r['annual_usd'])}</div>
            </div>
            <div class="risk-card-body">
                <div class="risk-section">
                    <div class="risk-section-title">📋 Current Status Note</div>
                    <div class="risk-comment">{comment_display}</div>
                </div>
                <div class="risk-section">
                    <div class="risk-section-title">📊 YTD Revenue Recorded</div>
                    <div style="margin-top:6px">{ytd_display}</div>
                    <div style="display:flex;gap:16px;margin-top:8px;flex-wrap:wrap">
                        {''.join(f'<span>{MONTH_NAMES[mi]}: {fmt_usd(r["monthly"][mi]) if r["monthly"][mi] else "–"}{" <em style=font-size:10px>(partial)</em>" if mi == last_data_month and is_partial else ""}</span>' for mi in range(num_months_with_data))}
                    </div>
                </div>
                <div class="risk-section">
                    <div class="risk-section-title">🎯 Recommended Actions</div>
                    <ul class="rec-list">{recs_html}</ul>
                </div>
            </div>
        </div>"""

    off_track_cards = ""
    for r in off_track_deals:
        recs = build_recommendations(r)
        recs_html = ''.join(f'<li>{rec}</li>' for rec in recs)
        comment_display = r['comment'] if r['comment'] else 'No notes recorded.'
        off_track_cards += f"""
        <div class="risk-card" style="border-left-color:#ef4444">
            <div class="risk-card-header">
                <div>
                    <div class="risk-deal-name">{r['name']}</div>
                    <div class="risk-deal-meta">{r['rail']} &nbsp;·&nbsp; Annual value: <strong>{fmt_usd(r['annual_usd'])}</strong></div>
                </div>
                <div class="risk-amount" style="color:#ef4444">{fmt_usd(r['annual_usd'])}</div>
            </div>
            <div class="risk-card-body">
                <div class="risk-section">
                    <div class="risk-section-title">📋 Status Note</div>
                    <div class="risk-comment">{comment_display}</div>
                </div>
                <div class="risk-section">
                    <div class="risk-section-title">🎯 Recommended Actions</div>
                    <ul class="rec-list">{recs_html}</ul>
                </div>
            </div>
        </div>"""

    # Landing zone bar widths (capped at 100%)
    realistic_bar_w   = min(100, realistic_pct)
    conservative_bar_w = min(100, conservative_pct)
    run_rate_bar_w    = min(100, run_rate_pct)

    # Customer concentration: aggregate annual value per deal name
    client_values = {}
    for r in revenues:
        if r['annual_usd'] > 0:
            client_values[r['name']] = client_values.get(r['name'], 0) + r['annual_usd']
    total_pipeline = sum(client_values.values())
    top_clients = sorted(client_values.items(), key=lambda x: -x[1])[:5]

    # Build concentration rows HTML
    conc_rows_html = ""
    for client_name, val in top_clients:
        pct = (val / total_pipeline * 100) if total_pipeline > 0 else 0
        bar_color = '#ef4444' if pct > 25 else '#f59e0b' if pct > 15 else '#3b82f6'
        risk_label = '<span style="color:#ef4444;font-size:11px;font-weight:700">HIGH</span>' if pct > 25 else '<span style="color:#f59e0b;font-size:11px;font-weight:700">MEDIUM</span>' if pct > 15 else '<span style="color:var(--text-secondary);font-size:11px">LOW</span>'
        conc_rows_html += f"""
    <div style="margin-bottom:14px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:5px">
        <span style="font-size:13px;font-weight:500;color:var(--text-primary)">{client_name}</span>
        <span style="display:flex;align-items:center;gap:10px">
          {risk_label}
          <span style="font-size:13px;color:var(--text-secondary)">{fmt_usd(val)}</span>
          <span style="font-size:13px;font-weight:700;color:{bar_color}">{pct:.1f}%</span>
        </span>
      </div>
      <div style="background:var(--border-main);border-radius:4px;height:8px;overflow:hidden">
        <div style="width:{min(100,pct):.1f}%;height:100%;background:{bar_color};border-radius:4px"></div>
      </div>
    </div>"""

    theme_css = get_base_css()
    toggle_html = get_toggle_html()
    theme_js = get_theme_js()

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Seamfix — Pipeline Intelligence Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  {theme_css}
  :root {{
    --bg:        var(--bg-body);
    --surface:   var(--bg-card);
    --surface2:  var(--bg-table-header);
    --border:    var(--border-main);
    --text:      var(--text-primary);
    --muted:     var(--text-secondary);
    --green:     #10b981;
    --yellow:    #f59e0b;
    --red:       #ef4444;
    --blue:      #3b82f6;
    --purple:    #8b5cf6;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: var(--bg); color: var(--text); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size: 15px; line-height: 1.5; }}

  /* NAV */
  .nav {{ background: var(--bg-nav); border-bottom: 1px solid var(--border); padding: 0 24px; display: flex; align-items: center; gap: 0; height: 48px; overflow-x: auto; position: sticky; top: 0; z-index: 200; }}
  .nav-brand {{ color: var(--text); font-weight: 700; font-size: 15px; margin-right: 24px; white-space: nowrap; }}
  .nav-link {{ color: var(--muted); text-decoration: none; padding: 0 14px; height: 48px; display: flex; align-items: center; font-size: 13px; border-bottom: 2px solid transparent; white-space: nowrap; transition: color .2s; }}
  .nav-link:hover {{ color: var(--text); }}
  .nav-link.active {{ color: #fff; border-bottom-color: var(--blue); font-weight: 500; }}

  /* HEADER */
  .page-header {{ padding: 28px 28px 0; max-width: 1600px; margin: 0 auto; }}
  .page-title {{ font-size: 22px; font-weight: 700; }}
  .page-sub {{ color: var(--muted); font-size: 13px; margin-top: 4px; }}
  .generated {{ color: var(--muted); font-size: 12px; margin-top: 4px; }}

  /* SECTION */
  .section {{ padding: 24px 28px 0; max-width: 1600px; margin: 0 auto; }}
  .section-title {{ font-size: 13px; font-weight: 700; text-transform: uppercase; letter-spacing: .08em; color: var(--muted); margin-bottom: 14px; display: flex; align-items: center; gap: 8px; }}
  .section-title::after {{ content:''; flex:1; height:1px; background: var(--border); }}

  /* STATUS CARDS */
  .status-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }}
  @media (max-width: 900px) {{ .status-grid {{ grid-template-columns: repeat(2, 1fr); }} }}
  .status-card {{ background: var(--surface); border-radius: 12px; padding: 20px; border: 1px solid var(--border); position: relative; overflow: hidden; }}
  .status-card::before {{ content:''; position:absolute; top:0; left:0; right:0; height:3px; }}
  .status-card.on-track::before  {{ background: var(--green);  }}
  .status-card.at-risk::before   {{ background: var(--yellow); }}
  .status-card.off-track::before {{ background: var(--red);    }}
  .status-card.closed::before    {{ background: var(--blue);   }}
  .status-label {{ font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: .08em; margin-bottom: 10px; }}
  .status-card.on-track  .status-label {{ color: var(--green);  }}
  .status-card.at-risk   .status-label {{ color: var(--yellow); }}
  .status-card.off-track .status-label {{ color: var(--red);    }}
  .status-card.closed    .status-label {{ color: var(--blue);   }}
  .status-value {{ font-size: 28px; font-weight: 800; line-height: 1; margin-bottom: 4px; }}
  .status-count {{ font-size: 12px; color: var(--muted); }}
  .status-naira {{ font-size: 11px; color: var(--muted); margin-top: 6px; }}

  /* PROJECTION / LANDING ZONE */
  .lz-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
  @media (max-width: 900px) {{ .lz-grid {{ grid-template-columns: 1fr; }} }}
  .lz-card {{ background: var(--surface); border-radius: 12px; padding: 24px; border: 1px solid var(--border); }}
  .lz-title {{ font-size: 13px; font-weight: 600; color: var(--muted); margin-bottom: 18px; }}
  .lz-scenario {{ margin-bottom: 18px; }}
  .lz-scenario-header {{ display: flex; justify-content: space-between; margin-bottom: 6px; font-size: 13px; }}
  .lz-scenario-label {{ color: var(--text); font-weight: 500; }}
  .lz-scenario-value {{ font-weight: 700; }}
  .lz-bar-track {{ background: var(--border); border-radius: 4px; height: 10px; overflow: hidden; }}
  .lz-bar-fill {{ height: 100%; border-radius: 4px; transition: width .8s ease; }}
  .lz-scenario-sub {{ font-size: 11px; color: var(--muted); margin-top: 4px; }}
  .lz-insight {{ background: var(--surface2); border-radius: 8px; padding: 14px; margin-top: 4px; font-size: 13px; line-height: 1.6; color: var(--text); }}
  .lz-insight strong {{ color: var(--yellow); }}

  /* STALLED ALERT */
  .alert-box {{ background: var(--warning-bg); border: 1px solid var(--warning); border-radius: 12px; padding: 20px; }}
  .alert-box-title {{ color: var(--warning-bright); font-weight: 700; font-size: 14px; margin-bottom: 4px; }}
  .alert-box-sub {{ color: var(--text-secondary); font-size: 12px; margin-bottom: 16px; }}

  /* TABLES */
  .table-wrap {{ background: var(--surface); border-radius: 12px; border: 1px solid var(--border); overflow: hidden; }}
  table {{ width: 100%; border-collapse: collapse; }}
  thead th {{ background: var(--bg-table-header); padding: 11px 14px; text-align: left; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; color: var(--muted); border-bottom: 1px solid var(--border); }}
  tbody tr {{ border-bottom: 1px solid var(--border-light); }}
  tbody tr:last-child {{ border-bottom: none; }}
  tbody tr:hover {{ background: var(--bg-table-hover); }}
  tbody td {{ padding: 11px 14px; font-size: 13px; vertical-align: middle; }}
  td.amount {{ text-align: right; font-variant-numeric: tabular-nums; }}

  /* BADGE */
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: 11px; font-weight: 600; }}

  /* CHARTS */
  .chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
  @media (max-width: 900px) {{ .chart-grid {{ grid-template-columns: 1fr; }} }}
  .chart-card {{ background: var(--surface); border-radius: 12px; padding: 20px; border: 1px solid var(--border); }}
  .chart-title {{ font-size: 13px; font-weight: 600; color: var(--muted); margin-bottom: 16px; }}

  /* RISK CARDS */
  .risk-card {{ background: var(--surface); border-radius: 12px; border: 1px solid var(--border); border-left: 4px solid var(--yellow); margin-bottom: 16px; overflow: hidden; }}
  .risk-card-header {{ background: var(--surface2); padding: 16px 20px; display: flex; justify-content: space-between; align-items: flex-start; gap: 12px; }}
  .risk-deal-name {{ font-size: 15px; font-weight: 700; }}
  .risk-deal-meta {{ font-size: 12px; color: var(--muted); margin-top: 3px; }}
  .risk-amount {{ font-size: 22px; font-weight: 800; color: var(--yellow); white-space: nowrap; }}
  .risk-card-body {{ padding: 16px 20px; display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; }}
  @media (max-width: 900px) {{ .risk-card-body {{ grid-template-columns: 1fr; }} }}
  .risk-section-title {{ font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: .08em; color: var(--muted); margin-bottom: 8px; }}
  .risk-comment {{ font-size: 12px; color: var(--text); line-height: 1.6; background: var(--surface2); padding: 10px; border-radius: 6px; }}
  .rec-list {{ list-style: none; padding: 0; }}
  .rec-list li {{ font-size: 12px; color: var(--text); padding: 6px 0 6px 18px; position: relative; line-height: 1.5; border-bottom: 1px solid var(--border); }}
  .rec-list li:last-child {{ border-bottom: none; }}
  .rec-list li::before {{ content: '→'; position: absolute; left: 0; color: var(--yellow); font-weight: 700; }}

  /* FOOTER */
  .footer {{ padding: 20px 28px; color: var(--text-secondary); font-size: 12px; border-top: 1px solid var(--border); margin-top: 32px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 8px; max-width: 1600px; margin-left: auto; margin-right: auto; }}

  /* SUMMARY BAR */
  .summary-bar {{ background: var(--surface2); border-radius: 12px; padding: 16px 20px; display: flex; gap: 32px; flex-wrap: wrap; margin-bottom: 0; border: 1px solid var(--border); }}
  .summary-item {{ display: flex; flex-direction: column; gap: 2px; }}
  .summary-item-label {{ font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: .06em; }}
  .summary-item-value {{ font-size: 18px; font-weight: 700; }}
</style>
</head>
<body>
{toggle_html}

<!-- NAV -->
<nav class="nav">
  <div class="nav-brand">⚡ Seamfix</div>
  <a href="dashboard.html"         class="nav-link">Cash Overview</a>
  <a href="expense_dashboard.html" class="nav-link">Expense &amp; Vendor</a>
  <a href="budget_dashboard.html"  class="nav-link">Budget vs Actual</a>
  <a href="revenue_dashboard.html" class="nav-link">Revenue &amp; Fundability</a>
  <a href="pipeline_dashboard.html" class="nav-link active">Pipeline Intelligence</a>
</nav>

<!-- HEADER -->
<div class="page-header">
  <div class="page-title">Pipeline Intelligence</div>
  <div class="page-sub">Deal-level momentum tracking, landing zone analysis, and action recommendations</div>
  <div class="generated" style="margin-top:6px">Data as of: <strong style="color:var(--text-primary)">{ytd_label}</strong> &nbsp;·&nbsp; Generated: {generated_at}</div>
</div>

<!-- QUICK SUMMARY BAR -->
<div class="section">
  <div class="summary-bar">
    <div class="summary-item">
      <span class="summary-item-label">Annual Target</span>
      <span class="summary-item-value">{fmt_usd(LANDING_ZONE)}</span>
    </div>
    <div class="summary-item">
      <span class="summary-item-label">YTD Actual ({ytd_label})</span>
      <span class="summary-item-value" style="color:var(--green)">{fmt_usd(ytd_total)}</span>
    </div>
    <div class="summary-item">
      <span class="summary-item-label">Realistic Projection</span>
      <span class="summary-item-value" style="color:{'var(--green)' if realistic_pct >= 90 else 'var(--yellow)'}">{fmt_usd(realistic_proj)}</span>
    </div>
    <div class="summary-item">
      <span class="summary-item-label">{'Surplus over Target' if realistic_gap <= 0 else 'Gap to Target'}</span>
      <span class="summary-item-value" style="color:{'var(--green)' if realistic_gap <= 0 else 'var(--yellow)'}">{fmt_usd(abs(realistic_gap))}</span>
    </div>
    <div class="summary-item">
      <span class="summary-item-label">Stalled On-Track Deals</span>
      <span class="summary-item-value" style="color:var(--yellow)">{len(stalled)} deals / {fmt_usd(stalled_value)}</span>
    </div>
    <div class="summary-item">
      <span class="summary-item-label">YTD Annualised Run Rate</span>
      <span class="summary-item-value" style="color:var(--purple)">{fmt_usd(annual_run_rate)}</span>
    </div>
  </div>
</div>

<!-- STATUS CARDS -->
<div class="section">
  <div class="section-title">Pipeline Status Breakdown</div>
  <div class="status-grid">
    <div class="status-card on-track">
      <div class="status-label">✅ On Track</div>
      <div class="status-value" style="color:var(--green)">{fmt_usd(on_track_val, 1)}</div>
      <div class="status-count">{on_track_count} deals &nbsp;·&nbsp; {on_track_val/LANDING_ZONE*100:.0f}% of target</div>
      <div class="status-naira">≈ {fmt_naira(on_track_val * FX_RATE)} at ₦{FX_RATE:,}/$</div>
    </div>
    <div class="status-card at-risk">
      <div class="status-label">⚠️ At Risk</div>
      <div class="status-value" style="color:var(--yellow)">{fmt_usd(at_risk_val, 1)}</div>
      <div class="status-count">{at_risk_count} deals &nbsp;·&nbsp; {at_risk_val/LANDING_ZONE*100:.0f}% of target</div>
      <div class="status-naira">≈ {fmt_naira(at_risk_val * FX_RATE)} at ₦{FX_RATE:,}/$</div>
    </div>
    <div class="status-card off-track">
      <div class="status-label">🔴 Off Track</div>
      <div class="status-value" style="color:var(--red)">{fmt_usd(off_track_val, 1)}</div>
      <div class="status-count">{off_track_count} deals &nbsp;·&nbsp; {off_track_val/LANDING_ZONE*100:.0f}% of target</div>
      <div class="status-naira">≈ {fmt_naira(off_track_val * FX_RATE)} at ₦{FX_RATE:,}/$</div>
    </div>
    <div class="status-card closed">
      <div class="status-label">🔒 Closed / Secured</div>
      <div class="status-value" style="color:var(--blue)">{fmt_usd(closed_val, 1)}</div>
      <div class="status-count">{closed_count} deals &nbsp;·&nbsp; {closed_val/LANDING_ZONE*100:.0f}% of target</div>
      <div class="status-naira">≈ {fmt_naira(closed_val * FX_RATE)} at ₦{FX_RATE:,}/$</div>
    </div>
  </div>
</div>

<!-- CLIENT CONCENTRATION -->
<div class="section">
  <div class="section-title">🎯 Client Concentration Risk</div>
  <div style="background:var(--surface);border-radius:12px;border:1px solid var(--border);padding:24px">
    <div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:18px">
      <span style="font-size:13px;color:var(--muted)">Top 5 clients by pipeline value — total pipeline: <strong style="color:var(--text)">{fmt_usd(total_pipeline)}</strong></span>
      <span style="font-size:11px;color:var(--muted)">Risk: &gt;25% = HIGH &nbsp;·&nbsp; 15–25% = MEDIUM &nbsp;·&nbsp; &lt;15% = LOW</span>
    </div>
    {conc_rows_html}
  </div>
</div>

<!-- LANDING ZONE -->
<div class="section">
  <div class="section-title">Landing Zone Analysis</div>
  <div class="lz-grid">
    <div class="lz-card">
      <div class="lz-title">Revenue Projections vs ${LANDING_ZONE/1_000_000:.0f}M Landing Zone</div>

      <!-- Annual target reference line -->
      <div class="lz-scenario" style="margin-bottom:22px">
        <div class="lz-scenario-header">
          <span class="lz-scenario-label" style="color:var(--blue)">🎯 Annual Target (Landing Zone)</span>
          <span class="lz-scenario-value" style="color:var(--blue)">{fmt_usd(LANDING_ZONE)}</span>
        </div>
        <div class="lz-bar-track"><div class="lz-bar-fill" style="width:100%;background:var(--blue)"></div></div>
      </div>

      <div class="lz-scenario">
        <div class="lz-scenario-header">
          <span class="lz-scenario-label">Realistic (50% of At-Risk closes)</span>
          <span class="lz-scenario-value" style="color:var(--green)">{fmt_usd(realistic_proj)} &nbsp;<span style="font-weight:400;font-size:12px;color:var(--muted)">{realistic_pct:.0f}%</span></span>
        </div>
        <div class="lz-bar-track"><div class="lz-bar-fill" style="width:{realistic_bar_w:.1f}%;background:var(--green)"></div></div>
        <div class="lz-scenario-sub">{'Surplus: ' if realistic_gap <= 0 else 'Gap: '}<strong style="color:{'var(--green)' if realistic_gap <= 0 else 'var(--yellow)'}">{fmt_usd(abs(realistic_gap))}</strong></div>
      </div>

      <div class="lz-scenario" style="margin-top:16px">
        <div class="lz-scenario-header">
          <span class="lz-scenario-label">Conservative (At-Risk &amp; Off-Track unresolved)</span>
          <span class="lz-scenario-value" style="color:var(--red)">{fmt_usd(conservative_proj)} &nbsp;<span style="font-weight:400;font-size:12px;color:var(--muted)">{conservative_pct:.0f}%</span></span>
        </div>
        <div class="lz-bar-track"><div class="lz-bar-fill" style="width:{conservative_bar_w:.1f}%;background:var(--red)"></div></div>
        <div class="lz-scenario-sub">Gap to landing zone: <strong style="color:var(--red)">{fmt_usd(conservative_gap)}</strong></div>
      </div>

      <div class="lz-scenario" style="margin-top:16px">
        <div class="lz-scenario-header">
          <span class="lz-scenario-label" style="color:var(--purple)">Run Rate (Annualised)</span>
          <span class="lz-scenario-value" style="color:var(--purple)">{fmt_usd(annual_run_rate)} &nbsp;<span style="font-weight:400;font-size:12px;color:var(--muted)">{run_rate_pct:.0f}%</span></span>
        </div>
        <div class="lz-bar-track"><div class="lz-bar-fill" style="width:{run_rate_bar_w:.1f}%;background:var(--purple)"></div></div>
        <div class="lz-scenario-sub" style="color:var(--muted)">Based on complete months average; latest month scaled for days elapsed</div>
      </div>
    </div>

    <div class="lz-card">
      <div class="lz-title">What This Means</div>
      <div class="lz-insight">
        <p style="margin-bottom:12px">The <strong>realistic projection of {fmt_usd(realistic_proj)}</strong> assumes all On Track deals deliver in full and 50% of At Risk deals close. This puts Seamfix at <strong>{realistic_pct:.0f}% of the ${LANDING_ZONE/1_000_000:.0f}M target</strong>{f', a surplus of <strong style="color:var(--green)">{fmt_usd(abs(realistic_gap))}</strong> above target' if realistic_gap <= 0 else f', leaving a gap of <strong>{fmt_usd(realistic_gap)}</strong>'}.</p>
        <p style="margin-bottom:12px">If At Risk and Off Track deals are not resolved, the conservative outcome is <strong style="color:{'var(--green)' if conservative_gap <= 0 else 'var(--red)'}">{fmt_usd(conservative_proj)}</strong> — {'still above target' if conservative_gap <= 0 else f'a shortfall of <strong style="color:var(--red)">{fmt_usd(conservative_gap)}</strong>, which is {conservative_gap/LANDING_ZONE*100:.0f}% of the annual target unfunded'}.</p>
        <p style="margin-bottom:12px">The YTD annualised run rate of <strong style="color:var(--purple)">{fmt_usd(annual_run_rate)}</strong> {'is tracking below the target pace' if annual_run_rate < LANDING_ZONE else 'is tracking at or above target pace'} — {'later-stage deals (currently at $0 actuals) <em>must</em> convert to sustain this trajectory' if annual_run_rate < LANDING_ZONE else 'continued conversion of pipeline deals will solidify this position'}.</p>
        <p><strong>The {len(stalled)} On Track deals showing zero YTD revenue ({fmt_usd(stalled_value)} combined)</strong> are the most important item to verify. If even 40% of those activate, the realistic projection improves materially.</p>
      </div>
      <div style="margin-top:16px">
        <canvas id="donutChart" height="180"></canvas>
      </div>
    </div>
  </div>
</div>

<!-- MOMENTUM TRACKER — shown first per request -->
<div class="section">
  <div class="section-title">Month-on-Month Momentum ({len(movers)} active deals)</div>
  <div class="chart-card" style="margin-bottom:16px">
    <div class="chart-title">Top Revenue Movers — {ytd_label} Actuals (USD)</div>
    <canvas id="momentumChart" height="180"></canvas>
  </div>
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th style="min-width:130px;max-width:160px">Deal</th><th>Status</th>
          {''.join(f'<th style="text-align:right">{MONTH_NAMES[mi]}{" (partial)" if mi == last_data_month and is_partial else ""}</th>' for mi in range(num_months_with_data))}
          <th style="text-align:right">YTD Total</th>
          <th>Trend</th>
          <th style="min-width:180px">Notes</th>
        </tr>
      </thead>
      <tbody>{movers_rows}</tbody>
    </table>
  </div>
</div>

<!-- STALLED ON TRACK -->
<div class="section">
  <div class="section-title">⚠️ On Track — But No Revenue Yet ({len(stalled)} deals, {fmt_usd(stalled_value)} at stake)</div>
  <div class="alert-box" style="margin-bottom:16px">
    <div class="alert-box-title">Verification Required</div>
    <div class="alert-box-sub">These deals are labelled "On Track" but have recorded $0 revenue through April. They may be pre-revenue (future start dates) or genuinely stalled. Each should be confirmed this week.</div>
  </div>
  <div class="table-wrap" style="border-left:4px solid #3b82f6">
    <table>
      <thead><tr><th>Deal</th><th>Rail</th><th style="text-align:right">Annual Value</th><th>Status Notes</th></tr></thead>
      <tbody>{stalled_rows}</tbody>
    </table>
  </div>
</div>

<!-- AT RISK -->
<div class="section">
  <div class="section-title">🔴 At-Risk Deals — {at_risk_count} deals / {fmt_usd(at_risk_val)} at stake</div>
  <p style="color:var(--muted);font-size:13px;margin-bottom:16px">
    Resolving all At Risk deals recovers up to {fmt_usd(at_risk_val)}{f' — narrowing the gap from {fmt_usd(conservative_gap)} to {fmt_usd(max(0, realistic_gap))}' if conservative_gap > 0 else ' — strengthening an already above-target position'}.
  </p>
  {at_risk_cards}
</div>

<!-- OFF TRACK -->
<div class="section">
  <div class="section-title">⛔ Off-Track Deals — {off_track_count} deals / {fmt_usd(off_track_val)} at stake</div>
  {off_track_cards}
</div>

<!-- CLOSED / SECURED -->
<div class="section">
  <div class="section-title">🔒 Closed / Secured — {closed_count} deals / {fmt_usd(closed_val)}</div>
  <p style="color:var(--muted);font-size:13px;margin-bottom:16px">Deals confirmed and secured. Revenue is fully counted in projections.</p>
  <div class="table-wrap" style="border-left:4px solid #10b981">
    <table>
      <thead><tr><th>Deal</th><th>Rail</th><th style="text-align:right">Annual Value</th><th style="text-align:right">YTD Actual</th><th>Notes</th></tr></thead>
      <tbody>{closed_rows}</tbody>
    </table>
  </div>
</div>

<!-- UNKNOWN / UNTAGGED -->
<div class="section">
  <div class="section-title">❓ Untagged Deals — {unknown_count} deals / {fmt_usd(unknown_val)}</div>
  <div class="alert-box" style="margin-bottom:16px;border-color:var(--warning);background:var(--warning-bg)">
    <div class="alert-box-title" style="color:var(--warning-bright)">Status Missing</div>
    <div class="alert-box-sub">These deals have no status assigned in the revenue file. Some have YTD actuals already flowing. The revenue team should tag each with On Track / At Risk / Off Track / Closed to ensure they are correctly counted in projections.</div>
  </div>
  <div class="table-wrap">
    <table>
      <thead><tr><th>Deal</th><th>Rail</th><th style="text-align:right">Annual Value</th><th style="text-align:right">YTD Actual</th><th>Notes</th></tr></thead>
      <tbody>{unknown_rows}</tbody>
    </table>
  </div>
</div>

<div class="footer" style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px">
  <span>Seamfix Pipeline Intelligence &nbsp;·&nbsp; Powered by Claude Cowork</span>
  <span>Generated: {generated_at} &nbsp;&bull;&nbsp; $1 = ₦{FX_RATE:,}</span>
</div>

<script>
// ── DONUT CHART ─────────────────────────────────
const donutCtx = document.getElementById('donutChart').getContext('2d');
new Chart(donutCtx, {{
  type: 'doughnut',
  data: {{
    labels: {json.dumps(donut_labels)},
    datasets: [{{
      data: {json.dumps(donut_values)},
      backgroundColor: ['#10b981','#f59e0b','#ef4444','#3b82f6','#6b7280'],
      borderWidth: 0, borderRadius: 3,
    }}]
  }},
  options: {{
    responsive: true, cutout: '70%',
    plugins: {{
      legend: {{ position:'right', labels: {{ color:'#94a3b8', boxWidth:12, padding:12, font:{{ size:12 }} }} }},
      tooltip: {{
        callbacks: {{
          label: ctx => ` ${{ctx.label}}: ${{(ctx.parsed/1e6).toFixed(2)}}M (${{(ctx.parsed/8000000*100).toFixed(0)}}%)`
        }}
      }}
    }}
  }}
}});

// ── MOMENTUM CHART ───────────────────────────────
const momCtx = document.getElementById('momentumChart').getContext('2d');
new Chart(momCtx, {{
  type: 'bar',
  data: {{
    labels: {json.dumps(momentum_labels)},
    datasets: [
      {','.join(f"{{ label:'{md['label']}', data:{json.dumps(md['data'])}, backgroundColor:'{md['color']}80', borderColor:'{md['color']}', borderWidth:1, borderRadius:3 }}" for md in momentum_month_data)}
    ]
  }},
  options: {{
    responsive:true, indexAxis:'y',
    plugins: {{
      legend: {{ labels:{{ color:'#94a3b8', font:{{ size:12 }} }} }},
      tooltip: {{ callbacks: {{ label: ctx => ` ${{ctx.dataset.label}}: ${{ctx.parsed.x.toLocaleString('en-US',{{minimumFractionDigits:0,maximumFractionDigits:0}})}}` }} }}
    }},
    scales: {{
      x: {{ grid:{{color:'#1e293b'}}, ticks:{{ color:'#94a3b8', callback: v => '$'+v.toLocaleString() }} }},
      y: {{ grid:{{display:false}}, ticks:{{ color:'#94a3b8', font:{{ size:11 }} }} }}
    }}
  }}
}});

{theme_js}

// Apply theme colors to charts after they are created
(function() {{
  var saved = localStorage.getItem('seamfix-theme') || 'light';
  updateCharts(saved);
}})();
</script>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"✅  Pipeline dashboard → {output_path}")


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(os.path.abspath(__file__))
    rev_files = glob.glob(os.path.join(folder, '*Path to Revenue*'))
    if not rev_files:
        print("❌  Could not find '2026 Path to Revenue' file."); sys.exit(1)
    revenue_file = rev_files[0]
    print(f"📊  Revenue file: {os.path.basename(revenue_file)}")

    revenues = extract_revenue_data(revenue_file)
    print(f"    Loaded {len(revenues)} deal rows")

    output_path = os.path.join(folder, 'pipeline_dashboard.html')
    generate_html(revenues, output_path)


if __name__ == '__main__':
    main()
