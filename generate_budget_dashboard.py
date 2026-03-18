#!/usr/bin/env python3
"""
Seamfix Budget vs Actual Dashboard Generator
Compares budget from Excel file against actual spend from cash reports.
Usage: python3 generate_budget_dashboard.py [folder_path]
"""

import os, sys, json, re, glob
from datetime import datetime
from openpyxl import load_workbook
from difflib import SequenceMatcher


# --- Investment detection (reused from generate_dashboard.py) ---
INVESTMENT_OUT_KEYWORDS = ['investment in', 'funding']
INVESTMENT_IN_KEYWORDS = ['investment withdrawal', 'investment liquidation']
INVESTMENT_OUT_EXACT = ['Seamfix UAE funding']
INVESTMENT_IN_EXACT = []


def is_investment_outflow(category_name):
    """Check if an outflow category is an investment transfer (not operational)."""
    if not category_name:
        return False
    cat_lower = str(category_name).strip().lower()
    if category_name in INVESTMENT_OUT_EXACT:
        return True
    return any(kw in cat_lower for kw in INVESTMENT_OUT_KEYWORDS)


def is_investment_inflow(category_name):
    """Check if an inflow category is an investment liquidation/withdrawal."""
    if not category_name:
        return False
    cat_lower = str(category_name).strip().lower()
    if category_name in INVESTMENT_IN_EXACT:
        return True
    return any(kw in cat_lower for kw in INVESTMENT_IN_KEYWORDS)


def get_investment_outflows(outflow_items):
    """Sum all investment outflows from an outflow_items dict."""
    return sum(v for k, v in outflow_items.items() if is_investment_outflow(k))


def safe_row_dict(row):
    d = {}
    for c in row:
        try:
            d[c.column_letter] = c.value
        except (AttributeError, TypeError):
            pass
    return d


def parse_date(fn):
    m = re.search(r'(\d+)\w*\s+(January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})', fn, re.IGNORECASE)
    if not m:
        return None
    day, ms, year = m.groups()
    mm = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
          'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    return datetime(int(year), mm[ms.lower()], int(day))


def sf(val):
    if val is None: return 0.0
    try: return float(val)
    except: return 0.0


def fmt_naira(val):
    """Format as Naira with appropriate suffix"""
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000_000:
        return f"{sign}\u20A6{v/1_000_000_000:.2f}B"
    elif v >= 1_000_000:
        return f"{sign}\u20A6{v/1_000_000:.1f}M"
    elif v >= 1_000:
        return f"{sign}\u20A6{v/1_000:.0f}K"
    else:
        return f"{sign}\u20A6{v:,.0f}"


def extract_report(filepath):
    """Extract outflow data from a cash report file."""
    fn = os.path.basename(filepath)
    dt = parse_date(fn)
    if not dt:
        return None

    wb = load_workbook(filepath, data_only=True)
    rec = {'date': dt, 'date_str': dt.strftime('%d %b %Y'), 'filename': fn}

    # === BREAKDOWN OF INCOME AND EXPENDITURE SHEET ===
    try:
        ws_bd = wb['Breakdown of Income and Expendi']
    except KeyError:
        wb.close()
        return None

    outflow_items = {}
    in_outflows = False

    for row in ws_bd.iter_rows(min_row=1, max_row=ws_bd.max_row, values_only=False):
        cells = safe_row_dict(row)
        b_val = str(cells.get('B', '') or '').strip()
        c_val = cells.get('C')

        if 'OUTFLOW' in b_val.upper() and c_val is not None and str(c_val).strip() in ('Amount', ''):
            in_outflows = True
            continue
        if b_val.upper().startswith('OUTFLOW') and c_val is None:
            in_outflows = True
            continue

        if in_outflows and c_val is not None and isinstance(c_val, (int, float)):
            if 'Total' in b_val and 'Outflow' in b_val:
                in_outflows = False
            elif 'Net Cash' in b_val:
                continue
            elif b_val and b_val not in ('OUTFLOWS',):
                outflow_items[b_val] = sf(c_val)

    rec['outflow_items'] = outflow_items
    wb.close()
    return rec


def extract_budget(filepath):
    """Extract budget from 'Budget Summary' sheet."""
    wb = load_workbook(filepath, data_only=True)
    try:
        ws = wb['Budget Summary']
    except KeyError:
        wb.close()
        return None

    budget_items = {}
    for row in ws.iter_rows(min_row=5, max_row=36, values_only=False):
        cells = safe_row_dict(row)
        a_val = cells.get('A')
        b_val = cells.get('B')

        if a_val and b_val and isinstance(b_val, (int, float)):
            budget_items[str(a_val).strip()] = sf(b_val)

    wb.close()
    return budget_items


# Define all 32 budget categories from the specification
BUDGET_CATEGORIES = {
    'Payroll': 2862494488,
    'Staff welfare & benefits': 554171104.4,
    'Business Development Expenses': 179880000,
    'Software Tools': 223191465,
    'New Product Development - IAM': 350000000,
    'Others - AIF Coupon': 300375000,
    'Professional Fees': 118038333.7,
    'Marketing & Branding - Events': 132000000,
    'Hosting Expenses': 127457500,
    'Retreat expenses': 0,
    'Learning & Development': 94000,
    'Miscellaneous Expenses': 0,
    'Assets purchase': 45082105.6,
    'Marketing & Branding - Digital': 45000000,
    'Rental Expense': 59504984.58,
    'Business Expansion': 0,
    'Freight and Clearing': 12000000,
    'Repairs': 14522528.5,
    'Office Expense': 13174011.64,
    'Insurance Expenses': 13137380.48,
    'NED Fees': 13400000,
    'Tenders/Bids Purchase Expenses': 4000000,
    'Diesel': 6900000,
    'Staff Allowances': 7464000,
    'Corporate Gift expense': 0,
    'Transport Expenses': 4824000,
    'Motor Vehicle Expense': 4893480,
    'Printing': 4170953.1,
    'Dues & Subscriptions': 3156058.904,
    'Business License and Permit': 275000,
    'Property Maintenance Expenses': 0,
    'Charitable Donations/CSR': 0,
}

# Category mapping: cash report expense names -> budget categories
# --- Exact matches first (checked before keyword search) ---
EXACT_MAP = {
    'Bind Creative': 'New Product Development - IAM',  # IAM product dev agency
    'Nimc consultancy fee': 'Professional Fees',
    'NSITF Contributions': 'Staff welfare & benefits',
    'Pension Remittance': 'Payroll',
    'Pension': 'Payroll',
    'PAYE': 'Payroll',
    'Paye': 'Payroll',
    'WHT remitted': None,        # Tax remittance, not a budget expense
    'WHT': None,
    'VAT': None,                 # Tax remittance
    'VAT Remitted': None,
    'Tatvasoft': 'Professional Fees',
    'BTA': 'Business Development Expenses',
    'ID4Africa': 'Marketing & Branding - Events',
    'ID4Africa Registration': 'Marketing & Branding - Events',
    'MWC Registration': 'Marketing & Branding - Events',
    'AWS': 'Hosting Expenses',
    'Shipping & Freight': 'Freight and Clearing',
    'Training': 'Learning & Development',
    'Business Premises & Dev Levy': 'Business License and Permit',
    'Management Allowance': 'Payroll',  # Part of staff compensation
    'NIMC -Licenses Fee': 'Business Development Expenses',  # License procurement for resale
    'Exit  Salary': 'Payroll',
    'Exit Salary': 'Payroll',
    'Payment Batches': 'Payroll',
    'Office Supplies, Fuel & Letter Dispatch': 'Office Expense',
    'Office Supplies & Fuel': 'Office Expense',
    'Refund': None,              # Not an expense
}

# --- Keyword-based fallback (substring match, checked in order) ---
CATEGORY_MAP = {
    'Salary': 'Payroll',
    'Salaries': 'Payroll',
    'Pension': 'Payroll',
    'Staff welfare': 'Staff welfare & benefits',
    'Staff allowances': 'Staff Allowances',
    'Consulting': 'Professional Fees',
    'Consultancy': 'Professional Fees',
    'legal': 'Professional Fees',
    'Professional Fees': 'Professional Fees',
    'Office Rent': 'Rental Expense',
    'Rental': 'Rental Expense',
    'Amazon': 'Assets purchase',
    'Assets': 'Assets purchase',
    'Generator Diesel': 'Diesel',
    'Diesel': 'Diesel',
    'Office Supplies': 'Office Expense',
    'Internet': 'Software Tools',
    'Software': 'Software Tools',
    'Hosting': 'Hosting Expenses',
    'AWS': 'Hosting Expenses',
    'Bank Charges': 'Miscellaneous Expenses',
    'Travel': 'Business Development Expenses',
    'Flight': 'Business Development Expenses',
    'BTA': 'Business Development Expenses',
    'Marketing': 'Marketing & Branding - Digital',
    'Branding': 'Marketing & Branding - Digital',
    'Creative': 'Marketing & Branding - Digital',
    'Events': 'Marketing & Branding - Events',
    'Conference': 'Marketing & Branding - Events',
    'Registration': 'Marketing & Branding - Events',
    'Insurance': 'Insurance Expenses',
    'NSITF': 'Staff welfare & benefits',
    'NHF': 'Staff welfare & benefits',
    'HMO': 'Staff welfare & benefits',
    'Transport': 'Transport Expenses',
    'Motor Vehicle': 'Motor Vehicle Expense',
    'Printing': 'Printing',
    'Subscriptions': 'Dues & Subscriptions',
    'Dues': 'Dues & Subscriptions',
    'Repairs': 'Repairs',
    'Maintenance': 'Property Maintenance Expenses',
    'Training': 'Learning & Development',
    'Freight': 'Freight and Clearing',
    'Shipping': 'Freight and Clearing',
    'License': 'Business License and Permit',
    'Permit': 'Business License and Permit',
}


def map_expense_to_budget(expense_name, outflow_items_all):
    """Map a cash report expense name to a budget category using keyword matching."""
    if not expense_name:
        return None

    # Skip investment outflows
    if is_investment_outflow(expense_name):
        return None

    # 1. Exact match (handles specific items like 'Bind Creative', 'Pension', 'VAT')
    stripped = expense_name.strip()
    if stripped in EXACT_MAP:
        return EXACT_MAP[stripped]  # May return None for tax remittances etc.

    # 2. Keyword-based matching (substring search)
    expense_lower = stripped.lower()
    for keyword, budget_cat in CATEGORY_MAP.items():
        if keyword.lower() in expense_lower:
            return budget_cat

    # 3. Fuzzy match as last resort (raised threshold to 0.6 to reduce false positives)
    best_match = None
    best_ratio = 0.0
    for budget_cat in BUDGET_CATEGORIES.keys():
        ratio = SequenceMatcher(None, expense_lower, budget_cat.lower()).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = budget_cat

    if best_ratio > 0.6:
        return best_match

    return None


def generate_html(budget_items, reports, output_path):
    """Generate the budget vs actual dashboard HTML."""

    # Calculate weeks elapsed
    weeks_elapsed = len(reports)

    # Strategic spend buckets — classifies the 32 budget categories into 4 groups
    STRATEGIC_BUCKETS = {
        'Revenue-Generating': [
            'Business Development Expenses', 'Marketing & Branding - Events',
            'Marketing & Branding - Digital', 'New Product Development - IAM',
            'Tenders/Bids Purchase Expenses', 'Business Expansion', 'Corporate Gift expense',
        ],
        'People Costs': [
            'Payroll', 'Staff welfare & benefits', 'Staff Allowances',
            'Learning & Development', 'Retreat expenses',
        ],
        'Technology & Infrastructure': [
            'Software Tools', 'Hosting Expenses', 'Assets purchase', 'Dues & Subscriptions',
        ],
        'Operations & Overhead': [
            'Rental Expense', 'Diesel', 'Office Expense', 'Repairs',
            'Property Maintenance Expenses', 'Freight and Clearing', 'Transport Expenses',
            'Motor Vehicle Expense', 'Printing', 'Insurance Expenses',
            'Business License and Permit', 'NED Fees', 'Professional Fees',
            'Others - AIF Coupon', 'Miscellaneous Expenses', 'Charitable Donations/CSR',
        ],
    }
    BUCKET_COLORS = {
        'Revenue-Generating':       '#10b981',   # green
        'People Costs':             '#3b82f6',   # blue
        'Technology & Infrastructure': '#8b5cf6', # purple
        'Operations & Overhead':    '#f59e0b',   # orange
    }

    # Aggregate actual spend by budget category
    category_actual = {cat: 0 for cat in BUDGET_CATEGORIES.keys()}
    all_expense_lines = []   # For deep dive table
    unbudgeted_lines = []    # Expense lines with no matching budget category
    weekly_spends = []       # For trend chart

    for r in reports:
        week_spend = 0
        for expense_name, amount in r.get('outflow_items', {}).items():
            if amount <= 0:
                continue
            budget_cat = map_expense_to_budget(expense_name, r.get('outflow_items', {}))

            if budget_cat:
                category_actual[budget_cat] += amount
                week_spend += amount
                all_expense_lines.append({
                    'date': r['date_str'],
                    'item': expense_name,
                    'amount': amount,
                    'category': budget_cat
                })
            elif not is_investment_outflow(expense_name):
                # Not an investment transfer and has no budget category — governance flag
                unbudgeted_lines.append({
                    'date': r['date_str'],
                    'item': expense_name,
                    'amount': amount,
                })

        weekly_spends.append(week_spend)

    # Calculate KPIs
    total_budget = sum(BUDGET_CATEGORIES.values())
    total_actual = sum(category_actual.values())
    ytd_budget_pace = total_budget * weeks_elapsed / 52
    variance_amount = ytd_budget_pace - total_actual
    variance_pct = (variance_amount / ytd_budget_pace * 100) if ytd_budget_pace > 0 else 0

    # Year-end projection: extrapolate current spend rate across full 52 weeks
    projected_year_end = (total_actual / weeks_elapsed * 52) if weeks_elapsed > 0 else 0
    projected_vs_budget = total_budget - projected_year_end   # positive = under, negative = over
    projected_pct = (abs(projected_vs_budget) / total_budget * 100) if total_budget > 0 else 0
    projected_over = projected_year_end > total_budget

    # Calculate per-category metrics
    category_metrics = []
    for cat in sorted(BUDGET_CATEGORIES.keys()):
        annual_budget = BUDGET_CATEGORIES[cat]
        ytd_budget = annual_budget * weeks_elapsed / 52
        ytd_actual = category_actual[cat]
        variance = ytd_budget - ytd_actual
        pct_used = (ytd_actual / ytd_budget * 100) if ytd_budget > 0 else 0
        # Annual envelope view: how much of the full-year approved budget is consumed
        pct_of_annual = (ytd_actual / annual_budget * 100) if annual_budget > 0 else 0
        annual_remaining = annual_budget - ytd_actual

        category_metrics.append({
            'name': cat,
            'annual_budget': annual_budget,
            'ytd_budget': ytd_budget,
            'ytd_actual': ytd_actual,
            'variance': variance,
            'pct_used': pct_used,
            'pct_of_annual': pct_of_annual,
            'annual_remaining': annual_remaining,
            'status': 'over' if ytd_actual > ytd_budget else 'under' if ytd_actual < ytd_budget * 0.5 else 'on-track'
        })

    # Strategic bucket totals
    bucket_totals = {}
    for bucket, cats in STRATEGIC_BUCKETS.items():
        b_budget = sum(BUDGET_CATEGORIES.get(c, 0) for c in cats)
        b_actual = sum(category_actual.get(c, 0) for c in cats)
        b_remaining = b_budget - b_actual
        b_pct = (b_actual / b_budget * 100) if b_budget > 0 else 0
        bucket_totals[bucket] = {
            'budget': b_budget, 'actual': b_actual,
            'remaining': b_remaining, 'pct': b_pct,
            'color': BUCKET_COLORS[bucket],
        }

    # Unbudgeted spend summary
    total_unbudgeted = sum(x['amount'] for x in unbudgeted_lines)
    unbudgeted_lines.sort(key=lambda x: -x['amount'])

    # Sort by overspend amount
    category_metrics.sort(key=lambda x: -max(0, x['ytd_actual'] - x['ytd_budget']))

    # Top 15 for chart
    top_categories = sorted(category_metrics, key=lambda x: x['annual_budget'], reverse=True)[:15]

    # Generate HTML
    generated_at = datetime.now().strftime('%d %b %Y %H:%M')
    latest_date        = reports[-1]['date_str'] if reports else 'N/A'
    first_report_date  = reports[0]['date_str']  if reports else 'N/A'

    # Build category metrics table rows
    category_table_rows = ""
    for m in category_metrics:
        status_icon = "🔴" if m['status'] == 'over' else "🟢" if m['status'] == 'under' else "🟡"
        status_text = "Over Budget" if m['status'] == 'over' else "Under Budget" if m['status'] == 'under' else "On Track"

        remaining_class = 'negative' if m['annual_remaining'] < 0 else 'positive'
        remaining_label = f"({fmt_naira(abs(m['annual_remaining']))} over)" if m['annual_remaining'] < 0 else fmt_naira(m['annual_remaining'])
        category_table_rows += f"""<tr>
<td>{m['name']}</td>
<td>{fmt_naira(m['annual_budget'])}</td>
<td>{fmt_naira(m['ytd_budget'])}</td>
<td>{fmt_naira(m['ytd_actual'])}</td>
<td class="{'negative' if m['variance'] < 0 else 'positive'}">{fmt_naira(abs(m['variance']))}</td>
<td>{m['pct_used']:.1f}%</td>
<td>{m['pct_of_annual']:.1f}%</td>
<td class="{remaining_class}">{remaining_label}</td>
<td>{status_icon} {status_text}</td>
</tr>"""

    # Build deep dive table rows (all expense lines, sorted by date)
    all_expense_lines.sort(key=lambda x: x['date'], reverse=True)
    deep_dive_rows = ""
    for line in all_expense_lines:
        deep_dive_rows += f"""<tr>
<td>{line['date']}</td>
<td>{line['item']}</td>
<td>{fmt_naira(line['amount'])}</td>
<td>{line['category']}</td>
</tr>"""

    # Build bar chart data (top 15)
    chart_labels = [m['name'] for m in top_categories]
    chart_ytd_budget = [m['ytd_budget'] for m in top_categories]
    chart_ytd_actual = [m['ytd_actual'] for m in top_categories]

    # Weeks for trend chart
    trend_weeks = [r['date_str'] for r in reports]

    # Executive takeaways
    overspend_cats = [m for m in category_metrics if m['status'] == 'over']
    underspend_cats = [m for m in category_metrics if m['status'] == 'under' and m['ytd_budget'] > 0]
    critical_alerts = [m for m in category_metrics if m['ytd_actual'] > m['ytd_budget'] * 1.5]

    takeaway_html = ""

    # Budget Health
    # variance_pct = (ytd_budget_pace - total_actual) / ytd_budget_pace * 100
    # Positive = actual is UNDER pace (good), Negative = actual is OVER pace (bad)
    if variance_pct > 10:
        health_status = "HEALTHY"
        health_color = "00D4AA"
        health_text = f"Currently {abs(variance_pct):.0f}% under YTD budget pace. Spending is well-controlled."
    elif variance_pct < -10:
        health_status = "CAUTION"
        health_color = "FFE66D"
        health_text = f"Currently {abs(variance_pct):.0f}% over YTD budget pace. Monitor closely to avoid overruns."
    else:
        health_status = "ON TRACK"
        health_color = "4ECDC4"
        health_text = f"Currently within {abs(variance_pct):.0f}% of YTD budget pace. Spending is aligned."

    takeaway_html += f"""<div class="takeaway-card" style="border-left-color:#{health_color}">
<div class="tw-header">
<span class="tw-icon">📊</span>
<span class="tw-title">Budget Health</span>
<span class="tw-badge" style="background:rgba({','.join(str(int(x, 16)) for x in [health_color[i:i+2] for i in range(0, 6, 2)])},0.15);color:#{health_color}">{health_status}</span>
</div>
<div class="tw-headline" style="color:#{health_color}">YTD Performance Assessment</div>
<div class="tw-body">{health_text} YTD spend is {fmt_naira(total_actual)} against a pace of {fmt_naira(ytd_budget_pace)}.</div>
</div>"""

    # Top Overspend Areas
    if overspend_cats:
        def safe_pct(actual, budget):
            return f"(+{((actual/budget-1)*100):.0f}%)" if budget > 0 else ""
        overspend_list = ''.join([f"<li>{m['name']}: {fmt_naira(m['ytd_actual'] - m['ytd_budget'])} over {safe_pct(m['ytd_actual'], m['ytd_budget'])}</li>" for m in overspend_cats[:5]])
        takeaway_html += f"""<div class="takeaway-card" style="border-left-color:#FF6B6B">
<div class="tw-header">
<span class="tw-icon">⚠️</span>
<span class="tw-title">Top Overspend Areas</span>
<span class="tw-badge" style="background:rgba(255,107,107,0.15);color:#FF6B6B">ACTION</span>
</div>
<div class="tw-headline" style="color:#FF6B6B">Categories Exceeding YTD Pace</div>
<div class="tw-body"><ul style="margin:6px 0 0 16px;padding:0;list-style:disc">{overspend_list}</ul></div>
</div>"""

    # Savings Opportunities
    if underspend_cats:
        savings_list = ''.join([f"<li>{m['name']}: {fmt_naira(m['ytd_budget'] - m['ytd_actual'])} remaining ({100-m['pct_used']:.0f}% unspent)</li>" for m in underspend_cats[:5]])
        takeaway_html += f"""<div class="takeaway-card" style="border-left-color:#00D4AA">
<div class="tw-header">
<span class="tw-icon">💡</span>
<span class="tw-title">Savings Opportunities</span>
<span class="tw-badge" style="background:rgba(0,212,170,0.15);color:#00D4AA">POSITIVE</span>
</div>
<div class="tw-headline" style="color:#00D4AA">Categories Under Budget Pace</div>
<div class="tw-body"><ul style="margin:6px 0 0 16px;padding:0;list-style:disc">{savings_list}</ul></div>
</div>"""

    # Critical Alerts
    if critical_alerts:
        alert_list = ''.join([f"<li>{m['name']}: {((m['ytd_actual']/m['ytd_budget']-1)*100):.0f}% over pace</li>" for m in critical_alerts[:3] if m['ytd_budget'] > 0])
        if alert_list:
            takeaway_html += f"""<div class="takeaway-card" style="border-left-color:#FF6B6B">
<div class="tw-header">
<span class="tw-icon">🚨</span>
<span class="tw-title">Critical Alerts</span>
<span class="tw-badge" style="background:rgba(255,107,107,0.15);color:#FF6B6B">URGENT</span>
</div>
<div class="tw-headline" style="color:#FF6B6B">Categories >150% of YTD Budget Pace</div>
<div class="tw-body"><ul style="margin:6px 0 0 16px;padding:0;list-style:disc">{alert_list}</ul></div>
</div>"""

    # Build strategic spend breakdown HTML
    strategic_section_html = ""
    total_known_actual = sum(bt['actual'] for bt in bucket_totals.values())
    for bucket, bt in bucket_totals.items():
        color = bt['color']
        bar_pct = min(bt['pct'], 100)
        share_pct = (bt['actual'] / total_known_actual * 100) if total_known_actual > 0 else 0
        strategic_section_html += f"""<div class="strategic-bucket" style="border-left:4px solid {color};background:rgba(15,23,42,0.6);border-radius:8px;padding:16px 20px;margin-bottom:12px;">
<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:8px;">
  <span style="font-weight:600;font-size:14px;color:#e2e8f0">{bucket}</span>
  <span style="font-size:12px;color:#94a3b8">{share_pct:.1f}% of total spend</span>
</div>
<div style="display:flex;gap:20px;margin-bottom:10px;flex-wrap:wrap">
  <div><span style="font-size:11px;color:#94a3b8">YTD Actual</span><br><span style="font-size:18px;font-weight:700;color:{color}">{fmt_naira(bt['actual'])}</span></div>
  <div><span style="font-size:11px;color:#94a3b8">Annual Budget</span><br><span style="font-size:16px;color:#cbd5e1">{fmt_naira(bt['budget'])}</span></div>
  <div><span style="font-size:11px;color:#94a3b8">Remaining</span><br><span style="font-size:16px;color:{'#ef4444' if bt['remaining'] < 0 else '#10b981'}">{fmt_naira(abs(bt['remaining']))} {'over' if bt['remaining'] < 0 else 'left'}</span></div>
</div>
<div style="background:#1e293b;border-radius:4px;height:8px;overflow:hidden">
  <div style="height:100%;width:{bar_pct:.1f}%;background:{color};border-radius:4px;transition:width 0.5s"></div>
</div>
<div style="font-size:11px;color:#94a3b8;margin-top:4px">{bt['pct']:.1f}% of annual budget consumed</div>
</div>"""

    # Build unbudgeted spend rows HTML
    unbudgeted_rows_html = ""
    for line in unbudgeted_lines[:50]:   # cap at 50 rows for readability
        unbudgeted_rows_html += f"""<tr>
<td>{line['date']}</td>
<td>{line['item']}</td>
<td class="negative">{fmt_naira(line['amount'])}</td>
</tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Seamfix Budget vs Actual Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:linear-gradient(135deg,#0f172a,#1e293b);color:#e2e8f0;padding:0;min-height:100vh}}
.container{{max-width:1600px;margin:0 auto;padding:0 28px 28px}}
.header{{margin-bottom:32px;padding-bottom:20px;border-bottom:2px solid rgba(0,212,170,0.2)}}
.header h1{{font-size:2.4em;font-weight:700;background:linear-gradient(135deg,#00D4AA,#4ECDC4);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:4px}}
.header .sub{{color:#94a3b8;font-size:0.95em}}
.header .meta{{color:#64748b;font-size:0.8em;margin-top:6px}}
.nav-tabs{{display:none!important}}
.top-nav{{background:#0a0f1e;border-bottom:1px solid #334155;padding:0 24px;display:flex;align-items:center;height:48px;overflow-x:auto;position:sticky;top:0;z-index:200}}
.top-nav-brand{{color:#e2e8f0;font-weight:700;font-size:15px;margin-right:24px;white-space:nowrap;text-decoration:none}}
.top-nav-link{{color:#94a3b8;text-decoration:none;padding:0 14px;height:48px;display:flex;align-items:center;font-size:13px;border-bottom:2px solid transparent;white-space:nowrap;transition:color .2s}}
.top-nav-link:hover{{color:#e2e8f0}}
.top-nav-link.active{{color:#fff;border-bottom-color:#00D4AA;font-weight:500}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:16px;margin-bottom:36px}}
.kpi-card{{background:linear-gradient(135deg,rgba(30,41,59,0.9),rgba(15,23,42,0.9));border:1px solid rgba(0,212,170,0.15);border-radius:12px;padding:22px;transition:all 0.3s ease}}
.kpi-card:hover{{border-color:rgba(0,212,170,0.4);transform:translateY(-3px);box-shadow:0 8px 30px rgba(0,212,170,0.1)}}
.kpi-label{{font-size:0.78em;color:#94a3b8;text-transform:uppercase;letter-spacing:1.2px;margin-bottom:8px;font-weight:600}}
.kpi-value{{font-size:1.6em;font-weight:700;color:#00D4AA;margin-bottom:6px}}
.kpi-value.negative{{color:#FF6B6B}}
.kpi-change{{font-size:0.82em;display:flex;align-items:center;gap:4px}}
.kpi-change.positive{{color:#00D4AA}}
.kpi-change.negative{{color:#FF6B6B}}
.kpi-change.neutral{{color:#64748b}}
.charts-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(480px,1fr));gap:20px;margin-bottom:36px}}
.chart-box{{background:linear-gradient(135deg,rgba(30,41,59,0.9),rgba(15,23,42,0.9));border:1px solid rgba(0,212,170,0.15);border-radius:12px;padding:22px;min-height:380px}}
.chart-box h3{{font-size:1em;font-weight:600;margin-bottom:16px;color:#cbd5e1}}
.section{{background:linear-gradient(135deg,rgba(30,41,59,0.9),rgba(15,23,42,0.9));border:1px solid rgba(0,212,170,0.15);border-radius:12px;padding:24px;margin-bottom:24px}}
.section h2{{font-size:1.3em;margin-bottom:16px;color:#00D4AA}}
.search-box{{margin-bottom:16px;padding:12px;background:rgba(0,212,170,0.05);border:1px solid rgba(0,212,170,0.2);border-radius:8px;color:#e2e8f0;font-family:inherit;width:100%;max-width:400px}}
.search-box::placeholder{{color:#64748b}}
table{{width:100%;border-collapse:collapse;font-size:0.85em}}
thead{{background:rgba(0,212,170,0.08)}}
th{{padding:10px 12px;text-align:left;font-weight:600;color:#00D4AA;border-bottom:2px solid rgba(0,212,170,0.15)}}
td{{padding:10px 12px;border-bottom:1px solid rgba(0,212,170,0.06);color:#cbd5e1}}
tbody tr:hover{{background:rgba(0,212,170,0.04)}}
.positive{{color:#00D4AA}}
.negative{{color:#FF6B6B}}
.tabs{{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}}
.tab-btn{{padding:8px 18px;border:1px solid rgba(0,212,170,0.2);border-radius:20px;background:transparent;color:#94a3b8;cursor:pointer;font-size:0.85em;font-family:inherit;transition:all 0.2s}}
.tab-btn.active,.tab-btn:hover{{background:rgba(0,212,170,0.15);color:#00D4AA;border-color:rgba(0,212,170,0.4)}}
.tab-content{{display:none}}
.tab-content.active{{display:block}}
.takeaways-section{{margin-bottom:36px}}
.takeaways-section h2{{font-size:1.4em;margin-bottom:16px;color:#FFE66D;display:flex;align-items:center;gap:10px}}
.takeaways-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(440px,1fr));gap:16px}}
.takeaway-card{{background:linear-gradient(135deg,rgba(30,41,59,0.95),rgba(15,23,42,0.95));border:1px solid rgba(0,212,170,0.1);border-left:4px solid #64748b;border-radius:10px;padding:20px;transition:all 0.3s ease}}
.takeaway-card:hover{{transform:translateY(-2px);box-shadow:0 6px 24px rgba(0,0,0,0.2)}}
.tw-header{{display:flex;align-items:center;gap:8px;margin-bottom:10px}}
.tw-icon{{font-size:1.2em}}
.tw-title{{font-size:0.82em;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:#94a3b8}}
.tw-badge{{font-size:0.7em;font-weight:700;padding:3px 10px;border-radius:12px;letter-spacing:0.5px}}
.tw-headline{{font-size:1.05em;font-weight:600;margin-bottom:8px;line-height:1.4}}
.tw-body{{font-size:0.88em;color:#94a3b8;line-height:1.65}}
.tw-body ul{{margin:6px 0 0 16px}}
.tw-body li{{margin-bottom:4px}}
@media(max-width:1024px){{.charts-grid{{grid-template-columns:1fr}}.kpi-grid{{grid-template-columns:repeat(2,1fr)}}.takeaways-grid{{grid-template-columns:1fr}}}}
@media(max-width:640px){{.kpi-grid{{grid-template-columns:1fr}}.header h1{{font-size:1.6em}}.nav-tabs{{flex-wrap:wrap}}}}
.pdf-btn{{display:none!important}}
th.sortable{{cursor:pointer;user-select:none}}
th.sortable:hover{{color:#FFE66D}}
th.sort-asc::after{{content:' ▲';font-size:0.7em}}
th.sort-desc::after{{content:' ▼';font-size:0.7em}}

@media print{{
.pdf-btn{{display:none!important}}
*{{color-adjust:exact!important;-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}}
body{{background:#fff!important;color:#1e293b!important;padding:12px!important;font-size:9pt!important}}
.container{{max-width:100%}}
.header h1{{font-size:1.6em!important;background:none!important;-webkit-text-fill-color:#0f172a!important;color:#0f172a!important}}
.header .sub{{color:#475569!important}}
.header .meta{{color:#64748b!important}}
.nav-tabs{{display:none!important}}
.kpi-grid{{grid-template-columns:repeat(4,1fr)!important;gap:8px!important;margin-bottom:16px!important}}
.kpi-card{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:12px!important;border-radius:8px!important;box-shadow:none!important;page-break-inside:avoid}}
.kpi-label{{color:#475569!important;font-size:0.65em!important}}
.kpi-value{{color:#0f172a!important;font-size:1.2em!important}}
.kpi-value.negative{{color:#dc2626!important}}
.kpi-change{{font-size:0.7em!important}}
.kpi-change.positive{{color:#16a34a!important}}
.kpi-change.negative{{color:#dc2626!important}}
.kpi-change.neutral{{color:#64748b!important}}
.takeaways-section{{margin-bottom:16px!important;page-break-inside:avoid}}
.takeaways-section h2{{color:#0f172a!important;font-size:1.1em!important}}
.takeaways-grid{{grid-template-columns:repeat(2,1fr)!important;gap:10px!important}}
.takeaway-card{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:14px!important;border-radius:8px!important;page-break-inside:avoid}}
.tw-title{{color:#475569!important}}
.tw-body{{color:#475569!important}}
.tw-headline{{font-size:0.95em!important}}
.charts-grid{{grid-template-columns:1fr!important;gap:12px!important;margin-bottom:16px!important}}
.chart-box{{background:#fff!important;border:1px solid #e2e8f0!important;padding:14px!important;min-height:auto!important;page-break-inside:avoid}}
.chart-box h3{{color:#1e293b!important;font-size:0.85em!important}}
.section{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:16px!important;page-break-inside:avoid}}
.section h2{{color:#0f172a!important}}
table{{font-size:0.75em!important}}
thead{{background:#f1f5f9!important}}
th{{color:#0f172a!important;border-bottom-color:#cbd5e1!important}}
td{{color:#334155!important;border-bottom-color:#e2e8f0!important}}
.positive{{color:#16a34a!important}}
.negative{{color:#dc2626!important}}
.tabs{{display:none!important}}
.tab-content{{display:block!important}}
.search-box{{display:none!important}}
tbody tr:hover{{background:transparent!important}}
}}
.dashboard-footer{{margin-top:48px;padding:20px 28px;border-top:1px solid #334155;color:#94a3b8;font-size:12px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
.dashboard-footer span{{color:#94a3b8}}
/* ── HEADER: match Pipeline Intelligence style ── */
.header{{padding:24px 28px 16px!important;border-bottom:1px solid #334155!important;margin-bottom:24px!important;background:none!important}}
.header h1{{font-size:22px!important;font-weight:700!important;background:none!important;-webkit-background-clip:unset!important;-webkit-text-fill-color:#e2e8f0!important;color:#e2e8f0!important;margin-bottom:4px!important}}
.header .sub{{font-size:13px!important;color:#94a3b8!important}}
.header .meta{{font-size:12px!important;color:#94a3b8!important;margin-top:4px!important}}
.header p{{font-size:13px!important;color:#94a3b8!important;margin-top:4px!important}}
</style>
</head>
<body>
<nav class="top-nav"><span class="top-nav-brand">⚡ Seamfix</span><a href="dashboard.html" class="top-nav-link ">Cash Overview</a><a href="expense_dashboard.html" class="top-nav-link ">Expense &amp; Vendor</a><a href="budget_dashboard.html" class="top-nav-link active">Budget vs Actual</a><a href="revenue_dashboard.html" class="top-nav-link ">Revenue &amp; Fundability</a><a href="pipeline_dashboard.html" class="top-nav-link ">Pipeline Intelligence</a></nav>
<!-- PDF button hidden per user request -->

<div class="container">
<div class="header">
<h1>Seamfix Budget vs Actual</h1>
<div class="sub">Budget Performance Dashboard &mdash; 2026</div>
</div>

<!-- old nav-tabs replaced by top-nav -->

<div class="kpi-grid">
<div class="kpi-card">
<div class="kpi-label">Total Annual Budget</div>
<div class="kpi-value">{fmt_naira(total_budget)}</div>
<div class="kpi-change neutral">All 32 cost heads</div>
</div>
<div class="kpi-card">
<div class="kpi-label">YTD Actual Spend</div>
<div class="kpi-value">{fmt_naira(total_actual)}</div>
<div class="kpi-change neutral">Operational expenses (excl. investments)</div>
</div>
<div class="kpi-card">
<div class="kpi-label">YTD Budget Pace</div>
<div class="kpi-value">{fmt_naira(ytd_budget_pace)}</div>
<div class="kpi-change neutral">{weeks_elapsed} weeks / 52 weeks</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Budget Variance</div>
<div class="kpi-value {'negative' if variance_amount < 0 else ''}">{fmt_naira(abs(variance_amount))}</div>
<div class="kpi-change {'positive' if variance_amount > 0 else 'negative'}">{'+' if variance_amount > 0 else '-'}{abs(variance_pct):.1f}% {'under' if variance_amount > 0 else 'over'}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Projected Year-End Spend</div>
<div class="kpi-value {'negative' if projected_over else ''}">{fmt_naira(projected_year_end)}</div>
<div class="kpi-change {'negative' if projected_over else 'positive'}">{'▲ ' + fmt_naira(abs(projected_vs_budget)) + ' over annual budget' if projected_over else '▼ ' + fmt_naira(abs(projected_vs_budget)) + ' under annual budget'}</div>
</div>
</div>

<div class="takeaways-section">
<h2>📊 Executive Takeaways</h2>
<div class="takeaways-grid">
{takeaway_html}
</div>
</div>

<div class="charts-grid">
<div class="chart-box"><h3>Top 15 Budget Categories: YTD Budget Pace vs Actual Spend</h3><canvas id="budgetChart"></canvas></div>
<div class="chart-box"><h3>Weekly Operational Spend Trend</h3><canvas id="trendChart"></canvas></div>
</div>

<div class="section">
<h2>🎯 Strategic Spend Breakdown</h2>
<p style="color:#94a3b8;font-size:13px;margin:-8px 0 18px">Where is the money going? Spend classified into four strategic buckets to show revenue-generating investment vs operational overhead.</p>
<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(340px,1fr));gap:12px">
{strategic_section_html}
</div>
</div>

<div class="section">
<h2>Budget Variance Analysis</h2>
<div class="tabs">
<button class="tab-btn active" onclick="showTab('variance')">Variance Table</button>
<button class="tab-btn" onclick="showTab('deepdive')">Expense Details</button>
<button class="tab-btn" onclick="showTab('unbudgeted')">⚠ Unbudgeted Spend ({len(unbudgeted_lines)})</button>
</div>
<div id="tab-variance" class="tab-content active">
<table id="varianceTable">
<thead><tr><th class="sortable" onclick="sortTable('varianceTable',0,'text')">Category</th><th class="sortable" onclick="sortTable('varianceTable',1,'money')">Annual Budget</th><th class="sortable" onclick="sortTable('varianceTable',2,'money')">YTD Budget</th><th class="sortable" onclick="sortTable('varianceTable',3,'money')">YTD Actual</th><th class="sortable" onclick="sortTable('varianceTable',4,'money')">Variance</th><th class="sortable" onclick="sortTable('varianceTable',5,'pct')">% Used vs YTD</th><th class="sortable" onclick="sortTable('varianceTable',6,'pct')">Annual Used %</th><th class="sortable" onclick="sortTable('varianceTable',7,'money')">Annual Remaining</th><th class="sortable" onclick="sortTable('varianceTable',8,'status')">Status</th></tr></thead>
<tbody>{category_table_rows}</tbody>
</table>
</div>
<div id="tab-deepdive" class="tab-content">
<input type="text" class="search-box" id="searchBox" placeholder="Search expense items (e.g. laptop, MacBook, office)...">
<table id="deepDiveTable">
<thead><tr><th class="sortable" onclick="sortTable('deepDiveTable',0,'text')">Date</th><th class="sortable" onclick="sortTable('deepDiveTable',1,'text')">Expense Item</th><th class="sortable" onclick="sortTable('deepDiveTable',2,'money')">Amount</th><th class="sortable" onclick="sortTable('deepDiveTable',3,'text')">Mapped Budget Category</th></tr></thead>
<tbody>{deep_dive_rows}</tbody>
</table>
</div>
<div id="tab-unbudgeted" class="tab-content">
<div style="background:rgba(239,68,68,0.08);border:1px solid rgba(239,68,68,0.25);border-radius:8px;padding:14px 18px;margin-bottom:16px">
  <strong style="color:#ef4444">⚠ Governance Flag:</strong>
  <span style="color:#94a3b8;font-size:13px;margin-left:8px">These {len(unbudgeted_lines)} expense line(s) totalling <strong style="color:#ef4444">{fmt_naira(total_unbudgeted)}</strong> have no matching approved budget category. They are excluded from the Variance Table above but are real cash outflows. Review and either reclassify to an existing budget heading or add a new budget line item.</span>
</div>
<table id="unbudgetedTable">
<thead><tr><th class="sortable" onclick="sortTable('unbudgetedTable',0,'text')">Date</th><th class="sortable" onclick="sortTable('unbudgetedTable',1,'text')">Expense Item</th><th class="sortable" onclick="sortTable('unbudgetedTable',2,'money')">Amount</th></tr></thead>
<tbody>{unbudgeted_rows_html}</tbody>
</table>
</div>
</div>

</div>

<script>
function sortTable(tableId, colIdx, type) {{
    const table = document.getElementById(tableId);
    const tbody = table.querySelector('tbody');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    const headers = table.querySelectorAll('th.sortable');

    // Determine sort direction
    const header = headers[colIdx];
    const isAsc = header.classList.contains('sort-asc');

    // Clear previous sort indicators
    headers.forEach(h => {{
        h.classList.remove('sort-asc', 'sort-desc');
    }});

    // Sort rows
    rows.sort((a, b) => {{
        let aVal = a.cells[colIdx].textContent.trim();
        let bVal = b.cells[colIdx].textContent.trim();

        if (type === 'money') {{
            aVal = parseFloat(aVal.replace(/[₦$B M K(),]/g, '')) || 0;
            bVal = parseFloat(bVal.replace(/[₦$B M K(),]/g, '')) || 0;
        }} else if (type === 'pct') {{
            aVal = parseFloat(aVal) || 0;
            bVal = parseFloat(bVal) || 0;
        }} else if (type === 'status') {{
            const statusOrder = {{'On Track': 0, 'Closed': 1, 'At Risk': 2, 'Off Track': 3, 'Unknown': 4}};
            aVal = statusOrder[aVal] ?? 4;
            bVal = statusOrder[bVal] ?? 4;
        }} else {{
            aVal = aVal.toLowerCase();
            bVal = bVal.toLowerCase();
        }}

        return isAsc ? (aVal > bVal ? 1 : -1) : (aVal < bVal ? 1 : -1);
    }});

    // Re-append rows and set sort indicator
    rows.forEach(row => tbody.appendChild(row));
    header.classList.add(isAsc ? 'sort-desc' : 'sort-asc');
}}

function showTab(id) {{
    document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
    document.getElementById('tab-'+id).classList.add('active');
    event.target.classList.add('active');
}}

// Deep dive search
const searchBox = document.getElementById('searchBox');
const deepDiveTable = document.getElementById('deepDiveTable');
const allRows = deepDiveTable.innerHTML;

searchBox?.addEventListener('input', function() {{
    const query = this.value.toLowerCase();
    const rows = deepDiveTable.querySelectorAll('tr');
    rows.forEach(row => {{
        const text = row.textContent.toLowerCase();
        row.style.display = text.includes(query) ? '' : 'none';
    }});
}});

const C = {{p:'#00D4AA',s:'#4ECDC4',a:'#FFE66D',d:'#FF6B6B',g:'rgba(0,212,170,0.08)'}};
const baseOpts = {{
    responsive:true,
    maintainAspectRatio:true,
    plugins:{{legend:{{labels:{{color:'#94a3b8',font:{{family:"'Inter',sans-serif",size:11}}}}}}}},
    scales:{{x:{{grid:{{color:C.g}},ticks:{{color:'#64748b',maxRotation:45}}}},y:{{grid:{{color:C.g}},ticks:{{color:'#64748b',callback:function(v){{if(v>=1e9)return'\\u20A6'+(v/1e9).toFixed(1)+'B';if(v>=1e6)return'\\u20A6'+(v/1e6).toFixed(0)+'M';if(v>=1e3)return'\\u20A6'+(v/1e3).toFixed(0)+'K';return'\\u20A6'+v}}}}}}}}
}};

new Chart(document.getElementById('budgetChart'),{{
    type:'bar',
    data:{{
        labels:{json.dumps(chart_labels)},
        datasets:[
            {{label:'YTD Budget Pace',data:{json.dumps(chart_ytd_budget)},backgroundColor:'rgba(78,205,196,0.7)',borderRadius:4}},
            {{label:'YTD Actual Spend',data:{json.dumps(chart_ytd_actual)},backgroundColor:'rgba(0,212,170,0.7)',borderRadius:4}}
        ]
    }},
    options:{{...baseOpts,indexAxis:'y',scales:{{x:baseOpts.scales.y,y:baseOpts.scales.x}}}}
}});

// Weekly trend
const weeks = {json.dumps(trend_weeks)};
const weekly_spend = {json.dumps(weekly_spends)};
const weekly_pace = {total_budget} / 52;

new Chart(document.getElementById('trendChart'),{{
    type:'line',
    data:{{
        labels:weeks,
        datasets:[
            {{label:'Weekly Operational Spend',data:weekly_spend,borderColor:C.p,backgroundColor:'rgba(0,212,170,0.08)',borderWidth:2.5,fill:true,tension:0.3,pointRadius:4,pointBackgroundColor:C.p}},
            {{label:'Weekly Budget Pace',data:Array({weeks_elapsed}).fill(weekly_pace),borderColor:C.d,borderWidth:2.5,borderDash:[5,5],fill:false,tension:0.3,pointRadius:0}}
        ]
    }},
    options:baseOpts
}});

</script>
<div class="dashboard-footer">
<span>Seamfix Financial Intelligence &nbsp;·&nbsp; Powered by Claude Cowork</span>
<span>Available Data Range: {first_report_date} &ndash; {latest_date} &nbsp;&bull;&nbsp; Generated: {generated_at} &nbsp;&bull;&nbsp; $1 = ₦1,450</span>
</div>
</body>
</html>"""

    with open(output_path, 'w') as f:
        f.write(html)


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else "."
    if not os.path.isdir(folder):
        print(f"Error: {folder} not found")
        sys.exit(1)

    # Find budget file
    budget_files = glob.glob(os.path.join(folder, "*BUDGET*.xlsx")) + glob.glob(os.path.join(folder, "*budget*.xlsx"))
    if not budget_files:
        print("Error: No budget file (*BUDGET*.xlsx) found")
        sys.exit(1)

    budget_file = budget_files[0]
    print(f"Budget file: {os.path.basename(budget_file)}")

    # Find cash report files
    cash_reports = glob.glob(os.path.join(folder, "Cash Report as at*.xlsx"))
    if not cash_reports:
        print("Error: No cash report files found")
        sys.exit(1)

    print(f"Found {len(cash_reports)} cash reports")

    # Extract budget
    budget_items = extract_budget(budget_file)
    if not budget_items:
        print("Error: Could not extract budget data")
        sys.exit(1)

    print(f"Extracted {len(budget_items)} budget items")

    # Extract reports
    reports = []
    for fn in sorted(cash_reports):
        print(f"  Processing: {os.path.basename(fn)}...", end=' ')
        rec = extract_report(fn)
        if rec:
            reports.append(rec)
            print(f"OK ({rec['date_str']})")
        else:
            print("SKIPPED")

    reports.sort(key=lambda x: x['date'])

    if not reports:
        print("Error: No reports extracted")
        sys.exit(1)

    print(f"\nExtracted {len(reports)} weeks\n")

    # Generate HTML
    out_path = os.path.join(folder, 'budget_dashboard.html')
    generate_html(budget_items, reports, out_path)
    print(f"Dashboard: {out_path}")


if __name__ == '__main__':
    main()
