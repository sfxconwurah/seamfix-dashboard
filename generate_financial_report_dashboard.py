#!/usr/bin/env python3
"""
Seamfix Group Financial Report Dashboard Generator
Reads the consolidated Group Financial Report ("Summary" tab) and renders an
executive P&L / profitability dashboard.

Targets the GROUP INCOME STATEMENT (right-hand block of the Summary tab):
  col K = line-item label   col L = current-period NGN    col M = prior-period NGN
  col N = current-period USD col O = prior-period USD      col P = variance %

Critical targets the executive team tracks:
  - Net Profit Margin (profitability)  ... target 10%
  - Gross Profit Margin                ... target 70%

USD figures come straight from the report's own USD columns (the report uses a
period-average FX rate), NOT the dashboard-wide FX_RATE constant.

Usage: python3 generate_financial_report_dashboard.py [folder_path]
Output: <folder>/financial_report_dashboard.html
"""

import os, sys, re, glob
from datetime import datetime
from openpyxl import load_workbook
from theme import get_base_css, get_toggle_html, get_theme_js

# Executive targets
GROSS_MARGIN_TARGET = 70.0   # %
NET_MARGIN_TARGET = 10.0     # %

# Economic Value Added (EVA) assumption — weighted average cost of capital.
# EVA = NOPAT - (Invested Capital x WACC). Finance's board-approved hurdle rate.
# Adjust here if it changes. Invested capital is proxied by Total Assets (the
# Summary tab has no clean debt+equity capital line).
WACC_PCT = 37.0              # %

# Nigeria statutory company income tax rate. Used to normalise NOPAT for EVA
# when the report books little/no tax YTD (a 0% booked rate would overstate
# economic profit). If the report books a real effective rate, that is used
# instead (see eff_tax below).
TAX_RATE = 30.0              # %

# GROUP income-statement columns on the Summary tab
COL_LABEL = 11   # K
COL_CUR_NGN = 12  # L
COL_PRIOR_NGN = 13  # M
COL_CUR_USD = 14  # N
COL_PRIOR_USD = 15  # O


def norm(s):
    """Normalise a label for matching: lowercase, collapse whitespace."""
    return re.sub(r'\s+', ' ', str(s or '').strip().lower())


def sf(val):
    """Safe float conversion (returns 0.0 for blanks / errors / #DIV/0!)."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def fmt_naira(val):
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000_000:
        return f"{sign}\u20a6{v/1_000_000_000:.2f}B"
    elif v >= 1_000_000:
        return f"{sign}\u20a6{v/1_000_000:.1f}M"
    elif v >= 1_000:
        return f"{sign}\u20a6{v/1_000:.0f}K"
    else:
        return f"{sign}\u20a6{v:,.0f}"


def fmt_usd(val):
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000:
        return f"{sign}${v/1_000_000:.2f}M"
    elif v >= 1_000:
        return f"{sign}${v/1_000:.1f}K"
    else:
        return f"{sign}${v:,.0f}"


def fmt_pct(frac):
    """Format a fraction (0.78) as a percentage string."""
    return f"{frac*100:.1f}%"


def yoy_pct(cur, prior):
    """Year-over-year growth as a percentage. Handles sign flips / zero prior."""
    if prior == 0:
        return None
    return (cur - prior) / abs(prior) * 100.0


_MONTHS = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
           'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def _report_period_key(path):
    """Sort key from a report filename so the LATEST period wins.

    Parses a month token + 2/4-digit year anywhere in the name (e.g. 'May-26',
    'Apr-26', '18th Jun-26', 'Jun 2026') plus an optional version suffix ('_v2').
    Returns (year, month, version, mtime). The regex anchors on a real month
    abbreviation (not just any 3 letters) so words like 'Report 18' don't get
    mistaken for a period. Files with no recognisable period fall back to mtime
    only so they never outrank a dated report.
    """
    name = os.path.basename(path)
    mtime = os.path.getmtime(path)
    month_alt = '|'.join(_MONTHS.keys())
    m = re.search(rf'({month_alt})[a-z]*[-_ ]?(\d{{2,4}})', name, re.IGNORECASE)
    if not m:
        return (-1, -1, -1, mtime)
    month = _MONTHS[m.group(1).lower()]
    yr = int(m.group(2))
    year = yr + 2000 if yr < 100 else yr
    v = re.search(r'[_-]v(\d+)', name, re.IGNORECASE)
    version = int(v.group(1)) if v else 0
    return (year, month, version, mtime)


def find_file(folder, pattern):
    """Return the LATEST matching report so weekly drops auto-update the tab."""
    files = glob.glob(os.path.join(folder, f"*{pattern}*"))
    files = [f for f in files if f.lower().endswith('.xlsx') and not os.path.basename(f).startswith('~$')]
    if not files:
        return None
    return max(files, key=_report_period_key)


def extract_financials(report_file):
    """Parse the GROUP income statement + ratios + revenue breakdowns from the Summary tab.

    Returns a dict of metrics, or None if the Summary tab / GROUP block isn't found.
    Parsing is label-driven (column K) so it survives row insertions/deletions.
    """
    wb = load_workbook(report_file, data_only=True)
    if 'Summary' not in wb.sheetnames:
        return None
    ws = wb['Summary']

    # Locate the GROUP BALANCE SHEET header — everything above it is the income
    # statement region we scan for P&L lines, ratios and ARR.
    bs_row = ws.max_row
    for r in range(1, ws.max_row + 1):
        if 'balance sheet' in norm(ws.cell(r, COL_LABEL).value):
            bs_row = r
            break

    def row_vals(r):
        return {
            'ngn': sf(ws.cell(r, COL_CUR_NGN).value),
            'ngn_prior': sf(ws.cell(r, COL_PRIOR_NGN).value),
            'usd': sf(ws.cell(r, COL_CUR_USD).value),
            'usd_prior': sf(ws.cell(r, COL_PRIOR_USD).value),
        }

    # ── Build label → values map for the income-statement region ──
    # Labels in the P&L / ratios block are unique, so a simple first-wins map works.
    pl = {}
    for r in range(1, bs_row):
        label = norm(ws.cell(r, COL_LABEL).value)
        if label and label not in pl:
            pl[r] = label
    by_label = {}
    for r, label in pl.items():
        by_label.setdefault(label, row_vals(r))

    def get(label, default=None):
        return by_label.get(norm(label), default or {'ngn': 0, 'ngn_prior': 0, 'usd': 0, 'usd_prior': 0})

    metrics = {}
    metrics['revenue'] = get('total revenue')
    metrics['cogs'] = get('cost of sales')
    metrics['gross_profit'] = get('gross profit')
    metrics['other_income'] = get('other income')
    metrics['opex'] = get('total operating expenses')
    metrics['ebitda'] = get('ebitda')
    metrics['da'] = get('depreciation & amortisation')
    metrics['ebit'] = get('ebit')
    metrics['interest'] = get('interest expense')
    metrics['pbt'] = get('profit before tax')
    metrics['tax'] = get('tax expense')
    metrics['pat'] = get('profit after tax')

    # Ratios (stored as fractions in the report)
    metrics['gross_margin'] = get('gross profit margin')
    metrics['ebitda_margin'] = get('ebitda margin')
    metrics['net_margin'] = get('net income margin')
    metrics['opex_ratio'] = get('opex to revenue ratio')
    metrics['payroll'] = get('payroll expense')
    metrics['payroll_pct'] = get('payroll % of revenue')
    metrics['marketing'] = get('marketing expenses')
    metrics['marketing_pct'] = get('marketing % of revenue')
    metrics['arr'] = get('arr')
    metrics['arr_pct'] = get('arr (%)')

    # ── Section breakdowns (vertical / customer / country) ──
    def extract_section(header_kw):
        """Read rows after a 'Revenue by X' header until the section's 'Total' row."""
        items = []
        start = None
        for r in range(1, bs_row):
            if header_kw in norm(ws.cell(r, COL_LABEL).value):
                start = r + 1
                break
        if start is None:
            return items
        for r in range(start, bs_row):
            label = ws.cell(r, COL_LABEL).value
            nlabel = norm(label)
            if not nlabel:
                continue
            if nlabel == 'total':
                break
            # stop if we hit the next 'revenue by' / 'ratios' header
            if nlabel.startswith('revenue by') or nlabel == 'ratios' or 'annual recurring' in nlabel:
                break
            v = row_vals(r)
            if v['ngn'] or v['ngn_prior'] or v['usd']:
                items.append({'name': str(label).strip(), **v})
        return items

    metrics['by_vertical'] = extract_section('revenue by vertical')
    metrics['by_customer'] = extract_section('revenue by customer')
    metrics['by_country'] = extract_section('revenue by country')

    # ── Balance-sheet highlights ──
    bs = {}
    for r in range(bs_row, ws.max_row + 1):
        label = norm(ws.cell(r, COL_LABEL).value)
        if label and label not in bs:
            bs[label] = row_vals(r)
    metrics['cash'] = bs.get('cash & cash equivalents', {'ngn': 0, 'usd': 0})
    metrics['receivables'] = bs.get('account receivables', {'ngn': 0, 'usd': 0})
    metrics['total_assets'] = bs.get('total asset', {'ngn': 0, 'usd': 0})
    # Components of invested capital (Current Assets - Current Liabilities + Net Fixed Assets)
    metrics['current_assets'] = bs.get('total current asset', {'ngn': 0, 'usd': 0})
    metrics['current_liabilities'] = bs.get('total current liabilities', {'ngn': 0, 'usd': 0})
    metrics['net_fixed_assets'] = bs.get('total non current asset', {'ngn': 0, 'usd': 0})
    metrics['current_ratio'] = bs.get('current ratio', {'ngn': 0})
    metrics['cash_ratio'] = bs.get('cash ratio', {'ngn': 0})

    # ── Reporting period (dates live in row 6 of the GROUP block) ──
    cur_date = prior_date = None
    for r in range(1, bs_row):
        cv = ws.cell(r, COL_CUR_NGN).value
        pv = ws.cell(r, COL_PRIOR_NGN).value
        if isinstance(cv, datetime) and isinstance(pv, datetime):
            cur_date, prior_date = cv, pv
            break
    metrics['cur_date'] = cur_date
    metrics['prior_date'] = prior_date

    return metrics


def extract_monthly_trend(report_file):
    """Read the 'MoM' tab to build a month-by-month trend of the key P&L lines.

    Row 1 holds month-end date headers (real dates only — the 'FY 25'/'FY 26'
    summary columns are skipped because they aren't datetimes); column C holds
    the line-item labels. Parsing is label-driven so it survives row/column
    shifts. Limited to the CURRENT fiscal year (the latest year with data) so the
    trend tracks this year's months. Returns {'months':[...], 'revenue':[...],
    'gross_profit':[...], 'ebitda':[...], 'net_profit':[...]} (NGN, one entry per
    month with data), or None if the tab is absent/unreadable.
    """
    try:
        wb = load_workbook(report_file, data_only=True)
        if 'MoM' not in wb.sheetnames:
            return None
        ws = wb['MoM']
    except Exception:
        return None

    # Month columns = those whose row-1 header is a real date (skips FY totals/blanks).
    month_cols = [(c, ws.cell(1, c).value) for c in range(1, ws.max_column + 1)
                  if isinstance(ws.cell(1, c).value, datetime)]
    if not month_cols:
        return None

    # Locate each metric row by scanning column C labels (first match wins).
    # NOTE: EBITDA is DERIVED using the SAME formula as the Summary tab —
    # EBITDA = Gross Profit + Other Income - Total Operating Expenses — NOT read
    # from the MoM 'EBITDA' row, which is corrupt in the source workbook (wrong
    # sign/magnitude, e.g. Jan-26 shows -193.7M when true EBITDA is +188.5M). The
    # derived value reconciles exactly to PAT every month. So we read the gross
    # margin, other income and operating-expense rows here.
    want = {'total revenue': 'revenue', 'gross margin': 'gross_profit',
            'other income': 'other_income', 'total operating expenses': 'opex',
            'pat': 'net_profit'}
    row_for = {}
    for r in range(1, ws.max_row + 1):
        label = norm(ws.cell(r, 3).value)
        if label in want and want[label] not in row_for:
            row_for[want[label]] = r
    if 'revenue' not in row_for:
        return None

    rev_r = row_for['revenue']
    # Months that actually have revenue reported.
    live = [(c, dt) for c, dt in month_cols if sf(ws.cell(rev_r, c).value) != 0]
    if not live:
        return None
    # Restrict to the current fiscal year (the latest year with data) so the
    # trend shows this year's months only (auto-advances each year).
    cur_year = max(dt.year for _, dt in live)
    live = [(c, dt) for c, dt in live if dt.year == cur_year]

    months = []
    series = {k: [] for k in ('revenue', 'gross_profit', 'ebitda', 'net_profit')}
    for c, dt in live:
        months.append(dt.strftime('%b %y'))
        gp = sf(ws.cell(row_for['gross_profit'], c).value) if 'gross_profit' in row_for else 0
        oth = sf(ws.cell(row_for['other_income'], c).value) if 'other_income' in row_for else 0
        opex = sf(ws.cell(row_for['opex'], c).value) if 'opex' in row_for else 0
        series['revenue'].append(round(sf(ws.cell(row_for['revenue'], c).value)) if 'revenue' in row_for else 0)
        series['gross_profit'].append(round(gp))
        # EBITDA = Gross Profit + Other Income - OpEx (Summary-tab formula);
        # the MoM 'EBITDA' row is corrupt so it is never read.
        series['ebitda'].append(round(gp + oth - opex))
        series['net_profit'].append(round(sf(ws.cell(row_for['net_profit'], c).value)) if 'net_profit' in row_for else 0)
    if not months:
        return None
    return {'months': months, **series}


def gauge_html(actual_pct, target_pct, good_above=True):
    """Render a 0-100% gauge with the actual fill and a target marker tick."""
    fill = max(0.0, min(actual_pct, 100.0))
    marker = max(0.0, min(target_pct, 100.0))
    meets = (actual_pct >= target_pct) if good_above else (actual_pct <= target_pct)
    fill_color = 'var(--accent)' if meets else 'var(--danger)'
    return (
        f'<div class="t-gauge">'
        f'<div class="t-gauge-fill" style="width:{fill:.1f}%;background:{fill_color}"></div>'
        f'<div class="t-gauge-marker" style="left:{marker:.1f}%" title="Target {target_pct:.0f}%"></div>'
        f'</div>'
    )


def build_html(m, report_file, output_path):
    generated_at = datetime.now().strftime('%d %b %Y %H:%M')

    cur_date = m['cur_date']
    prior_date = m['prior_date']
    if cur_date:
        period_label = f"YTD as at {cur_date.strftime('%d %b %Y')}"
        cur_yr = cur_date.strftime('%b %Y')
    else:
        period_label = "YTD"
        cur_yr = ""
    prior_yr = prior_date.strftime('%b %Y') if prior_date else "prior year"

    rev = m['revenue']
    revenue_yoy = yoy_pct(rev['ngn'], rev['ngn_prior'])

    gm = m['gross_margin']['ngn'] * 100.0       # gross margin %
    nm = m['net_margin']['ngn'] * 100.0         # net income margin %
    ebitda_m = m['ebitda_margin']['ngn'] * 100.0
    opex_r = m['opex_ratio']['ngn'] * 100.0
    payroll_r = m['payroll_pct']['ngn'] * 100.0
    marketing_r = m['marketing_pct']['ngn'] * 100.0
    arr_r = m['arr_pct']['ngn'] * 100.0
    arr_r_prior = m['arr_pct']['ngn_prior'] * 100.0

    # ARR KPI card — mirror the Revenue & Fundability dashboard's "ARR As At"
    # card: the recurring share of revenue actually EARNED YTD vs the 50% target.
    # Falls back to the report's own ARR line if the revenue file is absent.
    arr_ext = m.get('arr_ext')
    if arr_ext:
        _pct = arr_ext['pct']
        _tgt = arr_ext['target_pct']
        _below = _tgt - _pct
        _delta = (f"\u25bc {_below:.0f}pts below target" if _pct < _tgt else "\u25b2 on/above target")
        arr_value_str = f"{_pct:.0f}%"
        arr_value_neg = _pct < _tgt
        arr_secondary_str = f"Target {_tgt:.0f}% &middot; {_delta}"
        arr_change_str = f"{fmt_usd(arr_ext['usd'])} recurring of {fmt_usd(arr_ext['rev_usd'])} earned YTD"
        arr_kpi_label = f"ARR As At {arr_ext['last_month']} 2026"
    else:
        arr_value_str = fmt_usd(m['arr']['usd'])
        arr_secondary_str = f"{fmt_naira(m['arr']['ngn'])} &middot; {arr_r:.0f}% of revenue"
        arr_change_str = f"{prior_yr}: {fmt_usd(m['arr']['usd_prior'])}"
        arr_value_neg = False
        arr_kpi_label = "Annual Recurring Revenue (ARR)"

    gm_delta = gm - GROSS_MARGIN_TARGET
    nm_delta = nm - NET_MARGIN_TARGET

    pat = m['pat']
    pat_turnaround = pat['ngn'] > 0 and pat['ngn_prior'] < 0

    # ── Critical financial ratios ──
    # Cost lines (cogs/opex/da/interest/tax) are stored negative on the Summary
    # tab; subtotals (ebit/ebitda/pbt/pat) are positive. Use abs() on costs.
    rev_n = rev['ngn'] or 1
    rev_np = rev['ngn_prior'] or 1
    ebit_n = m['ebit']['ngn']
    ebit_m = ebit_n / rev_n * 100.0                      # operating margin
    ebit_m_prior = m['ebit']['ngn_prior'] / rev_np * 100.0
    pbt_n = m['pbt']['ngn']
    tax_n = abs(m['tax']['ngn'])
    # Booked effective rate; fall back to the 30% statutory rate when no tax is
    # booked (else NOPAT/EVA would overstate economic profit).
    eff_tax_booked = (tax_n / pbt_n * 100.0) if (pbt_n > 0 and tax_n > 0) else 0.0
    eff_tax = eff_tax_booked if eff_tax_booked > 0 else TAX_RATE
    tax_basis = "booked effective" if eff_tax_booked > 0 else "Nigeria statutory CIT"
    int_n = abs(m['interest']['ngn'])
    int_cover = (ebit_n / int_n) if int_n > 0 else None   # interest coverage (x)

    ta_n = m['total_assets']['ngn'] or 0

    # Invested capital = Current Assets - Current Liabilities + Net Fixed Assets
    # (i.e. net working capital + net fixed/non-current assets). Equivalent to
    # Total Assets - Current Liabilities. Used as the EVA capital base.
    inv_cap_n = m['current_assets']['ngn'] - m['current_liabilities']['ngn'] + m['net_fixed_assets']['ngn']
    inv_cap_u = m['current_assets']['usd'] - m['current_liabilities']['usd'] + m['net_fixed_assets']['usd']

    # Period length (YTD) — annualise period flows for ROA / EVA charge.
    months_elapsed = cur_date.month if cur_date else 12
    year_frac = max(months_elapsed, 1) / 12.0
    roa = (pat['ngn'] / ta_n / year_frac * 100.0) if ta_n else 0.0   # annualised ROA

    # ── Economic Value Added (EVA) ──
    # EVA = NOPAT - capital charge. NOPAT = EBIT x (1 - effective tax rate).
    # Capital charge = Invested Capital x WACC, prorated to the YTD period so it's
    # comparable with period NOPAT.
    nopat_n = ebit_n * (1 - eff_tax / 100.0)
    nopat_u = m['ebit']['usd'] * (1 - eff_tax / 100.0)
    capital_charge_n = inv_cap_n * (WACC_PCT / 100.0) * year_frac
    capital_charge_u = inv_cap_u * (WACC_PCT / 100.0) * year_frac
    eva_n = nopat_n - capital_charge_n
    eva_u = nopat_u - capital_charge_u
    eva_positive = eva_n > 0

    # Return on Invested Capital — annualised NOPAT over invested capital. Value
    # is created when ROIC exceeds WACC (consistent with EVA sign).
    roic = (nopat_n / inv_cap_n / year_frac * 100.0) if inv_cap_n else 0.0

    # Pre-built label strings (Python 3.9 forbids backslashes inside f-string expressions)
    UP, DOWN = '\u25b2', '\u25bc'
    nm_delta_label = (f"{UP} {nm_delta:+.1f}pts above target" if nm >= NET_MARGIN_TARGET
                      else f"{DOWN} {nm_delta:.1f}pts below target")
    gm_delta_label = (f"{UP} {gm_delta:+.1f}pts above target" if gm >= GROSS_MARGIN_TARGET
                      else f"{DOWN} {gm_delta:.1f}pts below target")
    def yoy_badge(val_dict, positive_is_good=True):
        """YoY badge using magnitude growth; flags loss/profit sign flips as turnarounds."""
        cur, prior = val_dict['ngn'], val_dict['ngn_prior']
        if prior == 0:
            return '<span style="color:var(--text-tertiary)">new</span>' if cur else ""
        # Sign flip (e.g. prior loss -> current profit)
        if (cur > 0) != (prior > 0):
            if cur > prior:
                return '<span class="positive">\u21ba turnaround</span>'
            return '<span class="negative">\u21ba reversed</span>'
        mag_y = (abs(cur) - abs(prior)) / abs(prior) * 100.0
        good = (mag_y >= 0) if positive_is_good else (mag_y <= 0)
        arrow = '\u25b2' if mag_y >= 0 else '\u25bc'
        cls = 'positive' if good else 'negative'
        return f'<span class="{cls}">{arrow} {abs(mag_y):.0f}% YoY</span>'

    pat_change = (f"{chr(0x1F680)} from {fmt_naira(pat['ngn_prior'])} loss"
                  if pat_turnaround else yoy_badge(pat))

    # ── Insights ──
    insights = []

    # Gross margin
    if gm >= GROSS_MARGIN_TARGET:
        insights.append(('GOOD', '\u2705', 'Gross margin above target',
                         f"Gross profit margin is <strong>{gm:.1f}%</strong>, "
                         f"<strong>{gm_delta:+.1f}pts</strong> versus the 70% target. "
                         f"Cost of sales is well contained at {fmt_pct(m['cogs']['ngn']/rev['ngn']*-1) if rev['ngn'] else 'n/a'} of revenue."))
    else:
        insights.append(('RISK', '\u26a0\ufe0f', 'Gross margin below target',
                         f"Gross profit margin is <strong>{gm:.1f}%</strong>, "
                         f"<strong>{gm_delta:.1f}pts</strong> short of the 70% target. "
                         f"Review cost of sales and delivery efficiency."))

    # Net margin
    if nm >= NET_MARGIN_TARGET:
        insights.append(('GOOD', '\u2705', 'Profitability above target',
                         f"Net profit margin is <strong>{nm:.1f}%</strong>, "
                         f"<strong>{nm_delta:+.1f}pts</strong> above the 10% profitability target "
                         f"(Net Profit {fmt_naira(pat['ngn'])} / {fmt_usd(pat['usd'])})."))
    else:
        insights.append(('RISK', '\u26a0\ufe0f', 'Profitability below target',
                         f"Net profit margin is <strong>{nm:.1f}%</strong>, "
                         f"<strong>{nm_delta:.1f}pts</strong> below the 10% target."))

    # Turnaround / revenue growth
    if pat_turnaround:
        insights.append(('GOOD', '\U0001F680', 'Loss-to-profit turnaround',
                         f"The group swung to a profit of <strong>{fmt_naira(pat['ngn'])}</strong> "
                         f"({fmt_usd(pat['usd'])}) from a loss of {fmt_naira(pat['ngn_prior'])} in {prior_yr}."))
    if revenue_yoy is not None:
        insights.append(('GOOD' if revenue_yoy >= 0 else 'RISK',
                         '\U0001F4C8' if revenue_yoy >= 0 else '\U0001F4C9',
                         'Revenue growth' if revenue_yoy >= 0 else 'Revenue decline',
                         f"Revenue {'grew' if revenue_yoy >= 0 else 'fell'} <strong>{abs(revenue_yoy):.0f}%</strong> "
                         f"YoY to <strong>{fmt_naira(rev['ngn'])}</strong> ({fmt_usd(rev['usd'])}) vs {prior_yr}."))

    # ARR drop — only when falling back to the report's own ARR line. When ARR
    # is sourced from the Revenue & Fundability sheet there is no prior-year
    # basis, so we skip this to avoid showing a conflicting ARR figure.
    arr_yoy = yoy_pct(m['arr']['ngn'], m['arr']['ngn_prior'])
    if not m.get('arr_ext') and arr_yoy is not None and arr_yoy < -10:
        insights.append(('RISK', '\u26a0\ufe0f', 'Recurring revenue contracted',
                         f"ARR fell <strong>{abs(arr_yoy):.0f}%</strong> YoY to {fmt_usd(m['arr']['usd'])} "
                         f"({fmt_naira(m['arr']['ngn'])}). Recurring revenue is now only <strong>{arr_r:.0f}%</strong> "
                         f"of total revenue (was {arr_r_prior:.0f}% in {prior_yr}) — revenue is increasingly "
                         f"one-off / project-based."))

    # Customer concentration
    customers = sorted(m['by_customer'], key=lambda x: x['ngn'], reverse=True)
    cust_total = sum(c['ngn'] for c in customers) or 1
    if customers:
        top = customers[0]
        conc = top['ngn'] / cust_total * 100.0
        if conc >= 40:
            insights.append(('RISK', '\u26a0\ufe0f', 'Customer concentration risk',
                             f"<strong>{top['name']}</strong> alone is <strong>{conc:.0f}%</strong> "
                             f"of customer revenue ({fmt_naira(top['ngn'])}). Heavy reliance on a single "
                             f"customer is a fundability and continuity risk."))

    # OpEx efficiency
    insights.append(('GOOD' if opex_r < 100 else 'RISK', '\U0001F4B8', 'Operating cost discipline',
                     f"Operating expenses are <strong>{opex_r:.0f}%</strong> of revenue "
                     f"(was {m['opex_ratio']['ngn_prior']*100:.0f}% in {prior_yr}). "
                     f"Payroll is {payroll_r:.0f}% and marketing {marketing_r:.0f}% of revenue."))

    # Economic Value Added (value creation vs cost of capital)
    if eva_positive:
        insights.append(('GOOD', '\U0001F4B0', 'Creating economic value',
                         f"After charging a <strong>{WACC_PCT:.0f}%</strong> cost of capital on "
                         f"{fmt_naira(ta_n)} of assets, the group still generated <strong>{fmt_naira(eva_n)}</strong> "
                         f"of Economic Value Added (EVA) YTD &mdash; returns are above the cost of capital."))
    else:
        insights.append(('RISK', '\u26a0\ufe0f', 'Returns below cost of capital',
                         f"Economic Value Added is <strong>{fmt_naira(eva_n)}</strong> YTD: NOPAT of "
                         f"{fmt_naira(nopat_n)} does not yet cover the {WACC_PCT:.0f}% capital charge "
                         f"({fmt_naira(capital_charge_n)}) on {fmt_naira(ta_n)} of assets."))

    insight_cards = ""
    badge_colors = {'GOOD': ('var(--accent)', 'var(--accent-bg)'),
                    'RISK': ('var(--danger)', 'var(--danger-bg)')}
    for kind, icon, title, body in insights:
        color, bg = badge_colors[kind]
        insight_cards += f"""<div class="takeaway-card" style="border-left-color:{color}">
<div class="tw-header"><span class="tw-icon">{icon}</span><span class="tw-title">{title}</span>
<span class="tw-badge" style="color:{color};background:{bg}">{kind}</span></div>
<div class="tw-body">{body}</div>
</div>"""

    # ── P&L table ──
    def pl_row(label, key, indent=False, bold=False, positive_is_good=True):
        d = m[key]
        ngn = d['ngn']; usd = d['usd']; ngn_p = d['ngn_prior']
        ystr = yoy_badge({'ngn': ngn, 'ngn_prior': ngn_p}, positive_is_good).replace(' YoY', '')
        name = f'<span style="padding-left:{16 if indent else 0}px">{label}</span>'
        weight = 'font-weight:700' if bold else ''
        return (f'<tr style="{weight}"><td>{name}</td>'
                f'<td class="num">{fmt_naira(ngn)}</td>'
                f'<td class="num">{fmt_usd(usd)}</td>'
                f'<td class="num">{fmt_naira(ngn_p)}</td>'
                f'<td class="num">{ystr}</td></tr>')

    pl_table = "".join([
        pl_row('Total Revenue', 'revenue', bold=True),
        pl_row('Cost of Sales', 'cogs', indent=True, positive_is_good=False),
        pl_row('Gross Profit', 'gross_profit', bold=True),
        pl_row('Other Income', 'other_income', indent=True),
        pl_row('Operating Expenses', 'opex', indent=True, positive_is_good=False),
        pl_row('EBITDA', 'ebitda', bold=True),
        pl_row('Depreciation & Amortisation', 'da', indent=True, positive_is_good=False),
        pl_row('EBIT', 'ebit', bold=True),
        pl_row('Interest Expense', 'interest', indent=True, positive_is_good=False),
        pl_row('Profit Before Tax', 'pbt', bold=True),
        pl_row('Tax Expense', 'tax', indent=True, positive_is_good=False),
        pl_row('Net Profit', 'pat', bold=True),
    ])

    # ── Expense breakdown table ──
    rev_ngn = rev['ngn'] or 1
    exp_items = [
        ('Cost of Sales', abs(m['cogs']['ngn']), abs(m['cogs']['usd'])),
        ('Payroll', m['payroll']['ngn'], m['payroll']['usd']),
        ('Marketing & Commercial', m['marketing']['ngn'], m['marketing']['usd']),
        ('Depreciation & Amortisation', abs(m['da']['ngn']), abs(m['da']['usd'])),
    ]
    # "Other operating" = total opex minus payroll & marketing (both are inside opex)
    other_opex = abs(m['opex']['ngn']) - m['payroll']['ngn'] - m['marketing']['ngn']
    other_opex_usd = abs(m['opex']['usd']) - m['payroll']['usd'] - m['marketing']['usd']
    if other_opex > 0:
        exp_items.append(('Other Operating Expenses', other_opex, other_opex_usd))
    exp_items.sort(key=lambda x: x[1], reverse=True)
    exp_rows = ""
    for name, ngn, usd in exp_items:
        pct = ngn / rev_ngn * 100.0
        exp_rows += (f'<tr><td>{name}</td><td class="num">{fmt_naira(ngn)}</td>'
                     f'<td class="num">{fmt_usd(usd)}</td><td class="num">{pct:.1f}%</td></tr>')

    # ── Revenue breakdown tables ──
    def breakdown_rows(items, total_key='ngn'):
        rows = sorted(items, key=lambda x: x[total_key], reverse=True)
        tot = sum(i[total_key] for i in rows) or 1
        out = ""
        for it in rows:
            share = it[total_key] / tot * 100.0
            y = yoy_pct(it['ngn'], it['ngn_prior'])
            ystr = f'<span class="{ "positive" if (y or 0) >= 0 else "negative" }">{y:+.0f}%</span>' if y is not None else '<span style="color:var(--text-tertiary)">new</span>'
            out += (f'<tr><td>{it["name"]}</td><td class="num">{fmt_naira(it["ngn"])}</td>'
                    f'<td class="num">{fmt_usd(it["usd"])}</td><td class="num">{share:.0f}%</td>'
                    f'<td class="num">{ystr}</td></tr>')
        return out

    vertical_rows = breakdown_rows(m['by_vertical'])
    country_rows = breakdown_rows(m['by_country'])
    customer_rows = breakdown_rows(customers[:8])

    # Chart data (verticals)
    v_sorted = sorted([i for i in m['by_vertical'] if i['ngn'] > 0], key=lambda x: x['ngn'], reverse=True)
    v_labels = [i['name'] for i in v_sorted]
    v_values = [round(i['ngn']) for i in v_sorted]

    # ── Monthly performance trend (from the MoM tab) ──
    trend = m.get('trend')
    if trend and trend['months']:
        t_months = trend['months']
        t_rev = [round(v / 1e6, 1) for v in trend['revenue']]
        t_gp = [round(v / 1e6, 1) for v in trend['gross_profit']]
        t_eb = [round(v / 1e6, 1) for v in trend['ebitda']]
        t_np = [round(v / 1e6, 1) for v in trend['net_profit']]
        trend_section = (
            '<div class="chart-box" style="margin-bottom:28px">'
            f'<h3>Monthly Performance Trend &mdash; {t_months[0]} to {t_months[-1]} (NGN millions)</h3>'
            '<div style="position:relative;height:340px"><canvas id="trendChart"></canvas></div>'
            '</div>')
        trend_js = f"""
new Chart(document.getElementById('trendChart'), {{
  type:'line',
  data:{{labels:{t_months!r}, datasets:[
    {{label:'Revenue', data:{t_rev!r}, borderColor:'#009E7E', backgroundColor:'rgba(0,158,126,.08)', borderWidth:2, tension:.3, fill:true, pointRadius:2}},
    {{label:'Gross Profit', data:{t_gp!r}, borderColor:'#00D4AA', borderWidth:2, tension:.3, pointRadius:2}},
    {{label:'EBITDA', data:{t_eb!r}, borderColor:'#FFB020', borderWidth:2, tension:.3, pointRadius:2}},
    {{label:'Net Profit', data:{t_np!r}, borderColor:'#FF6B6B', borderWidth:2, tension:.3, pointRadius:2}}
  ]}},
  options:{{responsive:true, maintainAspectRatio:false,
    interaction:{{mode:'index', intersect:false}},
    plugins:{{legend:{{position:'top', labels:{{color:'#475569', font:{{size:11}}, usePointStyle:true}}}},
      tooltip:{{callbacks:{{label:function(c){{return c.dataset.label+': \u20a6'+c.parsed.y.toFixed(0)+'M';}}}}}}}},
    scales:{{y:{{ticks:{{color:'#94a3b8', callback:function(v){{return '\u20a6'+v+'M';}}}}, grid:{{color:'rgba(148,163,184,.12)'}}}},
      x:{{ticks:{{color:'#94a3b8'}}, grid:{{display:false}}}}}}
  }}
}});"""
    else:
        trend_section = ''
        trend_js = ''

    # ── Key financial ratios grid ──
    gpm_prior = m['gross_margin']['ngn_prior'] * 100.0
    ebitda_prior = m['ebitda_margin']['ngn_prior'] * 100.0
    nm_prior = m['net_margin']['ngn_prior'] * 100.0
    opex_prior = m['opex_ratio']['ngn_prior'] * 100.0
    payroll_prior = m['payroll_pct']['ngn_prior'] * 100.0
    marketing_prior = m['marketing_pct']['ngn_prior'] * 100.0
    cur_ratio = m['current_ratio']['ngn']
    cash_ratio = m['cash_ratio']['ngn']
    int_cover_str = f"{int_cover:.1f}x" if int_cover else "n/a"
    int_cover_good = (int_cover >= 3) if int_cover else None

    ratios = [
        ('Profitability', 'Gross Margin', f"{gm:.1f}%", f"Target {GROSS_MARGIN_TARGET:.0f}% \u00b7 prior {gpm_prior:.1f}%", gm >= GROSS_MARGIN_TARGET),
        ('Profitability', 'EBITDA Margin', f"{ebitda_m:.1f}%", f"Prior {ebitda_prior:.1f}%", ebitda_m > 0),
        ('Profitability', 'Operating Margin (EBIT)', f"{ebit_m:.1f}%", f"Prior {ebit_m_prior:.1f}%", ebit_m > 0),
        ('Profitability', 'Net Profit Margin', f"{nm:.1f}%", f"Target {NET_MARGIN_TARGET:.0f}% \u00b7 prior {nm_prior:.1f}%", nm >= NET_MARGIN_TARGET),
        ('Returns', 'Return on Assets', f"{roa:.1f}%", "Annualised \u00b7 Net Profit \u00f7 total assets", roa > 0),
        ('Returns', 'Return on Invested Capital', f"{roic:.1f}%", f"Annualised \u00b7 NOPAT \u00f7 invested capital \u00b7 vs {WACC_PCT:.0f}% WACC", roic >= WACC_PCT),
        ('Efficiency', 'OpEx / Revenue', f"{opex_r:.0f}%", f"Prior {opex_prior:.0f}% \u00b7 lower is better", opex_r < 100),
        ('Efficiency', 'Payroll / Revenue', f"{payroll_r:.0f}%", f"Prior {payroll_prior:.0f}%", None),
        ('Efficiency', 'Marketing / Revenue', f"{marketing_r:.0f}%", f"Prior {marketing_prior:.0f}%", None),
        ('Efficiency', 'Effective Tax Rate', f"{eff_tax:.0f}%", tax_basis, None),
        ('Liquidity', 'Interest Coverage', int_cover_str, "EBIT \u00f7 interest \u00b7 \u22653x healthy", int_cover_good),
        ('Liquidity', 'Current Ratio', f"{cur_ratio:.2f}x", "\u22651.0x healthy", cur_ratio >= 1),
        ('Liquidity', 'Cash Ratio', f"{cash_ratio:.2f}x", "Cash \u00f7 current liabilities", None),
    ]
    ratio_cards = ""
    for cat, label, value, sub, good in ratios:
        vcls = '' if good is None else ('ratio-good' if good else 'ratio-bad')
        ratio_cards += (f'<div class="ratio-card"><div class="ratio-cat">{cat}</div>'
                        f'<div class="ratio-label">{label}</div>'
                        f'<div class="ratio-value {vcls}">{value}</div>'
                        f'<div class="ratio-sub">{sub}</div></div>')

    # ── EVA card text ──
    eva_sub = (f"Economic profit after a {WACC_PCT:.0f}% cost of capital on "
               f"{fmt_naira(inv_cap_n)} invested capital "
               f"(current assets &minus; current liabilities + net fixed assets; "
               f"YTD {months_elapsed}mo, annualised charge)")
    eva_break = (f"NOPAT {fmt_naira(nopat_n)} (EBIT after {eff_tax:.0f}% {tax_basis} tax) "
                 f"&minus; capital charge {fmt_naira(capital_charge_n)}")

    theme_css = get_base_css()
    theme_toggle = get_toggle_html()
    theme_js = get_theme_js()

    nav = ('<nav class="top-nav"><span class="top-nav-brand">\u26a1 Seamfix</span>'
           '<a href="dashboard.html" class="top-nav-link ">Cash Overview</a>'
           '<a href="expense_dashboard.html" class="top-nav-link ">Expense &amp; Vendor</a>'
           '<a href="budget_dashboard.html" class="top-nav-link ">Budget vs Actual</a>'
           '<a href="revenue_dashboard.html" class="top-nav-link ">Revenue &amp; Fundability</a>'
           '<a href="pipeline_dashboard.html" class="top-nav-link ">Pipeline Intelligence</a>'
           '<a href="financial_report_dashboard.html" class="top-nav-link active">Group Financials</a></nav>')

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Seamfix Group Financial Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
{theme_css}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:var(--bg-body);color:var(--text-primary);min-height:100vh}}
.container{{max-width:1600px;margin:0 auto;padding:0 28px 28px}}
.header{{padding:24px 28px 16px;border-bottom:1px solid var(--border-main);margin-bottom:24px}}
.header h1{{font-size:22px;font-weight:700;color:var(--text-primary);margin-bottom:4px}}
.header .sub{{font-size:13px;color:var(--text-secondary)}}
.header .meta{{font-size:12px;color:var(--text-secondary);margin-top:4px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;margin-bottom:28px}}
.kpi-card{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:22px;transition:all .3s}}
.kpi-card:hover{{transform:translateY(-3px);box-shadow:var(--shadow-hover)}}
.kpi-label{{font-size:.78em;color:var(--text-secondary);text-transform:uppercase;letter-spacing:1.2px;margin-bottom:8px;font-weight:600}}
.kpi-value{{font-size:1.4em;font-weight:700;color:var(--accent);margin-bottom:6px}}
.kpi-value.negative{{color:var(--danger)}}
.kpi-secondary{{font-size:.85em;color:var(--text-tertiary);margin-bottom:4px}}
.kpi-change{{font-size:.82em;color:var(--text-secondary)}}
.target-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(360px,1fr));gap:20px;margin-bottom:28px}}
.target-card{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:24px}}
.target-card h3{{font-size:.85em;text-transform:uppercase;letter-spacing:1px;color:var(--text-secondary);margin-bottom:6px;font-weight:600}}
.target-big{{font-size:2.2em;font-weight:700;margin-bottom:2px}}
.target-big.good{{color:var(--accent)}}
.target-big.bad{{color:var(--danger)}}
.target-sub{{font-size:.85em;color:var(--text-tertiary);margin-bottom:14px}}
.t-gauge{{position:relative;width:100%;height:18px;background:var(--bg-gauge);border-radius:9px;overflow:hidden;margin-bottom:6px}}
.t-gauge-fill{{height:100%;border-radius:9px;transition:width .4s}}
.t-gauge-marker{{position:absolute;top:-3px;width:3px;height:24px;background:var(--text-primary);opacity:.7}}
.t-legend{{display:flex;justify-content:space-between;font-size:.72em;color:var(--text-tertiary)}}
.ratio-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:14px;margin-bottom:20px}}
.ratio-card{{background:var(--bg-card);border:1px solid var(--border-light);border-radius:10px;padding:16px}}
.ratio-cat{{font-size:.62em;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:var(--text-tertiary);margin-bottom:6px}}
.ratio-label{{font-size:.8em;color:var(--text-secondary);margin-bottom:6px}}
.ratio-value{{font-size:1.5em;font-weight:700;color:var(--text-heading);margin-bottom:4px}}
.ratio-value.ratio-good{{color:var(--accent)}}
.ratio-value.ratio-bad{{color:var(--danger)}}
.ratio-sub{{font-size:.72em;color:var(--text-tertiary)}}
.eva-card{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:24px;margin-bottom:28px;border-left:4px solid var(--accent)}}
.eva-card.eva-neg{{border-left-color:var(--danger)}}
.eva-card h3{{font-size:.85em;text-transform:uppercase;letter-spacing:1px;color:var(--text-secondary);margin-bottom:6px;font-weight:600}}
.eva-big{{font-size:2.2em;font-weight:700;margin-bottom:2px}}
.eva-big.good{{color:var(--accent)}}.eva-big.bad{{color:var(--danger)}}
.eva-sub{{font-size:.85em;color:var(--text-tertiary);margin-bottom:8px}}
.eva-break{{font-size:.82em;color:var(--text-secondary)}}
.section{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:24px;margin-bottom:24px}}
.section h2{{font-size:1.2em;margin-bottom:16px;color:var(--accent)}}
.grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:24px;align-items:start}}
table{{width:100%;border-collapse:collapse;font-size:.85em}}
thead{{background:var(--bg-table-header)}}
th{{padding:9px 12px;text-align:left;font-weight:600;color:var(--accent);border-bottom:2px solid var(--border-accent)}}
th.num,td.num{{text-align:right}}
td{{padding:9px 12px;border-bottom:1px solid var(--border-light);color:var(--text-heading)}}
tbody tr:hover{{background:var(--bg-table-hover)}}
.positive{{color:var(--accent)}}
.negative{{color:var(--danger)}}
.takeaways-section{{margin-bottom:28px}}
.takeaways-section h2{{font-size:1.3em;margin-bottom:16px;color:var(--warning);display:flex;align-items:center;gap:10px}}
.takeaways-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(420px,1fr));gap:16px}}
.takeaway-card{{background:var(--bg-card);border:1px solid var(--border-light);border-left:4px solid var(--text-tertiary);border-radius:10px;padding:20px;transition:all .3s}}
.takeaway-card:hover{{transform:translateY(-2px);box-shadow:var(--shadow-hover)}}
.tw-header{{display:flex;align-items:center;gap:8px;margin-bottom:10px;flex-wrap:wrap}}
.tw-icon{{font-size:1.2em}}
.tw-title{{font-size:.82em;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:var(--text-secondary)}}
.tw-badge{{font-size:.7em;font-weight:700;padding:3px 10px;border-radius:12px;letter-spacing:.5px}}
.tw-body{{font-size:.88em;color:var(--text-secondary);line-height:1.65}}
.chart-box{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:22px;min-height:340px}}
.chart-box h3{{font-size:1em;font-weight:600;margin-bottom:16px;color:var(--text-heading)}}
.top-nav{{background:var(--bg-nav);border-bottom:1px solid var(--border-main);padding:0 24px;display:flex;align-items:center;height:48px;overflow-x:auto;position:sticky;top:0;z-index:200}}
.top-nav-brand{{color:var(--text-primary);font-weight:700;font-size:15px;margin-right:24px;white-space:nowrap}}
.top-nav-link{{color:var(--text-secondary);text-decoration:none;padding:0 14px;height:48px;display:flex;align-items:center;font-size:13px;border-bottom:2px solid transparent;white-space:nowrap}}
.top-nav-link:hover{{color:var(--text-primary)}}
.top-nav-link.active{{color:var(--text-primary);border-bottom-color:var(--accent);font-weight:500}}
.dashboard-footer{{margin-top:48px;padding:20px 28px;border-top:1px solid var(--border-main);color:var(--text-secondary);font-size:12px;display:flex;justify-content:space-between;flex-wrap:wrap;gap:8px}}
@media(max-width:1024px){{.kpi-grid{{grid-template-columns:repeat(2,1fr)}}.grid-2{{grid-template-columns:1fr}}.takeaways-grid{{grid-template-columns:1fr}}}}
@media(max-width:640px){{.kpi-grid{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
{theme_toggle}
{nav}
<div class="container">
<div class="header">
<h1>Seamfix Group Financial Report</h1>
<div class="sub">Consolidated profitability &amp; performance &mdash; {period_label} vs {prior_yr} (NGN with USD equivalents)</div>
<div class="meta">Source: {os.path.basename(report_file)} &nbsp;&bull;&nbsp; Generated: {generated_at}</div>
</div>

<div class="kpi-grid">
<div class="kpi-card">
<div class="kpi-label">Total Revenue</div>
<div class="kpi-value">{fmt_naira(rev['ngn'])}</div>
<div class="kpi-secondary">{fmt_usd(rev['usd'])}</div>
<div class="kpi-change">{yoy_badge(rev)} vs {prior_yr}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Gross Profit</div>
<div class="kpi-value">{fmt_naira(m['gross_profit']['ngn'])}</div>
<div class="kpi-secondary">{fmt_usd(m['gross_profit']['usd'])}</div>
<div class="kpi-change">{gm:.1f}% margin {yoy_badge(m['gross_profit'])}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">EBITDA</div>
<div class="kpi-value {'negative' if m['ebitda']['ngn'] < 0 else ''}">{fmt_naira(m['ebitda']['ngn'])}</div>
<div class="kpi-secondary">{fmt_usd(m['ebitda']['usd'])}</div>
<div class="kpi-change">{ebitda_m:.1f}% margin {yoy_badge(m['ebitda'])}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Net Profit</div>
<div class="kpi-value {'negative' if pat['ngn'] < 0 else ''}">{fmt_naira(pat['ngn'])}</div>
<div class="kpi-secondary">{fmt_usd(pat['usd'])}</div>
<div class="kpi-change">{pat_change}</div>
</div>
<div class="kpi-card">
<div class="kpi-label">{arr_kpi_label}</div>
<div class="kpi-value {'negative' if arr_value_neg else ''}">{arr_value_str}</div>
<div class="kpi-secondary">{arr_secondary_str}</div>
<div class="kpi-change">{arr_change_str}</div>
</div>
</div>

<div class="target-grid">
<div class="target-card">
<h3>Net Profit Margin &middot; Profitability</h3>
<div class="target-big {'good' if nm >= NET_MARGIN_TARGET else 'bad'}">{nm:.1f}%</div>
<div class="target-sub">Target {NET_MARGIN_TARGET:.0f}% &middot; {nm_delta_label} &middot; {prior_yr}: {m['net_margin']['ngn_prior']*100:.1f}%</div>
{gauge_html(nm, NET_MARGIN_TARGET)}
<div class="t-legend"><span>0%</span><span>Target {NET_MARGIN_TARGET:.0f}%</span><span>100%</span></div>
</div>
<div class="target-card">
<h3>Gross Profit Margin</h3>
<div class="target-big {'good' if gm >= GROSS_MARGIN_TARGET else 'bad'}">{gm:.1f}%</div>
<div class="target-sub">Target {GROSS_MARGIN_TARGET:.0f}% &middot; {gm_delta_label} &middot; {prior_yr}: {m['gross_margin']['ngn_prior']*100:.1f}%</div>
{gauge_html(gm, GROSS_MARGIN_TARGET)}
<div class="t-legend"><span>0%</span><span>Target {GROSS_MARGIN_TARGET:.0f}%</span><span>100%</span></div>
</div>
</div>

{trend_section}

<div class="section">
<h2>Key Financial Ratios &mdash; {period_label}</h2>
<div class="ratio-grid">{ratio_cards}</div>
</div>

<div class="eva-card {'' if eva_positive else 'eva-neg'}">
<h3>Economic Value Added (EVA) &middot; Value Creation</h3>
<div class="eva-big {'good' if eva_positive else 'bad'}">{fmt_naira(eva_n)}</div>
<div class="eva-sub">{fmt_usd(eva_u)} &middot; {eva_sub}</div>
<div class="eva-break">{eva_break} &nbsp;|&nbsp; {'Returns exceed the cost of capital &mdash; value created.' if eva_positive else 'Returns do not yet cover the cost of capital &mdash; value eroded.'}</div>
</div>

<div class="takeaways-section">
<h2>\U0001F4A1 Critical Insights</h2>
<div class="takeaways-grid">
{insight_cards}
</div>
</div>

<div class="section">
<h2>Income Statement &mdash; {period_label}</h2>
<table>
<thead><tr><th>Line Item</th><th class="num">{cur_yr} (NGN)</th><th class="num">{cur_yr} (USD)</th><th class="num">{prior_yr} (NGN)</th><th class="num">YoY</th></tr></thead>
<tbody>{pl_table}</tbody>
</table>
</div>

<div class="grid-2">
<div class="section">
<h2>Expense Analysis</h2>
<table>
<thead><tr><th>Cost Category</th><th class="num">NGN</th><th class="num">USD</th><th class="num">% of Revenue</th></tr></thead>
<tbody>{exp_rows}</tbody>
</table>
<p style="font-size:.8em;color:var(--text-tertiary);margin-top:12px">OpEx-to-Revenue: <strong>{opex_r:.0f}%</strong> (was {m['opex_ratio']['ngn_prior']*100:.0f}% in {prior_yr}) &middot; Payroll: <strong>{payroll_r:.0f}%</strong> of revenue &middot; Marketing: <strong>{marketing_r:.0f}%</strong> of revenue.</p>
</div>
<div class="chart-box">
<h3>Revenue by Vertical (NGN)</h3>
<div style="position:relative;height:300px"><canvas id="verticalChart"></canvas></div>
</div>
</div>

<div class="grid-2">
<div class="section">
<h2>Revenue by Vertical</h2>
<table>
<thead><tr><th>Vertical</th><th class="num">NGN</th><th class="num">USD</th><th class="num">Share</th><th class="num">YoY</th></tr></thead>
<tbody>{vertical_rows}</tbody>
</table>
</div>
<div class="section">
<h2>Revenue by Country</h2>
<table>
<thead><tr><th>Country</th><th class="num">NGN</th><th class="num">USD</th><th class="num">Share</th><th class="num">YoY</th></tr></thead>
<tbody>{country_rows}</tbody>
</table>
</div>
</div>

<div class="section">
<h2>Top Customers</h2>
<table>
<thead><tr><th>Customer</th><th class="num">NGN</th><th class="num">USD</th><th class="num">Share</th><th class="num">YoY</th></tr></thead>
<tbody>{customer_rows}</tbody>
</table>
</div>

<div class="section">
<h2>Balance Sheet &amp; Liquidity</h2>
<table>
<thead><tr><th>Item</th><th class="num">NGN</th><th class="num">USD</th></tr></thead>
<tbody>
<tr><td>Cash &amp; Cash Equivalents</td><td class="num">{fmt_naira(m['cash']['ngn'])}</td><td class="num">{fmt_usd(m['cash']['usd'])}</td></tr>
<tr><td>Account Receivables</td><td class="num">{fmt_naira(m['receivables']['ngn'])}</td><td class="num">{fmt_usd(m['receivables']['usd'])}</td></tr>
<tr><td>Total Assets</td><td class="num">{fmt_naira(m['total_assets']['ngn'])}</td><td class="num">{fmt_usd(m['total_assets']['usd'])}</td></tr>
<tr><td>Current Ratio</td><td class="num">{cur_ratio:.2f}x</td><td class="num">&mdash;</td></tr>
<tr><td>Cash Ratio</td><td class="num">{cash_ratio:.2f}x</td><td class="num">&mdash;</td></tr>
</tbody>
</table>
</div>

</div>
<script>
const vData = {{labels: {v_labels!r}, values: {v_values!r}}};
const palette = ['#009E7E','#3BA89E','#00D4AA','#4ECDC4','#FFE66D','#FF6B6B','#b45309'];
new Chart(document.getElementById('verticalChart'), {{
  type: 'doughnut',
  data: {{labels: vData.labels, datasets: [{{data: vData.values, backgroundColor: palette, borderWidth: 0}}]}},
  options: {{responsive:true, maintainAspectRatio:false, plugins:{{legend:{{position:'right', labels:{{color:'#475569', font:{{size:11}}}}}}, tooltip:{{callbacks:{{label: function(c){{var t=c.dataset.data.reduce((a,b)=>a+b,0); var p=t?(c.parsed/t*100).toFixed(0):0; return c.label+': \u20a6'+(c.parsed/1e6).toFixed(0)+'M ('+p+'%)';}}}}}}}}}}
}});
{trend_js}
{theme_js}
</script>
<div class="dashboard-footer">
<span>Seamfix Financial Intelligence &nbsp;\u00b7&nbsp; Powered by Claude Cowork</span>
<span>Generated: {generated_at}</span>
</div>
</body>
</html>"""

    with open(output_path, 'w') as f:
        f.write(html)


def write_placeholder(output_path, msg):
    """Write a friendly 'no data' page so the tab never crashes the app."""
    theme_css = get_base_css()
    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Seamfix Group Financial Report</title><style>{theme_css}
body{{font-family:'Inter',sans-serif;background:var(--bg-body);color:var(--text-primary);
display:flex;align-items:center;justify-content:center;min-height:90vh;text-align:center;padding:24px}}
.box{{max-width:560px}}.box h1{{font-size:1.4em;color:var(--accent);margin-bottom:12px}}
.box p{{color:var(--text-secondary);line-height:1.6}}</style></head>
<body><div class="box"><h1>Group Financial Report</h1><p>{msg}</p></div></body></html>"""
    with open(output_path, 'w') as f:
        f.write(html)


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else "./data"
    if not os.path.isdir(folder):
        print(f"Error: {folder} not found")
        sys.exit(1)

    output_path = os.path.join(folder, "financial_report_dashboard.html")

    report_file = find_file(folder, "Group Financial")
    if not report_file:
        report_file = find_file(folder, "Financial Report")
    if not report_file:
        print("Group Financial Report file not found — writing placeholder.")
        write_placeholder(output_path,
                          "The Group Financial Report has not been provided for this period. "
                          "Drop the latest <em>Group Financial Report</em> .xlsx into the data folder to populate this dashboard.")
        sys.exit(0)

    print(f"Financial report file: {report_file}")
    try:
        metrics = extract_financials(report_file)
    except Exception as e:
        print(f"Error parsing report: {e}")
        write_placeholder(output_path, f"Could not parse the Group Financial Report: {e}")
        sys.exit(0)

    if not metrics or metrics['revenue']['ngn'] == 0:
        print("Could not locate the GROUP income statement on the Summary tab — writing placeholder.")
        write_placeholder(output_path,
                          "The Group Financial Report was found but its <strong>Summary</strong> tab "
                          "did not contain a recognisable GROUP income statement.")
        sys.exit(0)

    # ARR sourced from the Revenue & Fundability dashboard's underlying data so
    # both tabs read from the same place: the ACTUAL recurring revenue EARNED
    # year-to-date (sum of monthly actuals, cols N-Y) for deals flagged
    # "Recurring" in the Path to Revenue sheet (x FX_RATE for NGN). This is the
    # actuals-based figure (not the planned/contracted annual target in col F).
    # Falls back to the report's own ARR line if the revenue file isn't present.
    try:
        import generate_revenue_dashboard as gen_rev
        rev_file = gen_rev.find_file(folder, "Path to Revenue")
        if rev_file:
            revenues = gen_rev.extract_revenue_data(rev_file)
            recurring = [r for r in revenues if r.get('recurring')]
            arr_actual_usd = sum(r['ytd_actual'] for r in recurring)
            total_ytd_usd = sum(r['ytd_actual'] for r in revenues)
            arr_pct = (arr_actual_usd / total_ytd_usd * 100.0) if total_ytd_usd > 0 else 0.0
            monthly_totals = [sum(r['monthly'][i] for r in revenues) for i in range(12)]
            with_data = [i for i, v in enumerate(monthly_totals) if v > 0]
            last_m = max(with_data) if with_data else 0
            metrics['arr_ext'] = {
                'usd': arr_actual_usd,
                'ngn': arr_actual_usd * gen_rev.FX_RATE,
                'rev_usd': total_ytd_usd,
                'pct': arr_pct,
                'target_pct': 50,  # mirrors ARR_TARGET_PCT in generate_revenue_dashboard
                'last_month': gen_rev.MONTH_NAMES[last_m],
                'count': len(recurring),
                'total': len(revenues),
            }
            print(f"  ARR As At {metrics['arr_ext']['last_month']} (recurring % of revenue, actual YTD): "
                  f"{arr_pct:.0f}% vs 50% target ({fmt_usd(arr_actual_usd)} of {fmt_usd(total_ytd_usd)})")
    except Exception as e:
        print(f"  ARR sync skipped ({e}) — falling back to report ARR line")

    # Month-by-month trend (from the 'MoM' tab) for the trend chart.
    try:
        metrics['trend'] = extract_monthly_trend(report_file)
        if metrics['trend']:
            print(f"  Monthly trend: {len(metrics['trend']['months'])} months "
                  f"({metrics['trend']['months'][0]} \u2192 {metrics['trend']['months'][-1]})")
    except Exception as e:
        metrics['trend'] = None
        print(f"  Monthly trend skipped ({e})")

    build_html(metrics, report_file, output_path)
    print(f"Generated: {output_path}")
    print(f"  Revenue: {fmt_naira(metrics['revenue']['ngn'])} ({fmt_usd(metrics['revenue']['usd'])})")
    print(f"  Gross margin: {metrics['gross_margin']['ngn']*100:.1f}% (target {GROSS_MARGIN_TARGET:.0f}%)")
    print(f"  Net margin: {metrics['net_margin']['ngn']*100:.1f}% (target {NET_MARGIN_TARGET:.0f}%)")
    print(f"  PAT: {fmt_naira(metrics['pat']['ngn'])}")


if __name__ == "__main__":
    main()
