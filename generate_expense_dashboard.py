#!/usr/bin/env python3
"""
Seamfix Expense & Vendor Analysis Dashboard Generator

Builds a detailed expense/vendor analysis dashboard from weekly cash reports.
Extracts category-level outflows and payment batch vendor details.

Usage: python3 generate_expense_dashboard.py /path/to/xlsx/folder
"""

import os
import sys
import json
import re
from datetime import datetime
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import load_workbook
from theme import get_base_css, get_toggle_html, get_theme_js


# Category grouping rules
CATEGORY_GROUPS = {
    "Salary & Benefits": [
        "Salary", "Salaries", "Exit Salary", "Management Allowance",
        "Pension", "Pension Remittance", "PAYE", "Paye", "NSITF Contributions"
    ],
    "Vendor Payments": ["Payment Batches"],
    "Software & Tools": ["Software Tools"],
    "Travel & Accommodation": ["Flight Ticket/Accommodation", "BTA", "Hotel/Acc"],
    "Bank & Financial": ["Bank Charges", "Interbank"],
    "Tax & Compliance": ["VAT", "WHT", "VAT Remitted", "WHT remitted", "Business Premises & Dev Levy"],
    "Office Operations": ["Office Supplies & Fuel", "Office Supplies Fuel & Letter Dispatch"],
    "Contractors & Consultants": ["Tatvasoft", "Bind Creative", "Amazon", "Nimc consultancy fee"],
    "Events & Conferences": ["ID4Africa", "ID4Africa Registration", "MWC Registration", "Training"],
    "Investment Outflows": ["Seamfix UAE funding", "Investment in SCB USD mutua fund"],
}


def safe_row_dict(row):
    """Extract column values from a row, handling merged cells."""
    d = {}
    for c in row:
        try:
            d[c.column_letter] = c.value
        except:
            pass
    return d


def parse_filename_date(filename):
    """Parse date from filename like 'Cash Report as at 13th February 2026.xlsx'"""
    match = re.search(r'(\d+)\w*\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', filename)
    if match:
        day, month_name, year = match.groups()
        months = {
            'January': 1, 'February': 2, 'March': 3, 'April': 4,
            'May': 5, 'June': 6, 'July': 7, 'August': 8,
            'September': 9, 'October': 10, 'November': 11, 'December': 12
        }
        month = months.get(month_name, 1)
        return datetime(int(year), month, int(day))
    return None


def get_standardized_category(original_category):
    """Map original category to standardized group."""
    if not original_category:
        return "Other"
    original = str(original_category).strip()
    original_lower = original.lower()
    # Exact match first
    for group, items in CATEGORY_GROUPS.items():
        if original in items:
            return group
    # Keyword-based fallback for investment categories
    # This catches new investment vehicles (e.g. "Investment in ARM MMF")
    if any(kw in original_lower for kw in ['investment in', 'funding']):
        return "Investment Outflows"
    return "Other"


def extract_expenses(file_path):
    """Extract category and vendor expenses from a single xlsx file."""
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb["Breakdown of Income and Expendi"]
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None, None, None

    # Parse date from filename
    filename = os.path.basename(file_path)
    report_date = parse_filename_date(filename)
    if not report_date:
        print(f"Could not parse date from {filename}")
        return None, None, None

    categories = {}
    vendors = []

    # Find OUTFLOWS section (row with "OUTFLOWS" in column B)
    outflow_start = None
    total_outflow_row = None

    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100, values_only=False), 1):
        b_val = row[1].value if len(row) > 1 else None
        if b_val and "OUTFLOWS" in str(b_val).upper():
            outflow_start = row_idx
        if b_val and "TOTAL CASH OUTFLOW" in str(b_val).upper():
            total_outflow_row = row_idx
            break

    if outflow_start and total_outflow_row:
        for row_idx in range(outflow_start + 1, total_outflow_row):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
            row_dict = safe_row_dict(row)
            b_val = row_dict.get('B')
            c_val = row_dict.get('C')

            # Skip "Net Cash" rows
            if b_val and "Net Cash" in str(b_val):
                continue

            if b_val and c_val:
                try:
                    amount = float(c_val) if isinstance(c_val, (int, float)) else 0
                    if amount > 0:
                        std_category = get_standardized_category(b_val)
                        categories[str(b_val).strip()] = {
                            "amount": amount,
                            "standardized": std_category
                        }
                except:
                    pass

    # Find BREAKDOWN OF PAYMENT BATCH section
    breakdown_start = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=200, values_only=False), 1):
        g_val = row[6].value if len(row) > 6 else None
        if g_val and "BREAKDOWN OF PAYMENT BATCH" in str(g_val).upper():
            breakdown_start = row_idx
            break

    if breakdown_start:
        for row_idx in range(breakdown_start + 1, breakdown_start + 100):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=False))[0]
            row_dict = safe_row_dict(row)
            g_val = row_dict.get('G')
            h_val = row_dict.get('H')
            i_val = row_dict.get('I')

            # Stop at Total row or empty vendor name
            if not g_val or "TOTAL" in str(g_val).upper():
                break

            # Skip header row
            if g_val and "BENEFICIARY" in str(g_val).upper():
                continue

            if g_val:
                try:
                    amount = float(h_val) if isinstance(h_val, (int, float)) else 0
                    if amount > 0:
                        vendors.append({
                            "name": str(g_val).strip(),
                            "amount": amount,
                            "description": str(i_val).strip() if i_val else "",
                            "date": report_date
                        })
                except:
                    pass

    wb.close()
    return report_date, categories, vendors


def process_all_files(folder_path):
    """Process all xlsx files in a folder."""
    xlsx_files = sorted(
        [f for f in os.listdir(folder_path) if f.endswith('.xlsx')],
        key=lambda x: parse_filename_date(x) or datetime.min
    )

    all_weekly_data = []  # List of (date, categories, vendors)
    all_vendors = {}  # vendor_name -> list of payments
    all_categories_by_week = defaultdict(dict)  # date -> {category -> amount}

    for xlsx_file in xlsx_files:
        file_path = os.path.join(folder_path, xlsx_file)
        report_date, categories, vendors = extract_expenses(file_path)

        if report_date and categories and vendors:
            all_weekly_data.append((report_date, categories, vendors))

            # Aggregate categories by week
            for cat_name, cat_data in categories.items():
                std_cat = cat_data["standardized"]
                all_categories_by_week[report_date][std_cat] = \
                    all_categories_by_week[report_date].get(std_cat, 0) + cat_data["amount"]

            # Aggregate vendor payments
            for vendor_payment in vendors:
                vendor_name = vendor_payment["name"]
                if vendor_name not in all_vendors:
                    all_vendors[vendor_name] = []
                all_vendors[vendor_name].append(vendor_payment)

    return all_weekly_data, all_vendors, all_categories_by_week


def calculate_kpis(all_weekly_data, all_vendors):
    """Calculate key performance indicators."""
    total_ytd = 0
    for report_date, categories, vendors in all_weekly_data:
        for vendor_payment in vendors:
            total_ytd += vendor_payment["amount"]

    unique_vendors = len(all_vendors)

    weeks = len(all_weekly_data)
    avg_burn = total_ytd / weeks if weeks > 0 else 0

    # Top vendor by total spend
    top_vendor = None
    top_vendor_spend = 0
    for vendor_name, payments in all_vendors.items():
        vendor_total = sum(p["amount"] for p in payments)
        if vendor_total > top_vendor_spend:
            top_vendor_spend = vendor_total
            top_vendor = vendor_name

    # Recurring vendors (3+ appearances)
    recurring_count = sum(1 for vendor_name, payments in all_vendors.items() if len(payments) >= 3)

    # Largest single payment
    largest_payment = 0
    for vendor_name, payments in all_vendors.items():
        for payment in payments:
            if payment["amount"] > largest_payment:
                largest_payment = payment["amount"]

    return {
        "total_ytd": total_ytd,
        "unique_vendors": unique_vendors,
        "avg_burn": avg_burn,
        "top_vendor": top_vendor,
        "top_vendor_spend": top_vendor_spend,
        "recurring_count": recurring_count,
        "largest_payment": largest_payment,
        "weeks": weeks
    }


def format_naira(value):
    """Format value as Naira with appropriate suffix."""
    if value >= 1_000_000_000:
        return f"₦{value / 1_000_000_000:.1f}B"
    elif value >= 1_000_000:
        return f"₦{value / 1_000_000:.1f}M"
    elif value >= 1_000:
        return f"₦{value / 1_000:.0f}K"
    else:
        return f"₦{value:.0f}"


def generate_html(all_weekly_data, all_vendors, all_categories_by_week, kpis, output_file):
    """Generate the HTML dashboard."""

    # Prepare data for charts
    weeks_list = sorted(all_categories_by_week.keys())
    week_labels = [d.strftime("%d %b") for d in weeks_list]

    # Date range for header
    earliest_date = weeks_list[0].strftime("%d %B %Y") if weeks_list else "N/A"
    latest_date = weeks_list[-1].strftime("%d %B %Y") if weeks_list else "N/A"
    generated_at = datetime.now().strftime("%d %B %Y at %H:%M")

    theme_css = get_base_css()
    toggle_html = get_toggle_html()
    theme_js = get_theme_js()

    # Expense category trend data
    all_std_categories = set()
    for week_categories in all_categories_by_week.values():
        all_std_categories.update(week_categories.keys())
    all_std_categories = sorted(list(all_std_categories))

    category_trend_datasets = []
    colors = ["#00D4AA", "#FF6B6B", "#4ECDC4", "#FFE66D", "#A8E6CF", "#95E1D3", "#F38181", "#AA96DA"]

    for idx, category in enumerate(all_std_categories):
        data = [all_categories_by_week.get(week, {}).get(category, 0) for week in weeks_list]
        category_trend_datasets.append({
            "label": category,
            "data": data,
            "backgroundColor": colors[idx % len(colors)],
            "borderColor": colors[idx % len(colors)],
            "borderWidth": 0
        })

    # Top 15 vendors
    vendor_totals = [(name, sum(p["amount"] for p in payments))
                     for name, payments in all_vendors.items()]
    vendor_totals.sort(key=lambda x: x[1], reverse=True)
    top_15 = vendor_totals[:15]

    top_vendors_labels = [v[0] for v in top_15]
    top_vendors_data = [v[1] for v in top_15]

    # Vendor frequency
    vendor_freq = Counter([len(all_vendors[name]) for name in all_vendors.keys()])
    freq_labels = [f"{k}x" for k in sorted(vendor_freq.keys())]
    freq_data = [vendor_freq[k] for k in sorted(vendor_freq.keys())]

    # Weekly volatility
    weekly_totals = []
    for week in weeks_list:
        week_total = sum(all_categories_by_week[week].values())
        weekly_totals.append(week_total)

    # Calculate moving average (3-week)
    moving_avg = []
    for i in range(len(weekly_totals)):
        start = max(0, i - 1)
        end = min(len(weekly_totals), i + 2)
        avg = sum(weekly_totals[start:end]) / (end - start)
        moving_avg.append(avg)

    # Category pie chart
    ytd_by_category = defaultdict(float)
    for week_categories in all_categories_by_week.values():
        for cat, amount in week_categories.items():
            ytd_by_category[cat] += amount

    pie_labels = sorted(ytd_by_category.keys())
    pie_data = [ytd_by_category[cat] for cat in pie_labels]

    # Vendor ledger (all vendors)
    vendor_ledger = []
    for vendor_name, payments in sorted(all_vendors.items(),
                                       key=lambda x: sum(p["amount"] for p in x[1]),
                                       reverse=True):
        total = sum(p["amount"] for p in payments)
        avg = total / len(payments) if payments else 0
        last_payment_date = max(p["date"] for p in payments).strftime("%d %b %Y")
        last_description = payments[-1]["description"] if payments else ""
        recurring = "Yes" if len(payments) >= 3 else "No"

        vendor_ledger.append({
            "name": vendor_name,
            "total": total,
            "count": len(payments),
            "avg": avg,
            "last_date": last_payment_date,
            "last_desc": last_description,
            "recurring": recurring
        })

    # Recurring vendors
    recurring_vendors = [v for v in vendor_ledger if v["recurring"] == "Yes"]

    # Large payments (>5M)
    large_payments = []
    for vendor_name, payments in all_vendors.items():
        for payment in payments:
            if payment["amount"] > 5_000_000:
                large_payments.append({
                    "date": payment["date"].strftime("%d %b %Y"),
                    "vendor": vendor_name,
                    "amount": payment["amount"],
                    "description": payment["description"]
                })
    large_payments.sort(key=lambda x: x["amount"], reverse=True)

    # Executive takeaways
    # Use sum of OPERATIONAL categories as the denominator — exclude "Investment Outflows"
    # which are asset transfers (not operational expenses). Including them would understate
    # salary %, vendor concentration, and bank charges as a share of real operating spend.
    total_all_categories = sum(v for k, v in ytd_by_category.items() if k != "Investment Outflows") or 1

    top_5_spend = sum(v[1] for v in vendor_totals[:5])
    vendor_concentration = (top_5_spend / total_all_categories * 100)

    salary_cat_total = ytd_by_category.get("Salary & Benefits", 0)
    salary_pct = (salary_cat_total / total_all_categories * 100)

    bank_charges_total = ytd_by_category.get("Bank & Financial", 0)
    bank_pct = (bank_charges_total / total_all_categories * 100)

    # Check for large payments this week (last week in data)
    large_payment_alert = False
    if all_weekly_data:
        last_week_date, last_categories, last_vendors = all_weekly_data[-1]
        for payment in last_vendors:
            if payment["amount"] > 10_000_000:
                large_payment_alert = True
                break

    # Generate HTML
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Seamfix Expense & Vendor Analysis Dashboard</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        {theme_css}
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Inter', sans-serif;
            background: var(--bg-body);
            color: var(--text-primary);
            min-height: 100vh;
            padding: 0;
            transition: background 0.3s, color 0.3s;
        }}

        .container {{
            max-width: 1600px;
            margin: 0 auto;
            padding: 0 28px 28px;
        }}

        .header {{
            margin-bottom: 40px;
        }}

        .header h1 {{
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 10px;
            color: var(--text-primary);
        }}

        .header p {{
            color: var(--text-secondary);
            font-size: 0.95rem;
        }}

        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }}

        .kpi-card {{
            background: var(--bg-card);
            border: 1px solid var(--border-main);
            border-radius: 12px;
            padding: 24px;
        }}

        .kpi-card.highlight {{
            border-color: var(--accent);
            background: var(--accent-bg);
        }}

        .kpi-label {{
            font-size: 0.85rem;
            color: var(--text-secondary);
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 10px;
            font-weight: 600;
        }}

        .kpi-value {{
            font-size: 1.8rem;
            font-weight: 700;
            color: var(--accent);
            margin-bottom: 8px;
        }}

        .kpi-detail {{
            font-size: 0.8rem;
            color: var(--text-tertiary);
        }}

        .charts-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
            gap: 30px;
            margin-bottom: 40px;
        }}

        .chart-card {{
            background: var(--bg-card);
            border: 1px solid var(--border-main);
            border-radius: 12px;
            padding: 24px;
        }}

        .chart-card h3 {{
            font-size: 1.1rem;
            margin-bottom: 20px;
            color: var(--text-primary);
            font-weight: 600;
        }}

        .chart-container {{
            position: relative;
            height: 400px;
        }}

        .table-card {{
            background: var(--bg-card);
            border: 1px solid var(--border-main);
            border-radius: 12px;
            padding: 24px;
            margin-bottom: 30px;
        }}

        .table-card h3 {{
            font-size: 1.2rem;
            margin-bottom: 20px;
            color: var(--text-primary);
            font-weight: 600;
        }}

        .search-box {{
            margin-bottom: 20px;
        }}

        .search-box input {{
            width: 100%;
            max-width: 400px;
            padding: 10px 15px;
            background: var(--bg-input);
            border: 1px solid var(--border-main);
            border-radius: 8px;
            color: var(--text-primary);
            font-size: 0.9rem;
        }}

        .search-box input::placeholder {{
            color: var(--text-tertiary);
        }}

        .search-box input:focus {{
            outline: none;
            border-color: var(--accent);
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.9rem;
        }}

        th {{
            background: var(--bg-table-header);
            padding: 12px 15px;
            text-align: left;
            font-weight: 600;
            color: var(--text-heading);
            border-bottom: 1px solid var(--border-main);
        }}

        td {{
            padding: 12px 15px;
            border-bottom: 1px solid var(--border-light);
        }}

        tr:hover {{
            background: var(--bg-table-hover);
        }}

        .badge {{
            display: inline-block;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.75rem;
            font-weight: 600;
            text-transform: uppercase;
        }}

        .badge-yes {{
            background: var(--accent-bg);
            color: var(--accent);
        }}

        .badge-no {{
            background: var(--danger-bg);
            color: var(--danger);
        }}

        .amount {{
            color: var(--accent);
            font-weight: 600;
        }}

        .takeaway-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
            margin-bottom: 40px;
        }}

        .takeaway-card {{
            background: var(--bg-card);
            border: 1px solid var(--border-main);
            border-radius: 12px;
            padding: 0;
        }}

        .takeaway-card h4 {{
            font-size: 0.95rem;
            color: var(--text-heading);
            margin-bottom: 10px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}

        .takeaway-value {{
            font-size: 1.5rem;
            color: var(--warning);
            font-weight: 700;
            margin-bottom: 8px;
        }}

        .takeaway-insight {{
            font-size: 0.85rem;
            color: var(--text-secondary);
            line-height: 1.5;
        }}

        .alert {{
            background: var(--danger-bg);
            border-left: 3px solid var(--danger);
            padding: 12px 15px;
            border-radius: 6px;
            margin-bottom: 15px;
            color: var(--danger);
            font-size: 0.9rem;
        }}

        .hidden {{
            display: none;
        }}

        .footer {{
            text-align: center;
            color: var(--text-tertiary);
            font-size: 0.85rem;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid var(--border-main);
        }}

        .print-button{{display:none!important}}
        .top-nav{{background:var(--bg-nav);border-bottom:1px solid var(--border-main);padding:0 24px;display:flex;align-items:center;height:48px;overflow-x:auto;position:sticky;top:0;z-index:200}}
.top-nav-brand{{color:var(--text-primary);font-weight:700;font-size:15px;margin-right:24px;white-space:nowrap;text-decoration:none}}
.top-nav-link{{color:var(--text-secondary);text-decoration:none;padding:0 14px;height:48px;display:flex;align-items:center;font-size:13px;border-bottom:2px solid transparent;white-space:nowrap;transition:color .2s}}
.top-nav-link:hover{{color:var(--text-primary)}}
        .top-nav-link.active{{color:var(--text-primary);border-bottom-color:var(--accent);font-weight:500}}

        @media (max-width: 768px) {{
            .charts-grid {{
                grid-template-columns: 1fr;
            }}

            .kpi-grid {{
                grid-template-columns: 1fr;
            }}

            .header h1 {{
                font-size: 1.8rem;
            }}

            table {{
                font-size: 0.8rem;
            }}

            th, td {{
                padding: 8px 10px;
            }}
        }}

        .nav-bar{{display:none!important}}

        th.sortable {{
            cursor: pointer;
            user-select: none;
        }}

        th.sortable:hover {{
            color: var(--warning);
        }}

        th.sort-asc::after {{
            content: ' ▲';
            font-size: 0.7em;
        }}

        th.sort-desc::after {{
            content: ' ▼';
            font-size: 0.7em;
        }}

        @media print {{
            * {{
                color-adjust: exact !important;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }}

            body {{
                background: #fff !important;
                color: #1e293b !important;
                padding: 12px !important;
                font-size: 9pt !important;
            }}

            .container {{
                max-width: 100%;
            }}

            .print-button {{
                display: none !important;
            }}

            .nav-bar {{
                display: none !important;
            }}

            .header h1 {{
                font-size: 1.6em !important;
                background: none !important;
                -webkit-text-fill-color: #0f172a !important;
                color: #0f172a !important;
            }}

            .header p, .header .sub {{
                color: #475569 !important;
            }}

            .header .meta {{
                color: #64748b !important;
            }}

            /* KPI Cards */
            .kpi-grid {{
                grid-template-columns: repeat(3, 1fr) !important;
                gap: 8px !important;
                margin-bottom: 16px !important;
            }}

            .kpi-card {{
                background: #f8fafc !important;
                border: 1px solid #e2e8f0 !important;
                padding: 12px !important;
                border-radius: 8px !important;
                box-shadow: none !important;
                backdrop-filter: none !important;
                page-break-inside: avoid;
            }}

            .kpi-card.highlight {{
                border-color: #10b981 !important;
                background: #f0fdf4 !important;
            }}

            .kpi-label {{
                color: #475569 !important;
                font-size: 0.65em !important;
            }}

            .kpi-value {{
                color: #0f172a !important;
                font-size: 1.2em !important;
            }}

            .kpi-detail {{
                color: #64748b !important;
            }}

            /* Takeaway Cards */
            .takeaway-grid {{
                grid-template-columns: repeat(2, 1fr) !important;
                gap: 10px !important;
                margin-bottom: 16px !important;
            }}

            .takeaway-card {{
                background: #f8fafc !important;
                border: 1px solid #e2e8f0 !important;
                padding: 14px !important;
                border-radius: 8px !important;
                backdrop-filter: none !important;
                page-break-inside: avoid;
            }}

            .takeaway-card h4 {{
                color: #475569 !important;
            }}

            .takeaway-value {{
                color: #0f172a !important;
                font-size: 1.3em !important;
            }}

            .takeaway-insight {{
                color: #475569 !important;
            }}

            /* Alert */
            .alert {{
                background: #fef2f2 !important;
                border-left-color: #dc2626 !important;
                color: #dc2626 !important;
            }}

            /* Charts */
            .charts-grid {{
                grid-template-columns: repeat(2, 1fr) !important;
                gap: 12px !important;
                margin-bottom: 16px !important;
            }}

            .chart-card {{
                background: #fff !important;
                border: 1px solid #e2e8f0 !important;
                padding: 14px !important;
                backdrop-filter: none !important;
                page-break-inside: avoid;
            }}

            .chart-card h3 {{
                color: #1e293b !important;
                font-size: 0.85em !important;
            }}

            .chart-container {{
                height: 280px !important;
            }}

            /* Tables */
            .table-card {{
                background: #f8fafc !important;
                border: 1px solid #e2e8f0 !important;
                padding: 16px !important;
                backdrop-filter: none !important;
                page-break-inside: avoid;
            }}

            .table-card h3 {{
                color: #0f172a !important;
                font-size: 1em !important;
            }}

            .search-box {{
                display: none !important;
            }}

            table {{
                font-size: 0.75em !important;
                color: #1e293b !important;
            }}

            thead {{
                background: #f1f5f9 !important;
            }}

            th {{
                background: #f1f5f9 !important;
                color: #0f172a !important;
                border-bottom: 1px solid #cbd5e1 !important;
            }}

            td {{
                color: #334155 !important;
                border-bottom: 1px solid #e2e8f0 !important;
            }}

            .amount {{
                color: #0f172a !important;
                font-weight: 700 !important;
            }}

            .badge {{
                font-size: 0.7em !important;
            }}

            .badge-yes {{
                background: #dcfce7 !important;
                color: #16a34a !important;
            }}

            .badge-no {{
                background: #fee2e2 !important;
                color: #dc2626 !important;
            }}

            tbody tr:hover {{
                background: transparent !important;
            }}

            .hidden {{
                display: none !important;
            }}

            /* Footer */
            .footer {{
                color: #64748b !important;
                border-top-color: #e2e8f0 !important;
            }}

            #themeToggle {{
                display: none !important;
            }}
        }}
    /* ── HEADER: match Pipeline Intelligence style ── */
.header{{padding:24px 28px 16px!important;border-bottom:1px solid var(--border-main)!important;margin-bottom:24px!important;background:none!important}}
.header h1{{font-size:22px!important;font-weight:700!important;background:none!important;-webkit-background-clip:unset!important;-webkit-text-fill-color:var(--text-primary)!important;color:var(--text-primary)!important;margin-bottom:4px!important}}
.header .sub{{font-size:13px!important;color:var(--text-secondary)!important}}
.header .meta{{font-size:12px!important;color:var(--text-secondary)!important;margin-top:4px!important}}
.header p{{font-size:13px!important;color:var(--text-secondary)!important;margin-top:4px!important}}
</style>
</head>
<body>
<nav class="top-nav"><span class="top-nav-brand">⚡ Seamfix</span><a href="dashboard.html" class="top-nav-link ">Cash Overview</a><a href="expense_dashboard.html" class="top-nav-link active">Expense &amp; Vendor</a><a href="budget_dashboard.html" class="top-nav-link ">Budget vs Actual</a><a href="revenue_dashboard.html" class="top-nav-link ">Revenue &amp; Fundability</a><a href="pipeline_dashboard.html" class="top-nav-link ">Pipeline Intelligence</a>{toggle_html}</nav>
    <div class="container">
        <div class="header">
            <h1>Seamfix Expense & Vendor Analysis</h1>
            <p class="sub">Detailed Outflow & Vendor Tracking &mdash; Weekly Cash Reports 2026</p>
            <p class="meta">Data as of: <strong>{latest_date}</strong> &nbsp;&bull;&nbsp; Generated: {generated_at}</p>
        </div>

        <!-- old nav-bar replaced -->

        <!-- PDF button hidden per user request -->

        <!-- KPI Cards -->
        <div class="kpi-grid">
            <div class="kpi-card highlight">
                <div class="kpi-label">Total YTD Expenses</div>
                <div class="kpi-value">{format_naira(kpis['total_ytd'])}</div>
                <div class="kpi-detail">{kpis['weeks']} weeks of data</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Unique Vendors Paid</div>
                <div class="kpi-value">{kpis['unique_vendors']}</div>
                <div class="kpi-detail">Across all payment batches</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Avg Weekly Burn Rate</div>
                <div class="kpi-value">{format_naira(kpis['avg_burn'])}</div>
                <div class="kpi-detail">Average per week</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Top Vendor by Spend</div>
                <div class="kpi-value" style="color: var(--warning); font-size: 1.3rem;">{kpis['top_vendor']}</div>
                <div class="kpi-detail">Total: {format_naira(kpis['top_vendor_spend'])}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Recurring Vendors</div>
                <div class="kpi-value">{kpis['recurring_count']}</div>
                <div class="kpi-detail">Appeared 3+ weeks</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-label">Largest Single Payment</div>
                <div class="kpi-value">{format_naira(kpis['largest_payment'])}</div>
                <div class="kpi-detail">YTD maximum</div>
            </div>
        </div>

        <!-- Executive Takeaways -->
        <div class="takeaway-grid">
            <div class="takeaway-card">
                <h4>Vendor Concentration</h4>
                <div class="takeaway-value">{vendor_concentration:.1f}%</div>
                <div class="takeaway-insight">of total vendor spend goes to top 5 vendors</div>
            </div>
            <div class="takeaway-card">
                <h4>Salary & Benefits</h4>
                <div class="takeaway-value">{salary_pct:.1f}%</div>
                <div class="takeaway-insight">of total outflows (YTD)</div>
            </div>
            <div class="takeaway-card">
                <h4>Bank & Financial Costs</h4>
                <div class="takeaway-value">{bank_pct:.2f}%</div>
                <div class="takeaway-insight">of total outflows</div>
            </div>
            <div class="takeaway-card">
                <h4>Vendor Sprawl</h4>
                <div class="takeaway-value">{kpis['unique_vendors']}</div>
                <div class="takeaway-insight">unique vendors across {kpis['weeks']} weeks</div>
            </div>
        </div>

        {('<div class="alert"><strong>Alert:</strong> Large payments >10M detected in latest week</div>' if large_payment_alert else '')}

        <!-- Charts -->
        <div class="charts-grid">
            <div class="chart-card">
                <h3>Expense Category Trend</h3>
                <div class="chart-container">
                    <canvas id="categoryTrendChart"></canvas>
                </div>
            </div>
            <div class="chart-card">
                <h3>Top 15 Vendors by Total Spend</h3>
                <div class="chart-container">
                    <canvas id="topVendorsChart"></canvas>
                </div>
            </div>
            <div class="chart-card">
                <h3>Vendor Frequency Distribution</h3>
                <div class="chart-container">
                    <canvas id="vendorFreqChart"></canvas>
                </div>
            </div>
            <div class="chart-card">
                <h3>Weekly Expense Volatility</h3>
                <div class="chart-container">
                    <canvas id="volatilityChart"></canvas>
                </div>
            </div>
            <div class="chart-card" style="grid-column: span 1;">
                <h3>YTD Spend by Category</h3>
                <div class="chart-container">
                    <canvas id="categoryPieChart"></canvas>
                </div>
            </div>
        </div>

        <!-- Vendor Ledger -->
        <div class="table-card">
            <h3>Complete Vendor Ledger</h3>
            <div class="search-box">
                <input type="text" id="vendorSearch" placeholder="Search vendor name..." onkeyup="filterTable('vendorTable', 'vendorSearch')">
            </div>
            <table id="vendorTable">
                <thead>
                    <tr>
                        <th class="sortable" onclick="sortTable('vendorTable',0,'text')">Vendor Name</th>
                        <th class="sortable" onclick="sortTable('vendorTable',1,'money')">Total YTD</th>
                        <th class="sortable" onclick="sortTable('vendorTable',2,'number')">Payments</th>
                        <th class="sortable" onclick="sortTable('vendorTable',3,'money')">Avg Payment</th>
                        <th class="sortable" onclick="sortTable('vendorTable',4,'text')">Last Payment</th>
                        <th class="sortable" onclick="sortTable('vendorTable',5,'text')">Last Description</th>
                        <th class="sortable" onclick="sortTable('vendorTable',6,'text')">Recurring?</th>
                    </tr>
                </thead>
                <tbody>
"""

    for vendor in vendor_ledger:
        html += f"""                    <tr>
                        <td>{vendor['name']}</td>
                        <td class="amount">{format_naira(vendor['total'])}</td>
                        <td>{vendor['count']}</td>
                        <td class="amount">{format_naira(vendor['avg'])}</td>
                        <td>{vendor['last_date']}</td>
                        <td>{vendor['last_desc'][:50]}...</td>
                        <td><span class="badge {'badge-yes' if vendor['recurring'] == 'Yes' else 'badge-no'}">{vendor['recurring']}</span></td>
                    </tr>
"""

    html += """                </tbody>
            </table>
        </div>

        <!-- Recurring Vendor Analysis -->
        <div class="table-card">
            <h3>Recurring Vendor Analysis (3+ Payments)</h3>
            <table id="recurringVendorTable">
                <thead>
                    <tr>
                        <th class="sortable" onclick="sortTable('recurringVendorTable',0,'text')">Vendor Name</th>
                        <th class="sortable" onclick="sortTable('recurringVendorTable',1,'money')">Total YTD</th>
                        <th class="sortable" onclick="sortTable('recurringVendorTable',2,'number')">Payments</th>
                        <th class="sortable" onclick="sortTable('recurringVendorTable',3,'money')">Avg Payment</th>
                        <th class="sortable" onclick="sortTable('recurringVendorTable',4,'text')">Frequency</th>
                    </tr>
                </thead>
                <tbody>
"""

    for vendor in recurring_vendors:
        frequency = f"Every {kpis['weeks'] / vendor['count']:.1f} weeks"
        html += f"""                    <tr>
                        <td>{vendor['name']}</td>
                        <td class="amount">{format_naira(vendor['total'])}</td>
                        <td>{vendor['count']}</td>
                        <td class="amount">{format_naira(vendor['avg'])}</td>
                        <td>{frequency}</td>
                    </tr>
"""

    html += """                </tbody>
            </table>
        </div>

        <!-- Large Payments Audit Trail -->
        <div class="table-card">
            <h3>Large Payment Audit Trail (>₦5M)</h3>
            <table id="largePaymentsTable">
                <thead>
                    <tr>
                        <th class="sortable" onclick="sortTable('largePaymentsTable',0,'text')">Date</th>
                        <th class="sortable" onclick="sortTable('largePaymentsTable',1,'text')">Vendor</th>
                        <th class="sortable" onclick="sortTable('largePaymentsTable',2,'money')">Amount</th>
                        <th class="sortable" onclick="sortTable('largePaymentsTable',3,'text')">Description</th>
                    </tr>
                </thead>
                <tbody>
"""

    for payment in large_payments:
        html += f"""                    <tr>
                        <td>{payment['date']}</td>
                        <td>{payment['vendor']}</td>
                        <td class="amount">{format_naira(payment['amount'])}</td>
                        <td>{payment['description']}</td>
                    </tr>
"""

    html += f"""                </tbody>
            </table>
        </div>

        <div class="footer" style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;color:var(--text-secondary);font-size:12px">
            <span>Seamfix Expense & Vendor Analysis &nbsp;·&nbsp; Powered by Claude Cowork</span>
            <span>Available Data Range: {earliest_date} &ndash; {latest_date} &nbsp;&bull;&nbsp; Generated: {generated_at} &nbsp;&bull;&nbsp; $1 = ₦1,450</span>
        </div>
    </div>

    <script>
        // Sort table function
        function sortTable(tableId, colIdx, type) {{
            const table = document.getElementById(tableId);
            const tbody = table.querySelector('tbody');
            if (!tbody) return;

            const rows = Array.from(tbody.querySelectorAll('tr'));
            const headers = table.querySelectorAll('th.sortable');

            // Determine sort direction
            const header = headers[colIdx];
            const isAsc = header.classList.contains('sort-asc');

            // Clear previous sort indicators
            headers.forEach(h => {{
                h.classList.remove('sort-asc', 'sort-desc');
            }});

            // Sort rows
            rows.sort((a, b) => {{
                let aVal = a.cells[colIdx].textContent.trim();
                let bVal = b.cells[colIdx].textContent.trim();

                if (type === 'money') {{
                    aVal = parseFloat(aVal.replace(/[₦$B M K(),\-]/g, '')) || 0;
                    bVal = parseFloat(bVal.replace(/[₦$B M K(),\-]/g, '')) || 0;
                }} else if (type === 'number') {{
                    aVal = parseFloat(aVal) || 0;
                    bVal = parseFloat(bVal) || 0;
                }} else if (type === 'status') {{
                    const statusOrder = {{'On Track': 0, 'Closed': 1, 'At Risk': 2, 'Off Track': 3, 'Unknown': 4}};
                    aVal = statusOrder[aVal] ?? 4;
                    bVal = statusOrder[bVal] ?? 4;
                }} else {{
                    aVal = aVal.toLowerCase();
                    bVal = bVal.toLowerCase();
                }}

                return isAsc ? (aVal > bVal ? 1 : -1) : (aVal < bVal ? 1 : -1);
            }});

            // Re-append rows
            rows.forEach(row => tbody.appendChild(row));
            header.classList.add(isAsc ? 'sort-desc' : 'sort-asc');
        }}

        // Filter table function
        function filterTable(tableId, searchId) {{
            const input = document.getElementById(searchId);
            const filter = input.value.toLowerCase();
            const table = document.getElementById(tableId);
            const rows = table.getElementsByTagName('tr');

            for (let i = 1; i < rows.length; i++) {{
                const text = rows[i].textContent.toLowerCase();
                rows[i].classList.toggle('hidden', !text.includes(filter));
            }}
        }}

        // Category Trend Chart
        const categoryTrendCtx = document.getElementById('categoryTrendChart').getContext('2d');
        new Chart(categoryTrendCtx, {{
            type: 'bar',
            data: {{
                labels: {json.dumps(week_labels)},
                datasets: {json.dumps(category_trend_datasets)}
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                indexAxis: 'x',
                scales: {{
                    x: {{ stacked: true, ticks: {{ color: '#94a3b8' }}, grid: {{ color: 'rgba(100,116,139,0.1)' }} }},
                    y: {{ stacked: true, ticks: {{ color: '#94a3b8' }}, grid: {{ color: 'rgba(100,116,139,0.1)' }} }}
                }},
                plugins: {{
                    legend: {{
                        labels: {{ color: '#94a3b8' }}
                    }}
                }}
            }}
        }});

        // Top Vendors Chart
        const topVendorsCtx = document.getElementById('topVendorsChart').getContext('2d');
        new Chart(topVendorsCtx, {{
            type: 'barH',
            type: 'bar',
            data: {{
                labels: {json.dumps(top_vendors_labels)},
                datasets: [{{
                    label: 'Total Spend',
                    data: {json.dumps(top_vendors_data)},
                    backgroundColor: '#00D4AA',
                    borderColor: '#00D4AA'
                }}]
            }},
            options: {{
                indexAxis: 'y',
                responsive: true,
                maintainAspectRatio: false,
                scales: {{
                    x: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: 'rgba(100,116,139,0.1)' }} }},
                    y: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ display: false }} }}
                }},
                plugins: {{
                    legend: {{ display: false }}
                }}
            }}
        }});

        // Vendor Frequency Chart
        const vendorFreqCtx = document.getElementById('vendorFreqChart').getContext('2d');
        new Chart(vendorFreqCtx, {{
            type: 'bar',
            data: {{
                labels: {json.dumps(freq_labels)},
                datasets: [{{
                    label: 'Number of Vendors',
                    data: {json.dumps(freq_data)},
                    backgroundColor: '#4ECDC4',
                    borderColor: '#4ECDC4'
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                scales: {{
                    x: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: 'rgba(100,116,139,0.1)' }} }},
                    y: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: 'rgba(100,116,139,0.1)' }} }}
                }},
                plugins: {{
                    legend: {{ display: false }}
                }}
            }}
        }});

        // Weekly Volatility Chart
        const volatilityCtx = document.getElementById('volatilityChart').getContext('2d');
        new Chart(volatilityCtx, {{
            type: 'line',
            data: {{
                labels: {json.dumps(week_labels)},
                datasets: [
                    {{
                        label: 'Weekly Total',
                        data: {json.dumps(weekly_totals)},
                        borderColor: '#FF6B6B',
                        backgroundColor: 'rgba(255, 107, 107, 0.1)',
                        tension: 0.4,
                        fill: true,
                        pointBackgroundColor: '#FF6B6B',
                        pointBorderColor: '#FF6B6B'
                    }},
                    {{
                        label: '3-Week Moving Avg',
                        data: {json.dumps(moving_avg)},
                        borderColor: '#FFE66D',
                        backgroundColor: 'transparent',
                        tension: 0.4,
                        borderDash: [5, 5],
                        pointBackgroundColor: '#FFE66D'
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                scales: {{
                    x: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: 'rgba(100,116,139,0.1)' }} }},
                    y: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: 'rgba(100,116,139,0.1)' }} }}
                }},
                plugins: {{
                    legend: {{ labels: {{ color: '#94a3b8' }} }}
                }}
            }}
        }});

        // Category Pie Chart
        const categoryPieCtx = document.getElementById('categoryPieChart').getContext('2d');
        new Chart(categoryPieCtx, {{
            type: 'doughnut',
            data: {{
                labels: {json.dumps(pie_labels)},
                datasets: [{{
                    data: {json.dumps(pie_data)},
                    backgroundColor: ['#00D4AA', '#FF6B6B', '#4ECDC4', '#FFE66D', '#A8E6CF', '#95E1D3', '#F38181', '#AA96DA', '#FCBAD3', '#A8D8EA'],
                    borderColor: 'rgba(30, 41, 59, 0.8)'
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{
                        labels: {{ color: '#94a3b8' }},
                        position: 'bottom'
                    }}
                }}
            }}
        }});
        {theme_js}
    </script>
</body>
</html>
"""

    with open(output_file, 'w') as f:
        f.write(html)

    print(f"Dashboard generated: {output_file}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 generate_expense_dashboard.py /path/to/xlsx/folder")
        sys.exit(1)

    folder_path = sys.argv[1]

    if not os.path.isdir(folder_path):
        print(f"Error: {folder_path} is not a valid directory")
        sys.exit(1)

    print(f"Processing xlsx files in {folder_path}...")
    all_weekly_data, all_vendors, all_categories_by_week = process_all_files(folder_path)

    if not all_weekly_data:
        print("Error: No valid xlsx files found")
        sys.exit(1)

    print(f"Found {len(all_weekly_data)} weeks of data")

    kpis = calculate_kpis(all_weekly_data, all_vendors)
    print(f"\nKPI Summary:")
    print(f"  Total YTD Expenses: {format_naira(kpis['total_ytd'])}")
    print(f"  Unique Vendors: {kpis['unique_vendors']}")
    print(f"  Top Vendor: {kpis['top_vendor']} ({format_naira(kpis['top_vendor_spend'])})")
    print(f"  Recurring Vendors (3+): {kpis['recurring_count']}")

    print(f"\nTop 10 Vendors by Spend:")
    vendor_totals = [(name, sum(p["amount"] for p in payments))
                     for name, payments in all_vendors.items()]
    vendor_totals.sort(key=lambda x: x[1], reverse=True)
    for idx, (name, total) in enumerate(vendor_totals[:10], 1):
        print(f"  {idx}. {name}: {format_naira(total)}")

    output_file = os.path.join(folder_path, "expense_dashboard.html")
    generate_html(all_weekly_data, all_vendors, all_categories_by_week, kpis, output_file)

    # Also save to outputs folder if it exists
    outputs_dir = "/sessions/relaxed-busy-archimedes/mnt/outputs"
    if os.path.isdir(outputs_dir):
        output_copy = os.path.join(outputs_dir, "expense_dashboard.html")
        generate_html(all_weekly_data, all_vendors, all_categories_by_week, kpis, output_copy)
        print(f"Dashboard also saved to: {output_copy}")

    print("\nDashboard generation complete!")


if __name__ == "__main__":
    main()
