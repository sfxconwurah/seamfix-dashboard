#!/usr/bin/env python3
"""
Seamfix Cash Reports Dashboard Generator v2
Reads all .xlsx files from a folder and generates an interactive HTML dashboard.
Includes USD investment portfolio in total cash position (fix: Mar 2026).
Usage: python3 generate_dashboard.py [folder_path]
"""

import os, sys, json, re
from datetime import datetime
from openpyxl import load_workbook
from theme import get_base_css, get_toggle_html, get_theme_js


# --- Investment detection (keyword-based, not exact match) ---
# These patterns catch any category containing investment-related keywords,
# so new investment vehicles (e.g. "Investment in ARM MMF") are auto-detected.
INVESTMENT_OUT_KEYWORDS = ['investment in', 'funding']  # outflow categories
INVESTMENT_IN_KEYWORDS = ['investment withdrawal', 'investment liquidation']  # inflow categories
# Exact matches to always include (even if keywords don't match)
INVESTMENT_OUT_EXACT = ['Seamfix UAE funding']
INVESTMENT_IN_EXACT = []


def is_investment_outflow(category_name):
    """Check if an outflow category is an investment transfer (not operational)."""
    if not category_name:
        return False
    cat_lower = str(category_name).strip().lower()
    if category_name in INVESTMENT_OUT_EXACT:
        return True
    return any(kw in cat_lower for kw in INVESTMENT_OUT_KEYWORDS)


def is_investment_inflow(category_name):
    """Check if an inflow category is an investment liquidation/withdrawal."""
    if not category_name:
        return False
    cat_lower = str(category_name).strip().lower()
    if category_name in INVESTMENT_IN_EXACT:
        return True
    return any(kw in cat_lower for kw in INVESTMENT_IN_KEYWORDS)


def get_investment_outflows(outflow_items):
    """Sum all investment outflows from an outflow_items dict."""
    return sum(v for k, v in outflow_items.items() if is_investment_outflow(k))


def get_investment_inflows(inflow_items):
    """Sum all investment inflows from an inflow_items dict."""
    return sum(v for k, v in inflow_items.items() if is_investment_inflow(k))


def safe_row_dict(row):
    d = {}
    for c in row:
        try:
            d[c.column_letter] = c.value
        except (AttributeError, TypeError):
            pass
    return d


def parse_date(fn):
    m = re.search(r'(\d+)\w*\s+(January|February|March|April|May|June|July|August|September|October|November|December|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})', fn, re.IGNORECASE)
    if not m:
        return None
    day, ms, year = m.groups()
    mm = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
          'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}
    return datetime(int(year), mm[ms.lower()], int(day))


def sf(val):
    if val is None: return 0.0
    try: return float(val)
    except: return 0.0


def fmt_naira(val):
    """Format as Naira with appropriate suffix"""
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000_000:
        return f"{sign}\u20A6{v/1_000_000_000:.2f}B"
    elif v >= 1_000_000:
        return f"{sign}\u20A6{v/1_000_000:.1f}M"
    elif v >= 1_000:
        return f"{sign}\u20A6{v/1_000:.0f}K"
    else:
        return f"{sign}\u20A6{v:,.0f}"


def fmt_usd(val):
    """Format as USD with appropriate suffix"""
    v = abs(val)
    sign = '-' if val < 0 else ''
    if v >= 1_000_000_000:
        return f"{sign}${v/1_000_000_000:.2f}B"
    elif v >= 1_000_000:
        return f"{sign}${v/1_000_000:.1f}M"
    elif v >= 1_000:
        return f"{sign}${v/1_000:.0f}K"
    else:
        return f"{sign}${v:,.0f}"


def extract_report(filepath):
    fn = os.path.basename(filepath)
    dt = parse_date(fn)
    if not dt:
        return None

    wb = load_workbook(filepath, data_only=True)
    rec = {'date': dt, 'date_str': dt.strftime('%d %b %Y'), 'filename': fn}

    # === CASH REPORT SHEET ===
    ws = wb['Cash Report']
    for row in ws.iter_rows(min_row=5, max_row=70, values_only=False):
        cells = safe_row_dict(row)
        c_str = str(cells.get('C', '') or '').strip()
        b_str = str(cells.get('B', '') or '').strip()

        if 'TOTAL CASH (NGN)' in c_str:
            rec['ngn_closing'] = sf(cells.get('J'))
            rec['ngn_opening'] = sf(cells.get('F'))
            rec['ngn_inflow'] = sf(cells.get('G'))
            rec['ngn_outflow'] = sf(cells.get('H'))

        if 'TOTAL CASH (USD' in b_str or 'TOTAL CASH (USD' in c_str:
            rec['usd_closing_ngn'] = sf(cells.get('J'))
            rec['usd_opening_ngn'] = sf(cells.get('F'))
            rec['usd_inflow'] = sf(cells.get('G'))
            rec['usd_outflow'] = sf(cells.get('H'))
            rec['usd_raw'] = sf(cells.get('D'))

        if 'TOTAL INVESTMENT' in b_str:
            if 'USD' in b_str.upper():
                # USD investments: closing balance is in column H (different layout from cash section)
                rec['investment_usd_raw'] = sf(cells.get('H'))
            else:
                # NGN investments: closing balance is in column J (standard layout)
                rec['investment_ngn'] = sf(cells.get('J'))

    # FX rate - scan USD rows for exchange rate
    rec['fx_rate'] = 1.0
    for row in ws.iter_rows(min_row=30, max_row=45, values_only=False):
        cells = safe_row_dict(row)
        c_str = str(cells.get('C', '') or '').strip()
        e_val = cells.get('E')
        if c_str and ('USD' in c_str or 'GTB' in c_str) and e_val and isinstance(e_val, (int, float)) and e_val > 100:
            rec['fx_rate'] = float(e_val)
            break

    # GBP balance
    rec['gbp_closing_ngn'] = 0.0
    for row in ws.iter_rows(min_row=28, max_row=35, values_only=False):
        cells = safe_row_dict(row)
        c_str = str(cells.get('C', '') or '').strip()
        if 'POUND' in c_str.upper():
            rec['gbp_closing_ngn'] = sf(cells.get('J'))
            break

    # USD investments converted to NGN (FX rate must be resolved first)
    rec['investment_usd_ngn'] = rec.get('investment_usd_raw', 0) * rec.get('fx_rate', 1)

    # Total cash position in NGN equivalent (liquid cash + all investments)
    rec['total_cash_ngn'] = (
        rec.get('ngn_closing', 0)
        + rec.get('usd_closing_ngn', 0)
        + rec.get('gbp_closing_ngn', 0)
        + rec.get('investment_ngn', 0)
        + rec.get('investment_usd_ngn', 0)
    )
    # Also keep a liquid-only view for reference
    rec['liquid_cash_ngn'] = rec.get('ngn_closing', 0) + rec.get('usd_closing_ngn', 0) + rec.get('gbp_closing_ngn', 0)

    # === BREAKDOWN OF INCOME AND EXPENDITURE SHEET ===
    ws_bd = wb['Breakdown of Income and Expendi']

    inflow_items = {}
    outflow_items = {}
    total_inflow_cash = None
    total_outflow_cash = None
    closing_balance = None
    total_opening = None

    in_inflows = False
    in_outflows = False

    for row in ws_bd.iter_rows(min_row=1, max_row=ws_bd.max_row, values_only=False):
        cells = safe_row_dict(row)
        b_val = str(cells.get('B', '') or '').strip()
        c_val = cells.get('C')
        g_val = str(cells.get('G', '') or '').strip()
        h_val = cells.get('H')
        j_val = cells.get('J')

        # Right side: opening/closing
        if 'Total Opening Balance' in g_val:
            total_opening = sf(j_val)
        if 'CASH BALANCE AS AT' in g_val:
            closing_balance = sf(j_val)

        # Left side: dynamic parsing of inflows
        if b_val.startswith('Inflow') and ':' in b_val:
            in_inflows = True
            in_outflows = False
            continue
        if 'OUTFLOW' in b_val.upper() and c_val is not None and str(c_val).strip() in ('Amount', ''):
            in_outflows = True
            in_inflows = False
            continue
        if b_val.upper().startswith('OUTFLOW') and c_val is None:
            in_outflows = True
            in_inflows = False
            continue

        if in_inflows and c_val is not None and isinstance(c_val, (int, float)):
            if 'Total' in b_val and 'Inflow' in b_val:
                total_inflow_cash = sf(c_val)
                in_inflows = False
            elif b_val and b_val not in ('Inflows:', 'INFLOWS'):
                inflow_items[b_val] = sf(c_val)

        if in_outflows and c_val is not None and isinstance(c_val, (int, float)):
            if 'Total' in b_val and 'Outflow' in b_val:
                total_outflow_cash = sf(c_val)
                in_outflows = False
            elif 'Net Cash' in b_val:
                continue  # Skip the net row
            elif b_val and b_val not in ('OUTFLOWS',):
                outflow_items[b_val] = sf(c_val)

    rec['inflow_items'] = inflow_items
    rec['outflow_items'] = outflow_items
    rec['total_inflow'] = total_inflow_cash or sum(inflow_items.values())
    rec['total_outflow'] = total_outflow_cash or sum(outflow_items.values())
    rec['closing_balance'] = closing_balance
    rec['total_opening'] = total_opening

    wb.close()
    return rec


def detect_anomalies(reports):
    anomalies = []
    # Category averages
    all_cats = {}
    for r in reports:
        for k, v in r.get('outflow_items', {}).items():
            all_cats.setdefault(k, []).append(v)
    cat_avgs = {k: sum(v)/len(v) for k, v in all_cats.items() if v}

    for i, r in enumerate(reports):
        wk = r['date_str']

        # Negative net flow
        net = r.get('total_inflow', 0) - r.get('total_outflow', 0)
        if net < 0:
            anomalies.append({'week': wk, 'type': 'Negative Net Flow', 'severity': 'high',
                'detail': f"Outflows ({fmt_naira(r['total_outflow'])}) exceeded inflows ({fmt_naira(r['total_inflow'])}) by {fmt_naira(abs(net))}"})

        # Category spikes (> 2x average)
        for cat, amt in r.get('outflow_items', {}).items():
            avg = cat_avgs.get(cat, 0)
            if avg > 0 and amt > avg * 2 and amt > 1_000_000:
                anomalies.append({'week': wk, 'type': f'Expense Spike: {cat}', 'severity': 'medium',
                    'detail': f"{fmt_naira(amt)} vs average {fmt_naira(avg)} ({amt/avg:.1f}x)"})

        # Bank charge spike
        bc = r.get('outflow_items', {}).get('Bank Charges', 0)
        if i > 0:
            prev_bc = reports[i-1].get('outflow_items', {}).get('Bank Charges', 0)
            if prev_bc > 0 and bc > prev_bc * 1.5 and bc > 500_000:
                # Check for concurrent investment outflow (transaction fee context)
                inv_amt = get_investment_outflows(r.get('outflow_items', {}))
                context = f" (likely SCB investment fee on {fmt_naira(inv_amt)} transaction)" if inv_amt > 100_000_000 else ""
                anomalies.append({'week': wk, 'type': 'Bank Charge Spike', 'severity': 'medium' if inv_amt > 100_000_000 else 'high',
                    'detail': f"{fmt_naira(bc)} vs prior week {fmt_naira(prev_bc)} ({bc/prev_bc:.1f}x increase){context}"})

        # Inflow drop > 30%
        if i > 0:
            prev_in = reports[i-1].get('total_inflow', 0)
            curr_in = r.get('total_inflow', 0)
            if prev_in > 0 and curr_in < prev_in * 0.7:
                anomalies.append({'week': wk, 'type': 'Inflow Drop', 'severity': 'high',
                    'detail': f"{fmt_naira(curr_in)} vs prior {fmt_naira(prev_in)} ({(1 - curr_in/prev_in)*100:.0f}% decrease)"})

        # New category detection
        if i > 0:
            prev_cats = set(reports[i-1].get('inflow_items', {}).keys())
            curr_cats = set(r.get('inflow_items', {}).keys())
            new_cats = curr_cats - prev_cats
            for nc in new_cats:
                amt = r['inflow_items'].get(nc, 0)
                if amt > 1_000_000:
                    anomalies.append({'week': wk, 'type': 'New Revenue Stream', 'severity': 'info',
                        'detail': f"{nc}: {fmt_naira(amt)} (first appearance)"})

    return anomalies


def validate_reports(reports):
    """
    Sanity-check parsed reports and return a list of data quality warnings.
    These surface in the dashboard as a visible banner so parsing failures
    are caught immediately — before leadership acts on wrong numbers.

    Returns a list of dicts: {level: 'error'|'warning', report: date_str, message: str}
    """
    warnings = []
    if not reports:
        return warnings

    totals = [r.get('total_cash_ngn', 0) for r in reports]
    median_total = sorted(totals)[len(totals) // 2]

    for i, r in enumerate(reports):
        wk = r['date_str']
        total = r.get('total_cash_ngn', 0)
        liquid = r.get('liquid_cash_ngn', 0)
        usd_raw = r.get('investment_usd_raw', 0)
        ngn_close = r.get('ngn_closing', 0)

        # 1. NGN closing balance not found — sheet structure may have changed
        if ngn_close == 0:
            warnings.append({
                'level': 'error', 'report': wk,
                'message': f"NGN closing balance is zero — 'TOTAL CASH (NGN)' row not found. "
                           f"The Cash Report sheet layout may have changed."
            })

        # 2. USD investment portfolio not detected
        if usd_raw == 0:
            warnings.append({
                'level': 'error', 'report': wk,
                'message': f"USD investment portfolio not found (expected ~$1.25M+). "
                           f"Check that 'TOTAL INVESTMENT (USD)' row is within rows 5–70 of the Cash Report sheet."
            })

        # 3. Total implausibly low vs historical median
        if median_total > 0 and total < median_total * 0.4 and total > 0:
            warnings.append({
                'level': 'error', 'report': wk,
                'message': f"Total cash position of {fmt_naira(total)} is less than 40% of the historical median "
                           f"({fmt_naira(median_total)}). Likely a parsing failure — verify the source file."
            })

        # 4. Implausible week-on-week drop (>40%) not explained by investment activity
        if i > 0:
            prev_total = reports[i-1].get('total_cash_ngn', 0)
            if prev_total > 0:
                wow_chg = (total - prev_total) / prev_total
                inv_out = get_investment_outflows(r.get('outflow_items', {}))
                inv_in = get_investment_inflows(r.get('inflow_items', {}))
                # Only flag if the drop can't be explained by a large investment outflow
                if wow_chg < -0.4 and inv_out < abs(total - prev_total) * 0.5:
                    warnings.append({
                        'level': 'warning', 'report': wk,
                        'message': f"Cash position dropped {abs(wow_chg)*100:.0f}% WoW "
                                   f"({fmt_naira(prev_total)} → {fmt_naira(total)}) without a matching "
                                   f"investment transfer. Confirm the source file is correct."
                    })

        # 5. FX rate looks implausible (outside ₦800–₦3000/$)
        fx = r.get('fx_rate', 0)
        if fx > 0 and (fx < 800 or fx > 3000):
            warnings.append({
                'level': 'warning', 'report': wk,
                'message': f"FX rate of ₦{fx:,.0f}/$ looks implausible. "
                           f"USD-denominated balances may be misconverted."
            })

    return warnings


def generate_insights(reports):
    if len(reports) < 2:
        return ["Insufficient data for trend analysis. Need at least 2 weeks."]

    insights = []
    first_cash = reports[0].get('total_cash_ngn', 0) or reports[0].get('closing_balance', 0)
    last_cash = reports[-1].get('total_cash_ngn', 0) or reports[-1].get('closing_balance', 0)
    peak_cash = max(r.get('total_cash_ngn', 0) or r.get('closing_balance', 0) for r in reports)
    peak_wk = [r for r in reports if (r.get('total_cash_ngn', 0) or r.get('closing_balance', 0)) == peak_cash][0]

    if last_cash > first_cash:
        pct = (last_cash - first_cash) / first_cash * 100
        insights.append(f"Cash position grew {pct:.0f}% over the period, from {fmt_naira(first_cash)} to {fmt_naira(last_cash)}. Peak was {fmt_naira(peak_cash)} on {peak_wk['date_str']}.")
    else:
        pct = (first_cash - last_cash) / first_cash * 100
        insights.append(f"Cash position declined {pct:.0f}% from {fmt_naira(first_cash)} to {fmt_naira(last_cash)}. Peak was {fmt_naira(peak_cash)} on {peak_wk['date_str']}.")

    # Revenue concentration analysis (exclude investment liquidations — asset conversions, not revenue)
    all_sources = {}
    for r in reports:
        for k, v in r.get('inflow_items', {}).items():
            if not is_investment_inflow(k):
                all_sources.setdefault(k, 0)
                all_sources[k] += v
    total_rev = sum(all_sources.values())
    if total_rev > 0:
        sorted_sources = sorted(all_sources.items(), key=lambda x: -x[1])
        top = sorted_sources[0]
        conc = top[1] / total_rev * 100
        if conc > 50:
            insights.append(f"Revenue is heavily concentrated: {top[0]} represents {conc:.0f}% of total inflows ({fmt_naira(top[1])}). Diversification is a strategic priority.")
        if len(sorted_sources) > 1:
            new_streams = [s for s in sorted_sources[1:5] if s[1] > 10_000_000]
            if new_streams:
                names = ', '.join([f"{s[0]} ({fmt_naira(s[1])})" for s in new_streams])
                insights.append(f"Emerging revenue streams showing promise: {names}.")

    # Expense analysis (exclude investment transfers — asset reallocation, not operational expense)
    all_expenses = {}
    for r in reports:
        for k, v in r.get('outflow_items', {}).items():
            if not is_investment_outflow(k):
                all_expenses.setdefault(k, 0)
                all_expenses[k] += v
    total_exp = sum(all_expenses.values())
    if total_exp > 0:
        sorted_exp = sorted(all_expenses.items(), key=lambda x: -x[1])
        top3 = sorted_exp[:3]
        parts = ', '.join([f"{e[0]} ({e[1]/total_exp*100:.0f}%)" for e in top3])
        insights.append(f"Top 3 operational expense categories over the period: {parts}.")

    # Average weekly burn and runway (using operational burn, excluding investment outflows)
    op_burns_ins = []
    for r in reports:
        items = r.get('outflow_items', {})
        total = r.get('total_outflow', 0)
        inv = get_investment_outflows(items)
        op_burns_ins.append(total - inv)
    if op_burns_ins:
        avg_op = sum(op_burns_ins) / len(op_burns_ins)
        runway_ins = last_cash / avg_op if avg_op > 0 else 0
        insights.append(f"Average weekly operational burn (excl. investment outflows) is {fmt_naira(avg_op)}. Total position (incl. investments): {fmt_naira(last_cash)}. Estimated runway: {runway_ins:.0f} weeks ({runway_ins/4.3:.0f} months).")

    # Investment activity
    inv_in = sum(r.get('inflow_items', {}).get('Investment Withdrawal', 0) + r.get('inflow_items', {}).get('Investment Liquidation', 0) + r.get('inflow_items', {}).get('Investment withdrawal', 0) for r in reports)
    if inv_in > 100_000_000:
        insights.append(f"Significant investment liquidation activity: {fmt_naira(inv_in)} withdrawn from investments over the period. Monitor whether this is strategic redeployment or liquidity pressure.")

    return insights


def generate_takeaways(reports):
    """Generate McKinsey-style executive takeaways — the 'so what' behind the numbers.
    These auto-update weekly as new data is added."""
    if len(reports) < 2:
        return []

    takeaways = []
    latest = reports[-1]
    prev = reports[-2]

    first_cash = reports[0].get('total_cash_ngn', 0) or reports[0].get('closing_balance', 0)
    last_cash = reports[-1].get('total_cash_ngn', 0) or reports[-1].get('closing_balance', 0)
    prev_cash = reports[-2].get('total_cash_ngn', 0) or reports[-2].get('closing_balance', 0)
    peak_cash = max(r.get('total_cash_ngn', 0) or r.get('closing_balance', 0) for r in reports)
    peak_wk = [r for r in reports if (r.get('total_cash_ngn', 0) or r.get('closing_balance', 0)) == peak_cash][0]
    trough_cash = min(r.get('total_cash_ngn', 0) or r.get('closing_balance', 0) for r in reports)

    # --- 1. CASH POSITION & TRAJECTORY ---
    cash_chg_period = ((last_cash - first_cash) / first_cash * 100) if first_cash else 0
    cash_chg_wow = ((last_cash - prev_cash) / prev_cash * 100) if prev_cash else 0
    vol = peak_cash - trough_cash

    # Detect investment transfers this week
    latest_inv_out = get_investment_outflows(latest.get('outflow_items', {}))
    latest_inv_in = get_investment_inflows(latest.get('inflow_items', {}))
    liquid_now = latest.get('liquid_cash_ngn', 0)
    liquid_prev = prev.get('liquid_cash_ngn', 0)
    liquid_chg_wow = ((liquid_now - liquid_prev) / liquid_prev * 100) if liquid_prev else 0

    if cash_chg_wow < -20:
        urgency = "urgent"
        headline = f"Total position dropped {abs(cash_chg_wow):.0f}% this week to {fmt_naira(last_cash)}"
        body = f"Down from {fmt_naira(prev_cash)} last week. "
        # Check what caused the drop
        latest_out = latest.get('outflow_items', {})
        big_items = sorted(latest_out.items(), key=lambda x: -x[1])[:3]
        if big_items:
            drivers = ', '.join([f"{k} ({fmt_naira(v)})" for k, v in big_items])
            body += f"Primary drivers: {drivers}. "
        body += "Determine whether these are one-time or recurring to assess true position."
    elif cash_chg_wow > 20:
        urgency = "positive"
        headline = f"Total position surged {cash_chg_wow:.0f}% this week to {fmt_naira(last_cash)}"
        body = f"Up from {fmt_naira(prev_cash)} last week. "
        latest_in = latest.get('inflow_items', {})
        big_items = sorted(latest_in.items(), key=lambda x: -x[1])[:3]
        if big_items:
            drivers = ', '.join([f"{k} ({fmt_naira(v)})" for k, v in big_items])
            body += f"Driven by: {drivers}. "
        if latest_inv_in > 0 and latest.get('total_inflow', 0) > 0 and latest_inv_in / latest['total_inflow'] > 0.3:
            body += f"Note: {fmt_naira(latest_inv_in)} ({latest_inv_in/latest['total_inflow']*100:.0f}% of inflows) came from investment liquidation — this inflates operational cash appearance."
    else:
        urgency = "neutral"
        direction = "grew" if cash_chg_wow >= 0 else "contracted"
        headline = f"Total position at {fmt_naira(last_cash)} ({direction} {abs(cash_chg_wow):.1f}% WoW)"
        body = f"Over the full reporting period, total position has {'grown' if cash_chg_period > 0 else 'declined'} {abs(cash_chg_period):.0f}% from {fmt_naira(first_cash)}. "
        body += f"Volatility band: {fmt_naira(trough_cash)} to {fmt_naira(peak_cash)}. "

    # Add investment transfer context when significant
    if latest_inv_out > 50_000_000:
        body += f"<br><strong>Investment transfer this week:</strong> {fmt_naira(latest_inv_out)} moved from bank accounts to investment (SCB mutual fund). Liquid cash is {fmt_naira(liquid_now)} (down {abs(liquid_chg_wow):.0f}% WoW), but the total position including investments remains {'stable' if abs(cash_chg_wow) < 10 else 'strong'}. This is an asset reallocation, not a loss."
    elif latest_inv_in > 50_000_000:
        body += f"<br><strong>Investment liquidation this week:</strong> {fmt_naira(latest_inv_in)} moved from investments into bank accounts. Liquid cash boosted but total position unchanged."

    takeaways.append({'title': 'Cash Position & Trajectory', 'headline': headline, 'body': body, 'urgency': urgency, 'icon': '\U0001F4B0'})

    # --- 2. REVENUE CONCENTRATION RISK ---
    # Exclude investment liquidations — these are asset conversions, not revenue
    all_sources = {}
    for r in reports:
        for k, v in r.get('inflow_items', {}).items():
            if not is_investment_inflow(k):
                all_sources.setdefault(k, 0)
                all_sources[k] += v
    total_rev = sum(all_sources.values())

    if total_rev > 0:
        sorted_sources = sorted(all_sources.items(), key=lambda x: -x[1])
        top = sorted_sources[0]
        top_pct = top[1] / total_rev * 100

        # Count distinct revenue sources > 1% of total
        meaningful_sources = [s for s in sorted_sources if s[1] / total_rev > 0.01]
        num_sources = len(meaningful_sources)

        # Latest week concentration
        latest_in = latest.get('inflow_items', {})
        latest_total = sum(latest_in.values())
        if latest_total > 0:
            latest_top = max(latest_in.items(), key=lambda x: x[1])
            latest_conc = latest_top[1] / latest_total * 100
        else:
            latest_conc = 0
            latest_top = ('N/A', 0)

        if top_pct > 50:
            urgency = "urgent"
            headline = f"{top[0]} represents {top_pct:.0f}% of total revenue — critical concentration risk"
        elif top_pct > 30:
            urgency = "caution"
            headline = f"Moderate concentration: {top[0]} at {top_pct:.0f}% of revenue"
        else:
            urgency = "positive"
            headline = f"Revenue well diversified — top source is only {top_pct:.0f}%"

        body = f"Across the period, {num_sources} meaningful revenue streams identified. "

        # New/emerging streams (exclude investment liquidations — asset conversions, not revenue)
        recent_new = set()
        for i in range(max(0, len(reports)-3), len(reports)):
            curr_keys = set(reports[i].get('inflow_items', {}).keys())
            if i > 0:
                prev_keys = set()
                for j in range(0, i):
                    prev_keys.update(reports[j].get('inflow_items', {}).keys())
                new = curr_keys - prev_keys
                for n in new:
                    if not is_investment_inflow(n) and reports[i]['inflow_items'].get(n, 0) > 5_000_000:
                        recent_new.add(n)
        if recent_new:
            body += f"Recently emerged: {', '.join(recent_new)}. These need nurturing to reduce dependency on {top[0]}. "

        body += f"This week: {latest_top[0]} was {latest_conc:.0f}% of inflows."
        takeaways.append({'title': 'Revenue Concentration Risk', 'headline': headline, 'body': body, 'urgency': urgency, 'icon': '\u26A0\uFE0F'})

    # --- 3. EXPENSE CONTROL ---
    # Use operational outflows (excluding investment transfers) for meaningful comparison
    all_expenses = {}
    for r in reports:
        for k, v in r.get('outflow_items', {}).items():
            all_expenses.setdefault(k, []).append(v)
    total_exp_all = sum(sum(v) for v in all_expenses.values())

    latest_items = latest.get('outflow_items', {})
    prev_items = prev.get('outflow_items', {})

    latest_inv_outflow = get_investment_outflows(latest_items)
    prev_inv_outflow = get_investment_outflows(prev_items)
    latest_op_outflow = latest.get('total_outflow', 0) - latest_inv_outflow
    prev_op_outflow = prev.get('total_outflow', 0) - prev_inv_outflow
    latest_total_outflow = latest.get('total_outflow', 0)

    op_chg = ((latest_op_outflow - prev_op_outflow) / prev_op_outflow * 100) if prev_op_outflow > 0 else 0

    # Find categories with biggest week-over-week increase (excluding investment transfers)
    spikes = []
    for cat, amt in latest_items.items():
        if is_investment_outflow(cat):
            continue  # Skip investment transfers from spike analysis
        prev_amt = prev_items.get(cat, 0)
        avg_amt = sum(all_expenses.get(cat, [0])) / max(len(all_expenses.get(cat, [1])), 1)
        if amt > avg_amt * 2 and amt > 1_000_000:
            spikes.append((cat, amt, avg_amt))
    spikes.sort(key=lambda x: -x[1])

    if latest_op_outflow > prev_op_outflow * 2:
        urgency = "urgent"
        headline = f"Operational outflows surged {op_chg:.0f}% to {fmt_naira(latest_op_outflow)} — investigate immediately"
    elif latest_op_outflow > prev_op_outflow * 1.3:
        urgency = "caution"
        headline = f"Operational outflows up {op_chg:.0f}% this week to {fmt_naira(latest_op_outflow)}"
    elif latest_op_outflow < prev_op_outflow * 0.7:
        urgency = "positive"
        headline = f"Operational outflows decreased {abs(op_chg):.0f}% to {fmt_naira(latest_op_outflow)} — improved cost control"
    else:
        urgency = "neutral"
        headline = f"Operational outflows at {fmt_naira(latest_op_outflow)} ({'up' if op_chg > 0 else 'down'} {abs(op_chg):.0f}% WoW)"

    body = ""
    # Note investment transfer if significant
    if latest_inv_outflow > 50_000_000:
        body += f"Total outflows were {fmt_naira(latest_total_outflow)}, of which {fmt_naira(latest_inv_outflow)} was an investment transfer (not operational). Operational spend: {fmt_naira(latest_op_outflow)}. "

    if spikes:
        spike_text = '; '.join([f"{s[0]}: {fmt_naira(s[1])} vs avg {fmt_naira(s[2])} ({s[1]/s[2]:.1f}x)" for s in spikes[:3]])
        body += f"Expense spikes this week: {spike_text}. "

    # Bank charges check
    bc = latest_items.get('Bank Charges', 0)
    bc_avg = sum(all_expenses.get('Bank Charges', [0])) / max(len(all_expenses.get('Bank Charges', [1])), 1)
    if bc > bc_avg * 2 and bc > 500_000:
        # Check if there's a concurrent investment outflow (SCB fee context)
        inv_this_week = get_investment_outflows(latest_items)
        if inv_this_week > 100_000_000:
            body += f"Bank charges at {fmt_naira(bc)} are {bc/bc_avg:.1f}x the average — likely driven by the {fmt_naira(inv_this_week)} investment transaction fee (2.5% + VAT). "
        else:
            body += f"Bank charges at {fmt_naira(bc)} are {bc/bc_avg:.1f}x the average — investigate with the bank. "

    # Salary cycle detection
    salary_keys = ['Salary', ' Salary', 'Salaries', ' Salaries']
    salary_this_week = sum(latest_items.get(k, 0) for k in salary_keys)
    if salary_this_week > 50_000_000:
        body += f"Salary batch of {fmt_naira(salary_this_week)} processed this week (monthly cycle). "

    if not body:
        body = "No unusual expense patterns detected this week. Continue monitoring."

    takeaways.append({'title': 'Expense Control', 'headline': headline, 'body': body, 'urgency': urgency, 'icon': '\U0001F4C9'})

    # --- 4. LIQUIDITY & RUNWAY ---
    # Use operational burn (excluding investment outflows) as the primary metric
    op_burns = []
    total_burns = []
    for r in reports:
        items = r.get('outflow_items', {})
        total = r.get('total_outflow', 0)
        inv = get_investment_outflows(items)
        op_burns.append(total - inv)
        total_burns.append(total)
    avg_op_burn = sum(op_burns) / len(op_burns) if op_burns else 0
    avg_total_burn = sum(total_burns) / len(total_burns) if total_burns else 0

    # Runway based on total position (including investments) and operational burn
    runway_weeks = last_cash / avg_op_burn if avg_op_burn > 0 else float('inf')
    runway_months = runway_weeks / 4.3

    # Investment balance (NGN + USD portfolios)
    inv_balance = latest.get('investment_ngn', 0) + latest.get('investment_usd_ngn', 0)
    liquid_cash = latest.get('liquid_cash_ngn', 0)

    if runway_weeks < 8:
        urgency = "urgent"
        headline = f"Cash runway at {runway_weeks:.0f} weeks — below 2-month safety threshold"
    elif runway_weeks < 13:
        urgency = "caution"
        headline = f"Cash runway at {runway_weeks:.0f} weeks ({runway_months:.0f} months) — monitor closely"
    else:
        urgency = "positive"
        headline = f"Healthy runway of {runway_weeks:.0f} weeks ({runway_months:.0f} months) at operational burn rate"

    body = f"Total position: {fmt_naira(last_cash)} (liquid: {fmt_naira(liquid_cash)}, investments: {fmt_naira(inv_balance)}). "
    body += f"Operational burn (excl. investment transfers): {fmt_naira(avg_op_burn)}/week. "

    # FX exposure
    usd_pct = latest.get('usd_closing_ngn', 0) / last_cash * 100 if last_cash > 0 else 0
    if usd_pct > 30:
        fx_first = reports[0].get('fx_rate', 1)
        fx_last = latest.get('fx_rate', 1)
        fx_chg = ((fx_last - fx_first) / fx_first * 100) if fx_first > 1 else 0
        body += f"USD holdings represent {usd_pct:.0f}% of total cash. FX rate moved {fx_chg:+.1f}% over the period ({fx_first:,.0f} to {fx_last:,.0f}) — currency risk is material."

    takeaways.append({'title': 'Liquidity & Runway', 'headline': headline, 'body': body, 'urgency': urgency, 'icon': '\U0001F6A8'})

    # --- 5. WHAT TO WATCH NEXT WEEK ---
    watch_items = []

    # Upcoming salary cycle?
    weeks_since_salary = 0
    for r in reversed(reports):
        sal = sum(r.get('outflow_items', {}).get(k, 0) for k in salary_keys)
        if sal > 50_000_000:
            break
        weeks_since_salary += 1
    if weeks_since_salary >= 3:
        watch_items.append("Salary batch likely due in the next 1-2 weeks — ensure sufficient NGN liquidity.")

    # Negative net trend (using operational flows, excluding investment transfers)
    recent_op_nets = []
    for r in reports[-3:]:
        op_out = r.get('total_outflow', 0) - get_investment_outflows(r.get('outflow_items', {}))
        op_in = r.get('total_inflow', 0) - get_investment_inflows(r.get('inflow_items', {}))
        recent_op_nets.append(op_in - op_out)
    neg_count = sum(1 for n in recent_op_nets if n < 0)
    if neg_count >= 2:
        watch_items.append(f"Operational cash flow was negative in {neg_count} of the last 3 weeks — this trend needs to reverse.")

    # Revenue concentration
    if total_rev > 0 and top_pct > 60:
        watch_items.append(f"Track whether {top[0]} payments arrive on schedule — any delay directly threatens operating cash.")

    # Large one-offs to check (excluding investment transfers)
    for cat, amt in latest_items.items():
        if is_investment_outflow(cat):
            continue  # Investment transfers are handled in the Cash Position takeaway
        avg = sum(all_expenses.get(cat, [0])) / max(len(all_expenses.get(cat, [1])), 1)
        if amt > avg * 3 and amt > 10_000_000:
            watch_items.append(f"Confirm {cat} payment of {fmt_naira(amt)} was authorized and is non-recurring.")

    if not watch_items:
        watch_items.append("No immediate flags — maintain standard monitoring cadence.")

    body = ''.join([f"<li>{item}</li>" for item in watch_items])
    takeaways.append({'title': 'Watch List for Next Week', 'headline': 'Action items requiring attention', 'body': f"<ul style='margin:8px 0 0 16px;padding:0;list-style:disc'>{body}</ul>", 'urgency': 'neutral', 'icon': '\U0001F440'})

    return takeaways


def generate_html(reports, anomalies, insights, takeaways, output_path, data_warnings=None):
    dates = [r['date_str'] for r in reports]
    cash_positions = [r.get('total_cash_ngn', 0) or r.get('closing_balance', 0) for r in reports]
    ngn_positions = [r.get('ngn_closing', 0) for r in reports]
    usd_positions = [r.get('usd_closing_ngn', 0) for r in reports]
    inflows = [r.get('total_inflow', 0) for r in reports]
    outflows = [r.get('total_outflow', 0) for r in reports]
    nets = [i - o for i, o in zip(inflows, outflows)]
    fx_rates = [r.get('fx_rate', 1) for r in reports]

    # Revenue source mix - gather all sources
    all_rev_sources = set()
    for r in reports:
        for k in r.get('inflow_items', {}):
            if k not in ('Interbank',):
                all_rev_sources.add(k)

    # Expense categories - standardize
    cat_map = {
        'Salary': ['Salary', ' Salary', 'Salaries', ' Salaries', 'Exit  Salary', 'Exit Salary'],
        'Payment Batches': ['Payment Batches'],
        'Software Tools': ['Software Tools'],
        'Travel & Flights': ['Flight Ticket/Accommodation', 'BTA'],
        'Bank Charges': ['Bank Charges'],
        'Tax & Statutory': ['Paye', 'PAYE', 'VAT', 'WHT', 'VAT Remitted', 'WHT remitted', 'Pension', 'Pension Remittance', 'NSITF Contributions', 'Business Premises & Dev Levy'],
        'Management Allowance': ['Management Allowance'],
        'Office & Supplies': ['Office Supplies & Fuel', 'Office Supplies, Fuel & Letter Dispatch'],
        'Vendors & Contractors': ['Tatvasoft', 'Bind Creative', 'Amazon', 'Nimc consultancy fee'],
        'Conferences & Events': ['ID4Africa', 'ID4Africa Registration', 'MWC Registration', 'Training'],
        'Investment Outflows': ['Seamfix UAE funding', 'Investment in SCB USD mutua fund'],
        'Other': [],
    }
    # Reverse map
    rev_cat = {}
    for std, originals in cat_map.items():
        for o in originals:
            rev_cat[o] = std

    std_expense_cats = list(cat_map.keys())

    expense_data = {cat: [] for cat in std_expense_cats}
    for r in reports:
        items = r.get('outflow_items', {})
        cat_totals = {cat: 0 for cat in std_expense_cats}
        for k, v in items.items():
            mapped = rev_cat.get(k)
            if not mapped:
                # Keyword fallback for new investment categories
                mapped = 'Investment Outflows' if is_investment_outflow(k) else 'Other'
            cat_totals[mapped] += v
        for cat in std_expense_cats:
            expense_data[cat].append(cat_totals[cat])

    # KPIs
    latest = reports[-1]
    prev = reports[-2] if len(reports) > 1 else latest
    latest_cash = cash_positions[-1]
    prev_cash = cash_positions[-2] if len(cash_positions) > 1 else latest_cash
    cash_chg = ((latest_cash - prev_cash) / prev_cash * 100) if prev_cash else 0

    # NGN vs USD balance split (naira-denominated vs dollar-denominated holdings, incl. investments)
    ngn_balance = latest.get('ngn_closing', 0) + latest.get('investment_ngn', 0)
    usd_balance = latest.get('usd_raw', 0) + latest.get('investment_usd_raw', 0)
    prev_ngn_balance = prev.get('ngn_closing', 0) + prev.get('investment_ngn', 0)
    prev_usd_balance = prev.get('usd_raw', 0) + prev.get('investment_usd_raw', 0)
    ngn_bal_chg = ((ngn_balance - prev_ngn_balance) / prev_ngn_balance * 100) if prev_ngn_balance else 0
    usd_bal_chg = ((usd_balance - prev_usd_balance) / prev_usd_balance * 100) if prev_usd_balance else 0

    latest_in = inflows[-1]
    prev_in = inflows[-2] if len(inflows) > 1 else latest_in
    in_chg = ((latest_in - prev_in) / prev_in * 100) if prev_in else 0

    latest_out = outflows[-1]
    prev_out = outflows[-2] if len(outflows) > 1 else latest_out
    out_chg = ((latest_out - prev_out) / prev_out * 100) if prev_out else 0

    net_flow = latest_in - latest_out

    # Operational burn (excluding investment outflows)
    op_outflows = []
    for r in reports:
        items = r.get('outflow_items', {})
        total = r.get('total_outflow', 0)
        inv = get_investment_outflows(items)
        op_outflows.append(total - inv)
    avg_op_burn = sum(op_outflows) / len(op_outflows) if op_outflows else 1
    avg_burn = sum(outflows) / len(outflows) if outflows else 1
    runway = latest_cash / avg_op_burn if avg_op_burn > 0 else 0

    # 4-week cash forecast — two views
    # Expected: uses avg net cash flow (op inflows minus op outflows)
    # Floor: burn-only, assumes zero inflows (worst case / stress test)
    op_inflows_list = []
    for r in reports:
        items = r.get('inflow_items', {})
        total_in = r.get('total_inflow', 0)
        inv_in = sum(v for k, v in items.items() if is_investment_inflow(k))
        op_inflows_list.append(total_in - inv_in)
    avg_op_inflow = sum(op_inflows_list) / len(op_inflows_list) if op_inflows_list else 0

    avg_weekly_net = avg_op_inflow - avg_op_burn
    forecast_4w_net = latest_cash + (avg_weekly_net * 4)
    forecast_4w_net_chg_pct = ((forecast_4w_net - latest_cash) / latest_cash * 100) if latest_cash > 0 else 0

    forecast_4w_floor = latest_cash - (avg_op_burn * 4)
    forecast_4w_floor_chg_pct = ((forecast_4w_floor - latest_cash) / latest_cash * 100) if latest_cash > 0 else 0

    # Revenue concentration for latest week
    latest_inflows = latest.get('inflow_items', {})
    if latest_inflows:
        top_src = max(latest_inflows.items(), key=lambda x: x[1])
        total_in_latest = sum(latest_inflows.values())
        concentration = (top_src[1] / total_in_latest * 100) if total_in_latest > 0 else 0
        conc_name = top_src[0]
    else:
        concentration = 0
        conc_name = "N/A"

    # Revenue datasets for chart
    rev_colors = ['#00D4AA', '#4ECDC4', '#FFE66D', '#FF6B6B', '#A8E6CF', '#95E1D3', '#F38181', '#AA96DA', '#FCbad3', '#FFFFD2', '#B5EAD7', '#C7CEEA']
    rev_datasets_js = []
    for idx, src in enumerate(sorted(all_rev_sources)):
        vals = [r.get('inflow_items', {}).get(src, 0) for r in reports]
        if sum(vals) > 0:
            color = rev_colors[idx % len(rev_colors)]
            rev_datasets_js.append(f'{{label:{json.dumps(src)},data:{json.dumps(vals)},backgroundColor:"{color}"}}')

    # Expense datasets for chart
    exp_colors = ['#FF6B6B', '#FFE66D', '#00D4AA', '#4ECDC4', '#A8E6CF', '#F38181', '#AA96DA', '#FCbad3', '#95E1D3', '#B5EAD7', '#C7CEEA', '#FFFFD2']
    exp_datasets_js = []
    for idx, cat in enumerate(std_expense_cats):
        vals = expense_data[cat]
        if sum(vals) > 0:
            color = exp_colors[idx % len(exp_colors)]
            exp_datasets_js.append(f'{{label:{json.dumps(cat)},data:{json.dumps(vals)},backgroundColor:"{color}"}}')

    # Takeaways HTML
    takeaway_cards = ""
    for tw in takeaways:
        urg = tw['urgency']
        if urg == 'urgent':
            border_color = 'var(--danger)'
            badge_bg = 'var(--danger-bg)'
            badge_color = 'var(--danger)'
            badge_text = 'NEEDS ACTION'
        elif urg == 'caution':
            border_color = 'var(--warning)'
            badge_bg = 'var(--warning-bg)'
            badge_color = 'var(--warning)'
            badge_text = 'MONITOR'
        elif urg == 'positive':
            border_color = 'var(--accent)'
            badge_bg = 'var(--accent-bg)'
            badge_color = 'var(--accent)'
            badge_text = 'HEALTHY'
        else:
            border_color = 'var(--text-tertiary)'
            badge_bg = 'rgba(100,116,139,0.15)'
            badge_color = 'var(--text-secondary)'
            badge_text = 'INFO'

        takeaway_cards += f'''<div class="takeaway-card" style="border-left-color:{border_color}">
<div class="tw-header"><span class="tw-icon">{tw['icon']}</span><span class="tw-title">{tw['title']}</span><span class="tw-badge" style="background:{badge_bg};color:{badge_color}">{badge_text}</span></div>
<div class="tw-headline" style="color:{border_color}">{tw['headline']}</div>
<div class="tw-body">{tw['body']}</div>
</div>'''

    # Anomaly HTML
    anomaly_html = ""
    if anomalies:
        items = ""
        for a in anomalies:
            sev_color = 'var(--danger)' if a['severity'] == 'high' else 'var(--warning)' if a['severity'] == 'medium' else 'var(--accent-secondary)'
            items += f'<div class="anomaly-item" style="border-left-color:{sev_color}"><div class="anomaly-week">{a["week"]}</div><div class="anomaly-type" style="color:{sev_color}">{a["type"]}</div><div class="anomaly-detail">{a["detail"]}</div></div>'
        anomaly_html = f'<div class="section anomaly-section"><h2>Anomaly Alerts</h2><div class="anomaly-grid">{items}</div></div>'

    # Insights HTML
    insights_html = ''.join(f'<div class="insight-item">{ins}</div>' for ins in insights)

    # Summary table
    summary_rows = ""
    for i, r in enumerate(reports):
        pos = cash_positions[i]
        inf = inflows[i]
        outf = outflows[i]
        nt = inf - outf
        fx = fx_rates[i]
        nc = 'positive' if nt >= 0 else 'negative'
        summary_rows += f'<tr><td>{r["date_str"]}</td><td>{fmt_naira(pos)}</td><td class="positive">{fmt_naira(inf)}</td><td class="negative">{fmt_naira(outf)}</td><td class="{nc}">{fmt_naira(nt)}</td><td>{fx:,.2f}</td></tr>'

    # Per-week inflow breakdown table
    inflow_detail_rows = ""
    for r in reports:
        items = r.get('inflow_items', {})
        if items:
            sorted_items = sorted(items.items(), key=lambda x: -x[1])
            detail = ', '.join([f"{k}: {fmt_naira(v)}" for k, v in sorted_items[:5]])
            inflow_detail_rows += f'<tr><td>{r["date_str"]}</td><td>{fmt_naira(r.get("total_inflow", 0))}</td><td style="font-size:0.85em;color:var(--text-secondary)">{detail}</td></tr>'

    # Per-week outflow breakdown table
    outflow_detail_rows = ""
    for r in reports:
        items = r.get('outflow_items', {})
        if items:
            sorted_items = sorted(items.items(), key=lambda x: -x[1])
            detail = ', '.join([f"{k}: {fmt_naira(v)}" for k, v in sorted_items[:5]])
            outflow_detail_rows += f'<tr><td>{r["date_str"]}</td><td>{fmt_naira(r.get("total_outflow", 0))}</td><td style="font-size:0.85em;color:var(--text-secondary)">{detail}</td></tr>'

    generated_at = datetime.now().strftime('%d %b %Y %H:%M')
    latest_report_date = reports[-1]['date_str']
    first_report_date  = reports[0]['date_str']

    theme_css = get_base_css()
    toggle_html = get_toggle_html()
    theme_js = get_theme_js()

    # ── Build data-quality warning banner ───────────────────────────────
    data_warnings_html = ""
    if data_warnings:
        error_items = [w for w in data_warnings if w['level'] == 'error']
        warn_items  = [w for w in data_warnings if w['level'] == 'warning']
        banner_bg   = "#7f1d1d" if error_items else "#78350f"   # dark red / dark amber
        banner_border = "#ef4444" if error_items else "#f59e0b"
        banner_icon = "🚨" if error_items else "⚠️"
        banner_label = "DATA QUALITY ERRORS" if error_items else "DATA QUALITY WARNINGS"
        rows_html = ""
        for w in data_warnings:
            icon = "🔴" if w['level'] == 'error' else "🟡"
            rows_html += (f"<div style='margin-top:6px;font-size:13px;color:#fef2f2'>"
                          f"{icon} <strong>[{w['report']}]</strong> {w['message']}</div>")
        data_warnings_html = (
            f"<div style='margin:0 0 20px 0;padding:16px 20px;background:{banner_bg};"
            f"border:1px solid {banner_border};border-radius:8px;'>"
            f"<div style='font-size:14px;font-weight:700;color:#fef2f2'>"
            f"{banner_icon} {banner_label} — Numbers below may be incorrect</div>"
            f"{rows_html}"
            f"<div style='margin-top:10px;font-size:12px;color:#fca5a5'>"
            f"Action: Check the source Excel file for the affected week(s) and click Regenerate.</div>"
            f"</div>"
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Seamfix Cash Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
{theme_css}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:var(--bg-body);color:var(--text-primary);padding:0;min-height:100vh;transition:background 0.3s,color 0.3s}}
.container{{max-width:1600px;margin:0 auto;padding:0 28px 28px}}
.header{{margin-bottom:32px;padding-bottom:20px;border-bottom:2px solid var(--border-accent)}}
.header h1{{font-size:2.4em;font-weight:700;color:var(--text-primary);margin-bottom:4px}}
.header .sub{{color:var(--text-secondary);font-size:0.95em}}
.header .meta{{color:var(--text-tertiary);font-size:0.8em;margin-top:6px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:16px;margin-bottom:36px}}
.kpi-card{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:22px;transition:all 0.3s ease}}
.kpi-card:hover{{border-color:rgba(0,212,170,0.4);transform:translateY(-3px);box-shadow:var(--shadow-hover)}}
.kpi-label{{font-size:0.78em;color:var(--text-secondary);text-transform:uppercase;letter-spacing:1.2px;margin-bottom:8px;font-weight:600}}
.kpi-value{{font-size:1.6em;font-weight:700;color:var(--accent);margin-bottom:6px}}
.kpi-value.negative{{color:var(--danger)}}
.kpi-change{{font-size:0.82em;display:flex;align-items:center;gap:4px}}
.kpi-change.positive{{color:var(--accent)}}
.kpi-change.negative{{color:var(--danger)}}
.kpi-change.neutral{{color:var(--text-tertiary)}}
.charts-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(480px,1fr));gap:20px;margin-bottom:36px}}
.chart-box{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:22px;min-height:380px}}
.chart-box h3{{font-size:1em;font-weight:600;margin-bottom:16px;color:var(--text-heading)}}
.section{{background:var(--bg-card);border:1px solid var(--border-accent);border-radius:12px;padding:24px;margin-bottom:24px}}
.section h2{{font-size:1.3em;margin-bottom:16px;color:var(--accent)}}
.anomaly-section{{border-color:var(--danger-bg)}}
.anomaly-section h2{{color:var(--danger)}}
.anomaly-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}}
.anomaly-item{{background:var(--danger-bg);border-left:3px solid var(--danger);padding:14px;border-radius:0 6px 6px 0}}
.anomaly-week{{font-weight:600;color:var(--warning);font-size:0.85em;margin-bottom:3px}}
.anomaly-type{{font-weight:600;font-size:0.95em}}
.anomaly-detail{{color:var(--text-secondary);font-size:0.85em;margin-top:4px}}
.insight-item{{padding:10px 0;border-bottom:1px solid var(--border-light);color:var(--text-heading);line-height:1.7;font-size:0.92em}}
.insight-item:last-child{{border-bottom:none}}
table{{width:100%;border-collapse:collapse;font-size:0.85em}}
thead{{background:var(--bg-table-header)}}
th{{padding:10px 12px;text-align:left;font-weight:600;color:var(--accent);border-bottom:2px solid var(--border-accent)}}
td{{padding:10px 12px;border-bottom:1px solid var(--border-light);color:var(--text-heading)}}
tbody tr:hover{{background:var(--bg-table-hover)}}
.positive{{color:#00D4AA}}
.negative{{color:#FF6B6B}}
th.sortable{{cursor:pointer;user-select:none}}
th.sortable:hover{{color:var(--warning)}}
th.sort-asc::after{{content:' ▲';font-size:0.7em}}
th.sort-desc::after{{content:' ▼';font-size:0.7em}}
.tabs{{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}}
.tab-btn{{padding:8px 18px;border:1px solid var(--border-accent);border-radius:20px;background:transparent;color:var(--text-secondary);cursor:pointer;font-size:0.85em;font-family:inherit;transition:all 0.2s}}
.tab-btn.active,.tab-btn:hover{{background:var(--accent-bg);color:var(--accent);border-color:rgba(0,212,170,0.4)}}
.tab-content{{display:none}}
.tab-content.active{{display:block}}
.takeaways-section{{margin-bottom:36px}}
.takeaways-section h2{{font-size:1.4em;margin-bottom:16px;color:var(--warning);display:flex;align-items:center;gap:10px}}
.takeaways-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(440px,1fr));gap:16px}}
.takeaway-card{{background:var(--bg-card);border:1px solid var(--border-accent);border-left:4px solid var(--text-tertiary);border-radius:10px;padding:20px;transition:all 0.3s ease}}
.takeaway-card:hover{{transform:translateY(-2px);box-shadow:var(--shadow-hover)}}
.tw-header{{display:flex;align-items:center;gap:8px;margin-bottom:10px}}
.tw-icon{{font-size:1.2em}}
.tw-title{{font-size:0.82em;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:var(--text-secondary)}}
.tw-badge{{font-size:0.7em;font-weight:700;padding:3px 10px;border-radius:12px;letter-spacing:0.5px}}
.tw-headline{{font-size:1.05em;font-weight:600;margin-bottom:8px;line-height:1.4}}
.tw-body{{font-size:0.88em;color:var(--text-secondary);line-height:1.65}}
.tw-body ul{{margin:6px 0 0 16px}}
.tw-body li{{margin-bottom:4px}}
.nav-bar{{display:none!important}}

@media(max-width:1024px){{.charts-grid{{grid-template-columns:1fr}}.kpi-grid{{grid-template-columns:repeat(2,1fr)}}.takeaways-grid{{grid-template-columns:1fr}}.anomaly-grid{{grid-template-columns:repeat(2,1fr)}}}}
@media(max-width:640px){{.kpi-grid{{grid-template-columns:1fr}}.header h1{{font-size:1.6em}}.anomaly-grid{{grid-template-columns:1fr}}.nav-bar{{flex-wrap:wrap}}}}
.top-nav{{background:var(--bg-nav);border-bottom:1px solid var(--border-main);padding:0 24px;display:flex;align-items:center;height:48px;overflow-x:auto;position:sticky;top:0;z-index:200}}
.top-nav-brand{{color:var(--text-primary);font-weight:700;font-size:15px;margin-right:24px;white-space:nowrap;text-decoration:none}}
.top-nav-link{{color:var(--text-secondary);text-decoration:none;padding:0 14px;height:48px;display:flex;align-items:center;font-size:13px;border-bottom:2px solid transparent;white-space:nowrap;transition:color .2s}}
.top-nav-link:hover{{color:var(--text-primary)}}
.top-nav-link.active{{color:var(--text-primary);border-bottom-color:var(--accent);font-weight:500}}
.pdf-btn{{display:none!important}}
@media print{{
.pdf-btn{{display:none!important}}
*{{color-adjust:exact!important;-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}}
body{{background:#fff!important;color:#1e293b!important;padding:12px!important;font-size:9pt!important}}
.container{{max-width:100%}}
.header h1{{font-size:1.6em!important;background:none!important;-webkit-text-fill-color:#0f172a!important;color:#0f172a!important}}
.header .sub{{color:#475569!important}}
.header .meta{{color:#64748b!important}}
.nav-bar{{display:none!important}}
.kpi-grid{{grid-template-columns:repeat(6,1fr)!important;gap:8px!important;margin-bottom:16px!important}}
.kpi-card{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:12px!important;border-radius:8px!important;box-shadow:none!important;page-break-inside:avoid}}
.kpi-label{{color:#475569!important;font-size:0.65em!important}}
.kpi-value{{color:#0f172a!important;font-size:1.2em!important}}
.kpi-value.negative{{color:#dc2626!important}}
.kpi-change{{font-size:0.7em!important}}
.kpi-change.positive{{color:#16a34a!important}}
.kpi-change.negative{{color:#dc2626!important}}
.kpi-change.neutral{{color:#64748b!important}}
.takeaways-section{{margin-bottom:16px!important;page-break-inside:avoid}}
.takeaways-section h2{{color:#0f172a!important;font-size:1.1em!important}}
.takeaways-grid{{grid-template-columns:repeat(2,1fr)!important;gap:10px!important}}
.takeaway-card{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:14px!important;border-radius:8px!important;page-break-inside:avoid}}
.tw-title{{color:#475569!important}}
.tw-body{{color:#475569!important}}
.tw-headline{{font-size:0.95em!important}}
.charts-grid{{grid-template-columns:repeat(2,1fr)!important;gap:12px!important;margin-bottom:16px!important}}
.chart-box{{background:#fff!important;border:1px solid #e2e8f0!important;padding:14px!important;min-height:auto!important;page-break-inside:avoid}}
.chart-box h3{{color:#1e293b!important;font-size:0.85em!important}}
.section{{background:#f8fafc!important;border:1px solid #e2e8f0!important;padding:16px!important;page-break-inside:avoid}}
.section h2{{color:#0f172a!important}}
.anomaly-section{{border-color:#fca5a5!important}}
.anomaly-section h2{{color:#dc2626!important}}
.anomaly-grid{{grid-template-columns:repeat(3,1fr)!important;gap:8px!important}}
.anomaly-item{{background:#fef2f2!important;padding:10px!important}}
.anomaly-week{{color:#92400e!important}}
.anomaly-type{{color:#dc2626!important}}
.anomaly-detail{{color:#475569!important}}
.insight-item{{color:#475569!important;border-bottom-color:#e2e8f0!important}}
table{{font-size:0.75em!important}}
thead{{background:#f1f5f9!important}}
th{{color:#0f172a!important;border-bottom-color:#cbd5e1!important}}
td{{color:#334155!important;border-bottom-color:#e2e8f0!important}}
.positive{{color:#16a34a!important}}
.negative{{color:#dc2626!important}}
.tabs{{display:none!important}}
.tab-content{{display:block!important}}
tbody tr:hover{{background:transparent!important}}
}}
.dashboard-footer{{margin-top:48px;padding:20px 28px;border-top:1px solid var(--border-main);color:var(--text-secondary);font-size:12px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
.dashboard-footer span{{color:var(--text-secondary)}}
/* ── HEADER: match Pipeline Intelligence style ── */
.header{{padding:24px 28px 16px!important;border-bottom:1px solid var(--border-main)!important;margin-bottom:24px!important;background:none!important}}
.header h1{{font-size:22px!important;font-weight:700!important;background:none!important;-webkit-background-clip:unset!important;-webkit-text-fill-color:var(--text-primary)!important;color:var(--text-primary)!important;margin-bottom:4px!important}}
.header .sub{{font-size:13px!important;color:var(--text-secondary)!important}}
.header .meta{{font-size:12px!important;color:var(--text-secondary)!important;margin-top:4px!important}}
.header p{{font-size:13px!important;color:var(--text-secondary)!important;margin-top:4px!important}}
</style>
</head>
<body>
{toggle_html}
<nav class="top-nav"><span class="top-nav-brand">⚡ Seamfix</span><a href="dashboard.html" class="top-nav-link active">Cash Overview</a><a href="expense_dashboard.html" class="top-nav-link ">Expense &amp; Vendor</a><a href="budget_dashboard.html" class="top-nav-link ">Budget vs Actual</a><a href="revenue_dashboard.html" class="top-nav-link ">Revenue &amp; Fundability</a><a href="pipeline_dashboard.html" class="top-nav-link ">Pipeline Intelligence</a></nav>
<!-- PDF button hidden per user request -->
<div class="container">
<div class="header">
<h1>Seamfix Cash Dashboard</h1>
<div class="sub">Executive Financial Monitoring &mdash; Weekly Cash Reports 2026</div>
<div class="meta">Data as of: <strong>{latest_report_date}</strong> &nbsp;&bull;&nbsp; Generated: {generated_at}</div>
</div>

<!-- old nav-bar replaced by top-nav -->

{data_warnings_html}
<div class="kpi-grid">
<div class="kpi-card">
<div class="kpi-label">Total Position (incl. Investments)</div>
<div class="kpi-value">{fmt_naira(latest_cash)}</div>
<div class="kpi-change {'positive' if cash_chg >= 0 else 'negative'}">{'&#9650;' if cash_chg >= 0 else '&#9660;'} {abs(cash_chg):.1f}% vs prior week</div>
</div>
<div class="kpi-card">
<div class="kpi-label">NGN Balance (incl. Investments)</div>
<div class="kpi-value">{fmt_naira(ngn_balance)}</div>
<div class="kpi-change {'positive' if ngn_bal_chg >= 0 else 'negative'}">{'&#9650;' if ngn_bal_chg >= 0 else '&#9660;'} {abs(ngn_bal_chg):.1f}% vs prior week</div>
</div>
<div class="kpi-card">
<div class="kpi-label">USD Balance (incl. Investments)</div>
<div class="kpi-value">{fmt_usd(usd_balance)}</div>
<div class="kpi-change {'positive' if usd_bal_chg >= 0 else 'negative'}">{'&#9650;' if usd_bal_chg >= 0 else '&#9660;'} {abs(usd_bal_chg):.1f}% vs prior week</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Weekly Inflow</div>
<div class="kpi-value">{fmt_naira(latest_in)}</div>
<div class="kpi-change {'positive' if in_chg >= 0 else 'negative'}">{'&#9650;' if in_chg >= 0 else '&#9660;'} {abs(in_chg):.1f}% vs prior week</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Weekly Outflow</div>
<div class="kpi-value">{fmt_naira(latest_out)}</div>
<div class="kpi-change {'positive' if out_chg <= 0 else 'negative'}">{'&#9650;' if out_chg > 0 else '&#9660;'} {abs(out_chg):.1f}% vs prior week</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Net Cash Flow</div>
<div class="kpi-value {'negative' if net_flow < 0 else ''}">{fmt_naira(net_flow)}</div>
<div class="kpi-change neutral">this week</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Operational Runway</div>
<div class="kpi-value">{runway:.0f} weeks</div>
<div class="kpi-change neutral">({runway/4.3:.0f} months at operational burn rate)</div>
</div>
<div class="kpi-card">
<div class="kpi-label">Revenue Concentration</div>
<div class="kpi-value {'negative' if concentration > 60 else ''}">{concentration:.0f}%</div>
<div class="kpi-change neutral">{conc_name} (latest week)</div>
</div>
<div class="kpi-card">
<div class="kpi-label">4-Week Cash Forecast</div>
<div class="kpi-value {'negative' if forecast_4w_net < 0 else ''}">{fmt_naira(forecast_4w_net)}</div>
<div class="kpi-change {'negative' if forecast_4w_net_chg_pct < -10 else 'positive' if forecast_4w_net_chg_pct > 5 else 'neutral'}">{'+' if forecast_4w_net_chg_pct >= 0 else ''}{forecast_4w_net_chg_pct:.1f}% — expected, based on avg weekly inflows &amp; burn</div>
<div style="margin-top:6px;font-size:0.78em;color:var(--text-secondary)">Floor if revenue stops: {fmt_naira(forecast_4w_floor)} ({forecast_4w_floor_chg_pct:.1f}%)</div>
</div>
</div>

<div class="takeaways-section">
<h2>\U0001F4CB Executive Takeaways</h2>
<div class="takeaways-grid">
{takeaway_cards}
</div>
</div>

<div class="charts-grid">
<div class="chart-box"><h3>Total Cash Position (NGN Equivalent)</h3><canvas id="cashChart"></canvas></div>
<div class="chart-box"><h3>Weekly Inflows vs Outflows</h3><canvas id="flowChart"></canvas></div>
<div class="chart-box"><h3>Revenue Source Mix</h3><canvas id="revChart"></canvas></div>
<div class="chart-box"><h3>Expense Category Breakdown</h3><canvas id="expChart"></canvas></div>
<div class="chart-box"><h3>NGN vs USD Cash Split</h3><canvas id="splitChart"></canvas></div>
<div class="chart-box"><h3>USD/NGN Exchange Rate</h3><canvas id="fxChart"></canvas></div>
</div>

{anomaly_html}

<div class="section">
<h2>Detailed Views</h2>
<div class="tabs">
<button class="tab-btn active" onclick="showTab('summary')">Weekly Summary</button>
<button class="tab-btn" onclick="showTab('inflows')">Inflow Details</button>
<button class="tab-btn" onclick="showTab('outflows')">Outflow Details</button>
</div>
<div id="tab-summary" class="tab-content active">
<table id="summaryTable">
<thead><tr><th class="sortable" onclick="sortTable('summaryTable',0,'text')">Week Of</th><th class="sortable" onclick="sortTable('summaryTable',1,'money')">Cash Position</th><th class="sortable" onclick="sortTable('summaryTable',2,'money')">Inflow</th><th class="sortable" onclick="sortTable('summaryTable',3,'money')">Outflow</th><th class="sortable" onclick="sortTable('summaryTable',4,'money')">Net Flow</th><th class="sortable" onclick="sortTable('summaryTable',5,'number')">FX Rate</th></tr></thead>
<tbody>{summary_rows}</tbody>
</table>
</div>
<div id="tab-inflows" class="tab-content">
<table id="inflowsTable">
<thead><tr><th class="sortable" onclick="sortTable('inflowsTable',0,'text')">Week Of</th><th class="sortable" onclick="sortTable('inflowsTable',1,'money')">Total Inflow</th><th class="sortable" onclick="sortTable('inflowsTable',2,'text')">Top Sources</th></tr></thead>
<tbody>{inflow_detail_rows}</tbody>
</table>
</div>
<div id="tab-outflows" class="tab-content">
<table id="outflowsTable">
<thead><tr><th class="sortable" onclick="sortTable('outflowsTable',0,'text')">Week Of</th><th class="sortable" onclick="sortTable('outflowsTable',1,'money')">Total Outflow</th><th class="sortable" onclick="sortTable('outflowsTable',2,'text')">Top Categories</th></tr></thead>
<tbody>{outflow_detail_rows}</tbody>
</table>
</div>
</div>
</div>

<script>
function sortTable(tableId, colIdx, type) {{
    const table = document.getElementById(tableId);
    const tbody = table.querySelector('tbody');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    const headers = table.querySelectorAll('th.sortable');

    // Determine sort direction
    const header = headers[colIdx];
    const isAsc = header.classList.contains('sort-asc');

    // Clear previous sort indicators
    headers.forEach(h => {{
        h.classList.remove('sort-asc', 'sort-desc');
    }});

    // Sort rows
    rows.sort((a, b) => {{
        let aVal = a.cells[colIdx].textContent.trim();
        let bVal = b.cells[colIdx].textContent.trim();

        if (type === 'money') {{
            aVal = parseFloat(aVal.replace(/[₦$B M K(),\-]/g, '')) || 0;
            bVal = parseFloat(bVal.replace(/[₦$B M K(),\-]/g, '')) || 0;
        }} else if (type === 'number') {{
            aVal = parseFloat(aVal) || 0;
            bVal = parseFloat(bVal) || 0;
        }} else if (type === 'status') {{
            const statusOrder = {{'On Track': 0, 'Closed': 1, 'At Risk': 2, 'Off Track': 3, 'Unknown': 4}};
            aVal = statusOrder[aVal] ?? 4;
            bVal = statusOrder[bVal] ?? 4;
        }} else {{
            aVal = aVal.toLowerCase();
            bVal = bVal.toLowerCase();
        }}

        return isAsc ? (aVal > bVal ? 1 : -1) : (aVal < bVal ? 1 : -1);
    }});

    // Re-append rows
    rows.forEach(row => tbody.appendChild(row));
    header.classList.add(isAsc ? 'sort-desc' : 'sort-asc');
}}

function showTab(id) {{
    document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
    document.getElementById('tab-'+id).classList.add('active');
    event.target.classList.add('active');
}}

const C = {{p:'#00D4AA',s:'#4ECDC4',a:'#FFE66D',d:'#FF6B6B',g:'rgba(0,212,170,0.08)'}};
const baseOpts = {{
    responsive:true,
    maintainAspectRatio:true,
    plugins:{{legend:{{labels:{{color:'#94a3b8',font:{{family:"'Inter',sans-serif",size:11}}}}}}}},
    scales:{{x:{{grid:{{color:C.g}},ticks:{{color:'#64748b',maxRotation:45}}}},y:{{grid:{{color:C.g}},ticks:{{color:'#64748b',callback:function(v){{if(v>=1e9)return'\\u20A6'+(v/1e9).toFixed(1)+'B';if(v>=1e6)return'\\u20A6'+(v/1e6).toFixed(0)+'M';if(v>=1e3)return'\\u20A6'+(v/1e3).toFixed(0)+'K';return'\\u20A6'+v}}}}}}}}
}};

const dates = {json.dumps(dates)};

new Chart(document.getElementById('cashChart'),{{
    type:'line',
    data:{{labels:dates,datasets:[{{label:'Total Cash (NGN eq.)',data:{json.dumps(cash_positions)},borderColor:C.p,backgroundColor:'rgba(0,212,170,0.08)',borderWidth:2.5,fill:true,tension:0.3,pointRadius:4,pointBackgroundColor:C.p}}]}},
    options:baseOpts
}});

new Chart(document.getElementById('flowChart'),{{
    type:'bar',
    data:{{labels:dates,datasets:[
        {{label:'Inflow',data:{json.dumps(inflows)},backgroundColor:'rgba(0,212,170,0.7)',borderRadius:4}},
        {{label:'Outflow',data:{json.dumps(outflows)},backgroundColor:'rgba(255,107,107,0.7)',borderRadius:4}},
        {{label:'Net Flow',data:{json.dumps(nets)},type:'line',borderColor:C.a,backgroundColor:'transparent',borderWidth:2,pointRadius:3,tension:0.3}}
    ]}},
    options:baseOpts
}});

new Chart(document.getElementById('revChart'),{{
    type:'bar',
    data:{{labels:dates,datasets:[{','.join(rev_datasets_js)}]}},
    options:{{...baseOpts,scales:{{...baseOpts.scales,x:{{...baseOpts.scales.x,stacked:true}},y:{{...baseOpts.scales.y,stacked:true}}}}}}
}});

new Chart(document.getElementById('expChart'),{{
    type:'bar',
    data:{{labels:dates,datasets:[{','.join(exp_datasets_js)}]}},
    options:{{...baseOpts,scales:{{...baseOpts.scales,x:{{...baseOpts.scales.x,stacked:true}},y:{{...baseOpts.scales.y,stacked:true}}}}}}
}});

new Chart(document.getElementById('splitChart'),{{
    type:'bar',
    data:{{labels:dates,datasets:[
        {{label:'NGN Balances',data:{json.dumps(ngn_positions)},backgroundColor:'rgba(0,212,170,0.7)',borderRadius:4}},
        {{label:'USD Balances (NGN eq.)',data:{json.dumps(usd_positions)},backgroundColor:'rgba(100,116,139,0.55)',borderRadius:4}}
    ]}},
    options:{{...baseOpts,scales:{{...baseOpts.scales,x:{{...baseOpts.scales.x,stacked:true}},y:{{...baseOpts.scales.y,stacked:true}}}}}}
}});

new Chart(document.getElementById('fxChart'),{{
    type:'line',
    data:{{labels:dates,datasets:[{{label:'USD/NGN',data:{json.dumps(fx_rates)},borderColor:C.s,backgroundColor:'rgba(78,205,196,0.08)',borderWidth:2.5,fill:true,tension:0.3,pointRadius:4}}]}},
    options:{{...baseOpts,scales:{{...baseOpts.scales,y:{{...baseOpts.scales.y,ticks:{{color:'#64748b',callback:function(v){{return'\\u20A6'+v.toLocaleString()}}}}}}}}}}
}});
{theme_js}
</script>
<div class="dashboard-footer">
<span>Seamfix Financial Intelligence &nbsp;·&nbsp; Powered by Claude Cowork</span>
<span>Available Data Range: {first_report_date} &ndash; {latest_report_date} &nbsp;&bull;&nbsp; Generated: {generated_at} &nbsp;&bull;&nbsp; $1 = ₦1,450</span>
</div>
</body>
</html>"""

    with open(output_path, 'w') as f:
        f.write(html)


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else "/sessions/relaxed-busy-archimedes/mnt/Seamfix Cash Reports 2026/"
    if not os.path.isdir(folder):
        print(f"Error: {folder} not found"); sys.exit(1)

    xlsx_files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and 'Cash Report' in f]
    if not xlsx_files:
        print("No Cash Report files found"); sys.exit(1)

    print(f"Found {len(xlsx_files)} reports")

    reports = []
    for fn in sorted(xlsx_files):
        fp = os.path.join(folder, fn)
        print(f"  Processing: {fn}...", end=' ')
        rec = extract_report(fp)
        if rec:
            reports.append(rec)
            print(f"OK ({rec['date_str']})")
        else:
            print("SKIPPED")

    reports.sort(key=lambda x: x['date'])

    # Investment outflows (e.g. SCB mutual fund) are asset transfers, not expenses.
    # We adjust total position: for any week with a net investment outflow,
    # add the outflow back (since the money is still an asset, just in a different form).
    # For weeks with net investment inflows (liquidations), no adjustment needed
    # because the cash already includes the liquidated amount.
    for r in reports:
        outflow_items = r.get('outflow_items', {})
        inflow_items = r.get('inflow_items', {})
        inv_out = get_investment_outflows(outflow_items)
        inv_in = get_investment_inflows(inflow_items)
        # Net investment outflow this week: money moved from bank to investment
        net_inv_out = max(0, inv_out - inv_in)
        r['investment_outflow'] = inv_out
        r['investment_inflow'] = inv_in
        # total_cash_ngn = liquid cash + NGN investments + USD investments (both closing balances
        # from the Cash Report sheet). net_inv_out is NOT added back here because the investment
        # portfolio closing balances already reflect any money transferred into them during the week.
        r['total_cash_ngn'] = r.get('liquid_cash_ngn', 0) + r.get('investment_ngn', 0) + r.get('investment_usd_ngn', 0)

    print(f"\nExtracted {len(reports)} weeks")

    # Print verification
    for r in reports:
        inf = r.get('total_inflow', 0)
        outf = r.get('total_outflow', 0)
        cash = r.get('total_cash_ngn', 0) or r.get('closing_balance', 0)
        print(f"  {r['date_str']}: Cash={fmt_naira(cash)} | In={fmt_naira(inf)} | Out={fmt_naira(outf)} | Net={fmt_naira(inf-outf)}")

    anomalies = detect_anomalies(reports)
    # Cap to last 2 weeks — historical flags are noise, not actionable
    if len(reports) > 2:
        recent_weeks = {r['date_str'] for r in reports[-2:]}
        anomalies = [a for a in anomalies if a['week'] in recent_weeks]
    insights = generate_insights(reports)
    takeaways = generate_takeaways(reports)

    # Data quality validation — catches parsing failures before numbers reach leadership
    data_warnings = validate_reports(reports)
    if data_warnings:
        print(f"\n{'='*60}")
        print(f"DATA QUALITY WARNINGS ({len(data_warnings)} issue(s)):")
        for w in data_warnings:
            prefix = "ERROR  " if w['level'] == 'error' else "WARNING"
            print(f"  [{prefix}] [{w['report']}] {w['message']}")
        print(f"{'='*60}")
    else:
        print(f"\nData validation: OK — all {len(reports)} reports look clean.")

    out_path = os.path.join(folder, 'dashboard.html')
    generate_html(reports, anomalies, insights, takeaways, out_path, data_warnings=data_warnings)
    print(f"\nDashboard: {out_path}")

    # Also copy to outputs if available
    outputs = "/sessions/relaxed-busy-archimedes/mnt/outputs/"
    if os.path.isdir(outputs):
        import shutil
        shutil.copy2(out_path, os.path.join(outputs, 'dashboard.html'))
        print(f"Also saved to: {outputs}dashboard.html")


if __name__ == '__main__':
    main()
